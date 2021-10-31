import math
import os
import sys
import random
import numpy as np
import copy
from enum import Enum
from single_engine_predict_noc_nop import *
from mesh_hetero import *
from matplotlib import pyplot as plt
from config import *
import openpyxl

degrade_ratio_list = []
excel_datas = []

def randomTest(GATest,iterTime, HW_param, memory_param, NoC_param, all_sim_node_num , if_multicast, filename):

	fitness_min_ran = 0
	fitness_list = []
	fitness_min_ran_list = []
	for i in range(iterTime):
		#---生成个代---
		for_list, act_wgt_dict, out_dict, parallel_dim_list, partition_list = GATest.GaGetChild()
		#---计算适应度---
		fitness, degrade_ratio, compuation_cycles, runtime_list,cp_list,utilization_ratio_list, energy_dram_list, energy_L2_list, energy_die2die, energy_MAC, worstlinks = \
			calFitness(for_list, act_wgt_dict, out_dict, parallel_dim_list, partition_list, GATest.network_param, HW_param, memory_param, NoC_param, if_multicast)
		#---比较适应度，并记录相关变量---
		if fitness_min_ran == 0 or fitness < fitness_min_ran:
			fitness_min_ran = fitness
			for_list_1 = copy.deepcopy(for_list)
			act_wgt_dict_1 = copy.deepcopy(act_wgt_dict)
			out_dict_1 = copy.deepcopy(out_dict)
			parallel_dim_list_1 = copy.deepcopy(parallel_dim_list)
			partition_list_1 = copy.deepcopy(partition_list)
			compuation_cycles_1 = compuation_cycles
			degrade_ratio_1 = degrade_ratio
		fitness_list.append(fitness)
		fitness_min_ran_list.append(fitness_min_ran)
		degrade_ratio_list.append (degrade_ratio)

		excel_datas.append([i, fitness, degrade_ratio, str(for_list[0]), \
			parallel_dim_list[0][0],parallel_dim_list[0][1],parallel_dim_list[0][2], \
			parallel_dim_list[1][0],parallel_dim_list[1][1],parallel_dim_list[1][2], \
			parallel_dim_list[0][0]*parallel_dim_list[1][0], \
			parallel_dim_list[0][1]*parallel_dim_list[1][1], \
			parallel_dim_list[0][2]*parallel_dim_list[1][2], \
			parallel_dim_list[0][0]*parallel_dim_list[1][0]*parallel_dim_list[0][1]*parallel_dim_list[1][1], \
			str(partition_list), runtime_list[0], runtime_list[1], runtime_list[2],  \
			runtime_list[3], runtime_list[4], runtime_list[5], runtime_list[6],\
			cp_list[0], cp_list[1], cp_list[2],cp_list[3], cp_list[4], cp_list[5] ,\
		    utilization_ratio_list[0], utilization_ratio_list[1], utilization_ratio_list[2],utilization_ratio_list[3], utilization_ratio_list[4], utilization_ratio_list[5], \
			energy_dram_list[0], energy_dram_list[1], energy_dram_list[2], energy_dram_list[3], \
			energy_L2_list[0], energy_L2_list[1], energy_L2_list[2], energy_L2_list[3], \
			sum(energy_dram_list), sum(energy_L2_list), energy_die2die,energy_MAC, sum(energy_dram_list)+sum(energy_L2_list)+energy_die2die+energy_MAC , str(worstlinks) ])
		print("######---------Times = ", i)
		print("fitness_min_ran = ",fitness_min_ran)
		print("compuation_cycles_1 = ",compuation_cycles_1)
		print("degrade_ratio_1 = ",degrade_ratio_1)
		print("######---------over")
		print("")
		
		#---生成task file
	createTaskFile(for_list_1, act_wgt_dict_1, out_dict_1, parallel_dim_list_1, partition_list_1,GATest.network_param, HW_param, memory_param, NoC_param, all_sim_node_num, if_multicast)
	workbook = openpyxl.Workbook()
	sheet = workbook.get_sheet_by_name('Sheet') 
	# 写入标题
	column_tite = ["index","fitness","degrade_ratio", "dataflow", \
		"PP2","PQ2","PK2","PP3","PQ3","PK3","PP","PQ","PKtotal","PPPQtotal", \
		"partition_list",\
		"runtimeP","runtimeQ", "runtimeC", "runtimeK", "runtimeChipNum", "runtimeCoreNum", "runtime_calNum",\
		"ol1_cp_id","al1_cp_id","wl1_cp_id","ol2_cp_id","al2_cp_id","wl2_cp_id", \
		"ol1_util","al1_util","wl1_util","ol2_util","al2_util","wl2_util", \
		"energy_wr_opt_dram", "energy_rd_opt_dram", "energy_rd_wgt_dram", "energy_rd_act_dram", \
		"energy_wr_opt_L2", "energy_rd_opt_L2", "energy_rd_wgt_L2", "energy_rd_act_L2", \
		"dram_energy", "L2_energy", "energy_die2die", "energy_MAC", "energy_sum" , "worstlinks"]
	for col,column in enumerate(column_tite):
		sheet.cell(1, col+1, column)
	# 写入每一行
	for row, data in enumerate(excel_datas):
		for col, column_data in enumerate(data):
			sheet.cell(row+2, col+1, column_data)

	workbook.save(filename)
	return compuation_cycles_1,degrade_ratio_1, fitness_min_ran_list

if __name__ == '__main__':

	network_param = {"P":224,"Q":224,"C":3,"K":64,"R":3,"S":3}
	HW_param = {"Chiplet":4,"PE":16,"intra_PE":{"C":8,"K":8}}
	memory_param = {"OL1":1.5,"OL2":1.5*16,"AL1":800/1024,"AL2":64,"WL1":18,"WL2":18*16}
	
	NoC_w = int(HW_param["PE"] ** 0.5) + 1
	NOC_NODE_NUM = NoC_w * (NoC_w-1)
	NoP_w = int(HW_param["Chiplet"] ** 0.5) + 1
	NOP_SIZE = NoP_w * (NoP_w-1)
	
	TOPO_param = {"NoC_w":NoC_w, "NOC_NODE_NUM": NOC_NODE_NUM, "NoP_w": NoP_w, "NOP_SIZE": NOP_SIZE,"nop_scale_ratio": nop_bandwidth/noc_bandwidth}
	
	filename = './randomTest_result_VGG-16 conv1-new '+str(HW_param["Chiplet"])+'_'+str(HW_param["PE"])+'.xls'

	# --- 生成noc-nop结构图
	NoC_param, all_sim_node_num = construct_noc_nop_topo(TOPO_param["NOC_NODE_NUM"],TOPO_param["NoC_w"], TOPO_param["NOP_SIZE"],TOPO_param["NoP_w"], TOPO_param["nop_scale_ratio"])
	debug=0
	if_multicast = 0
	GATest = GaEncode(network_param, HW_param, debug)

	iterTime = 100
	fitness_min_ran = 0
	index = range(iterTime)

	random_test_iter = 1
	f = open("./random_test_record.txt",'w')
	GATest.printBasicSetFile(f)
	f.close()

	for i in range(random_test_iter):
		print("###### test iteration = ",i)
		compuation_cycles_1,degrade_ratio_1, fitness_min_ran_list = randomTest(GATest, iterTime, HW_param, memory_param, NoC_param, all_sim_node_num, if_multicast, filename)
		print(fitness_min_ran_list[len(fitness_min_ran_list)-1])
		f = open("./random_test_record.txt",'a')
		print("###### test iteration = ",i, file = f)
		print("fitness:",fitness_min_ran_list[len(fitness_min_ran_list)-1], file=f)
		f.close()
