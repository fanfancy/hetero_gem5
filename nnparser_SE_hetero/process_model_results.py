import math
import os
import sys
import random
import numpy as np
import copy
from enum import Enum
from single_engine_predict_granularity import *
from mesh_hetero import *
from matplotlib import pyplot as plt
from config import *
import openpyxl


model_list = ["lenet", "alexnet", "VGG16" , "resnet_18", "resnet_50"]
layer_dict = {}
layer_times = {}
for model in model_list:
    model_file_name = './randomTest_result_'+model+'/'+model+'_all_granularity_128T'+'.xlsx'
    model_file_name_degrade = './randomTest_result_'+model+'/'+model+'_degrade_128T'+'.xlsx'    
    model_file_name_latency= './randomTest_result_'+model+'/'+model+'_latencty_128T'+'.xlsx'    
    model_file_name_energy = './randomTest_result_'+model+'/'+model+'_energy_128T'+'.xlsx'    
    
    print ("model = ", model)
    f = open("./nn_input_pqck/" + model + ".txt")
    lines = f.readlines()
    line_num = 0
    for line in lines:
        if line.startswith("#"):
            pass
        else:
            line_item = line.split(" ")
            real_layer_name = line_item[0]
            conv_layer_label = line_item[-1]
            conv_layer_label = conv_layer_label.replace("\n", "")
            
            if "conv_layer" in conv_layer_label:
                item = conv_layer_label.split("*")
                conv_layer_id = item[0]
                if len(item)>1:
                    conv_layer_time = int(item[1])
                else:
                    conv_layer_time = 1
                layer_dict[conv_layer_id] = real_layer_name
                print (conv_layer_id,conv_layer_time)
                layer_times[real_layer_name] = conv_layer_time
    print (layer_dict)
    print (layer_times)

    model_edp = []
    model_degrade = []
    model_latency = []
    model_energy = []
    column_tite  = []
    for item in layer_dict:
        
        degrade_list = []
        chip_num_list = []
        edp_list = []
        real_degrade_list = []
        latency_list = []
        energy_list = []
        e_mem_list = [];e_die2die_list=[];e_MAC_list =[]

        compuation_cycles_list = []
        layer_name = layer_dict[item]
        file_name = './randomTest_result_'+model+'/'+layer_name+'_granularity_128T'+'.xlsx'
        workbook=openpyxl.load_workbook(file_name)
        worksheet=workbook.worksheets[0]

        column_tite.append(layer_name) 
        for cell in list(worksheet.columns)[66]:
            if (cell.value == "num_chip"):
                print (cell.value)
            else:
                chip_num_list.append(cell.value)

        chip_degrade_rd_wgt = []
        chip_degrade_rd_act = []
        chip_degrade_wr_opt = []
        chip_degrade_rd_opt = []
        for cell in list(worksheet.columns)[61]:
            if (cell.value == "chip_degrade_rd_wgt"):
                print (cell.value)
            else:
                chip_degrade_rd_wgt.append(cell.value)
        for cell in list(worksheet.columns)[60]:
            if (cell.value == "chip_degrade_rd_act"):
                print (cell.value)
            else:
                chip_degrade_rd_act.append(cell.value)
        for cell in list(worksheet.columns)[59]:
            if (cell.value == "chip_degrade_wr_opt"):
                print (cell.value)
            else:
                chip_degrade_wr_opt.append(cell.value)
        for cell in list(worksheet.columns)[58]:
            if (cell.value == "chip_degrade_rd_opt"):
                print (cell.value)
            else:
                chip_degrade_rd_opt.append(cell.value)
        
        ############################ core ############################
        core_degrade_rd_wgt = []
        core_degrade_rd_act = []
        core_degrade_wr_opt = []
        core_degrade_rd_opt = []
        for cell in list(worksheet.columns)[57]:
            if (cell.value == "core_degrade_rd_wgt"):
                print (cell.value)
            else:
                core_degrade_rd_wgt.append(cell.value)
        for cell in list(worksheet.columns)[56]:
            if (cell.value == "core_degrade_rd_act"):
                print (cell.value)
            else:
                core_degrade_rd_act.append(cell.value)
        for cell in list(worksheet.columns)[55]:
            if (cell.value == "core_degrade_wr_opt"):
                print (cell.value)
            else:
                core_degrade_wr_opt.append(cell.value)
        for cell in list(worksheet.columns)[54]:
            if (cell.value == "core_degrade_rd_opt"):
                print (cell.value)
            else:
                core_degrade_rd_opt.append(cell.value)


        for cell in list(worksheet.columns)[1]:
            if (cell.value == "compuation_cycles"):
                print (cell.value)
            else:
                compuation_cycles_list.append(cell.value)
        
        for cell in list(worksheet.columns)[1]:
            if (cell.value == "compuation_cycles"):
                print (cell.value)
            else:
                compuation_cycles_list.append(cell.value)
        
        for cell in list(worksheet.columns)[48]:
            if (cell.value == "e_mem"):
                print (cell.value)
            else:
                e_mem_list.append(cell.value)
        
        for cell in list(worksheet.columns)[47]:
            if (cell.value == "e_die2die"):
                print (cell.value)
            else:
                e_die2die_list.append(cell.value)
        
        for cell in list(worksheet.columns)[46]:
            if (cell.value == "e_MAC"):
                print (cell.value)
            else:
                e_MAC_list.append(cell.value)

        for id in range (len(chip_num_list)):
            degrade = max(core_degrade_rd_wgt[id], core_degrade_rd_act[id], core_degrade_rd_opt[id]+core_degrade_wr_opt[id], \
                chip_degrade_rd_act[id] + chip_degrade_rd_opt[id] + chip_degrade_rd_wgt[id] + chip_degrade_wr_opt[id] )
            real_degrade = degrade /  chip_num_list[id]
            if real_degrade<1:
                real_degrade = 1
            delay = real_degrade * compuation_cycles_list[id] 

            e_sum = e_mem_list[id] + e_MAC_list[id]
            real_edp = delay * e_sum / (PE_freq * freq_1G)
            real_degrade_list.append(real_degrade)
            edp_list.append(real_edp)
            energy_list.append(e_sum)
            latency_list.append(delay)
        
        model_edp.append(edp_list)
        model_degrade.append(real_degrade_list)
        model_latency.append(latency_list)
        model_energy.append(energy_list)

    sum_list = np.zeros((len(model_edp[0])))
    for method_id in range (len(model_edp[0])):
        for layer_id in range (len(model_edp)):
            layer_name = column_tite[layer_id]
            times = layer_times[layer_name]
            sum_list[method_id] += times*model_edp[layer_id][method_id]
    model_edp.append(sum_list)

    sum_list = np.zeros((len(model_energy[0])))
    for method_id in range (len(model_energy[0])):
        for layer_id in range (len(model_energy)):
            layer_name = column_tite[layer_id]
            times = layer_times[layer_name]
            sum_list[method_id] += times*model_energy[layer_id][method_id]
    avg_value = sum(sum_list)/(len(model_energy[0]))
    sum_list_scaled = [i/avg_value for i in sum_list]
    model_energy.append(sum_list_scaled)

    sum_list = np.zeros((len(model_latency[0])))
    for method_id in range (len(model_latency[0])):
        for layer_id in range (len(model_latency)):
            layer_name = column_tite[layer_id]
            times = layer_times[layer_name]
            sum_list[method_id] += times*model_latency[layer_id][method_id]
    avg_value = sum(sum_list)/(len(model_latency[0]))
    sum_list_scaled = [i/avg_value for i in sum_list]
    model_latency.append(sum_list_scaled)


    model_edp= zip(*model_edp)
    column_tite.append("edp_sum")
    workbook2 = openpyxl.Workbook()
    sheet2 = workbook2.get_sheet_by_name('Sheet') 
    for col,column in enumerate(column_tite):
        sheet2.cell(1, col+1, column)

    for row, data in enumerate(model_edp):
        for col, column_data in enumerate(data):
            sheet2.cell(row+2, col+1, column_data)
    workbook2.save(model_file_name)

    #### model_degrade ####
    model_degrade= zip(*model_degrade)
    workbook3 = openpyxl.Workbook()
    sheet3 = workbook3.get_sheet_by_name('Sheet') 
    for col,column in enumerate(column_tite):
        sheet3.cell(1, col+1, column)

    for row, data in enumerate(model_degrade):
        for col, column_data in enumerate(data):
            sheet3.cell(row+2, col+1, column_data)
    workbook3.save(model_file_name_degrade)

    #### model_latency ####
    model_latency= zip(*model_latency)
    workbook4 = openpyxl.Workbook()
    sheet4 = workbook4.get_sheet_by_name('Sheet') 
    for col,column in enumerate(column_tite):
        sheet4.cell(1, col+1, column)

    for row, data in enumerate(model_latency):
        for col, column_data in enumerate(data):
            sheet4.cell(row+2, col+1, column_data)
    workbook4.save(model_file_name_latency)


    #### model_energy ######
    model_energy= zip(*model_energy)
    workbook5 = openpyxl.Workbook()
    sheet5 = workbook5.get_sheet_by_name('Sheet') 
    for col,column in enumerate(column_tite):
        sheet5.cell(1, col+1, column)

    for row, data in enumerate(model_energy):
        for col, column_data in enumerate(data):
            sheet5.cell(row+2, col+1, column_data)
    workbook5.save(model_file_name_energy)

