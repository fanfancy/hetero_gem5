import math
import os
import sys
import random
import numpy as np
import copy
from enum import Enum
from matplotlib import pyplot as plt
import openpyxl
import xlrd
from inter_layer_noc_nop import *

def getLayerParam(app_name):
	layer_num = 0
	layer_dict = {}
	layer_id_list = []
	f = open("./nn_input_noc_nop/" + app_name + ".txt")

	print("network model ----- " + app_name + " -------------")

	lines = f.readlines()
	for line in lines:
		if line.startswith("#") or line.startswith("*"):
			pass
		elif line != "\n":
			line_item = line.split(" ")
			layer_num += 1
			H = int(line_item[1])
			M = int(line_item[2])
			C = int(line_item[3])
			R = int(line_item[4])
			S = int(line_item[4])
			stride = int(line_item[5])
			padding = int(line_item[6])
			K = int(line_item[7])
			P = int(line_item[8])
			Q = int(line_item[9])

			layer_id_list.append(layer_num)
			layer_dict[layer_num] = {"H":H,"M":M,"P":P,"Q":Q,"C":C,"K":K,"R":R,"S":S, "stride":stride, "padding":padding}
			print("layer" + str(layer_num) + " : " + str(layer_dict[layer_num]))
	f.close()
	return layer_dict, layer_id_list

def getIntraLayerEDP(app_name, file_type = "intra_layer"):
	EDP_intra_layer = {}

	file_name = "./test/intra_layer_edp/" + app_name + "_" + file_type +".txt"
	f = open(file_name)
	lines = f.readlines()
	layer_num = 0

	for line in lines:
		if line.startswith("layer"):
			line_item = line.split("\t")
			layer_num += 1
			EDP_intra_layer[layer_num] = {}
			EDP_intra_layer[layer_num]["P"] = float(line_item[1])
			EDP_intra_layer[layer_num]["PK"] = float(line_item[2])
			EDP_intra_layer[layer_num]["K"] = float(line_item[3])
	print(file_type," : ", EDP_intra_layer)
	return EDP_intra_layer

def nop_noc_test(app_name):
	PENoCNum = 16*16*16
	NOC_NODE_NUM = 20
	NoC_w = 5

	EDP_sum_dict = {"P":0, "K":0, "PK":0}
	EDP_inter_dict = {"P":0, "K":0, "PK":0}
	energy_sum_dict = {"P":0, "K":0, "PK":0}
	delay_sum_dict = {"P":0, "K":0, "PK":0}
	comm_delay_sum_dict = {"P":0, "K":0, "PK":0}
	d2d_e_sum_dict = {"P":0, "K":0, "PK":0}
	dram_e_sum_dict = {"P":0, "K":0, "PK":0}
	parallel_type_dict = {"P":[],"K":[],"PK":[]}

	para_type = ["P","K","PK"]
	parallel_type_1 = {"P":{"P":16,"Q":1,"K":1},'K':{"P":1,"Q":1,"K":16},'PK':{"P":4,"Q":1,"K":4}}
	parallel_type_2 = {"P":{"P":16,"Q":1,"K":1},'K':{"P":1,"Q":1,"K":16},'PK':{"P":4,"Q":1,"K":4}}

	file_name = "./test/noc_nop_edp/" + app_name + "_noc_nop.txt"
	f = open(file_name,'w')

	# --- 获得神经网络参数
	layer_dict, layer_id_list = getLayerParam(app_name)
	# --- 获得片内EDP数值
	EDP_intra_layer = getIntraLayerEDP(app_name)
	energy_intra_layer = getIntraLayerEDP(app_name, "intra_layer_energy")
	delay_intra_layer = getIntraLayerEDP(app_name, "intra_layer_delay")

	excel_datas = []
	data_list = {"P":[],"PK":[],"K":[]}

	# ---初始（最下面一层的）NoC edp
	for para_type_2 in para_type:
		layer_id = layer_id_list[-1]
		EDP = EDP_intra_layer[layer_id][para_type_2]
		energy = energy_intra_layer[layer_id][para_type_2]
		delay = delay_intra_layer[layer_id][para_type_2]
		EDP_sum_dict[para_type_2] += EDP
		energy_sum_dict[para_type_2] += energy
		delay_sum_dict[para_type_2] += delay
		line = "parallel " + str(para_type_2) + "  noc_EDP " + str(EDP)
		print(line, file = f)
		data_list[para_type_2] = [EDP]
	print("energy_sum_dict",energy_sum_dict)
	print("delay_sum_dict",delay_sum_dict)

	for ii in range(12):
			data_list["P"].append(0)
			data_list["PK"].append(0)
			data_list["K"].append(0)
	
	excel_per_layer = ["layer"+str(layer_id_list[-1])] + data_list["P"] +data_list["PK"] + data_list["K"] + [str(energy_sum_dict), str(delay_sum_dict), str(EDP_inter_dict), str(EDP_sum_dict), str(parallel_type_dict)]

	excel_datas.append(excel_per_layer)
	

	for i in reversed(layer_id_list):
		if i < layer_id_list[-1]:
			network_param_2 = layer_dict[i+1]
			parallel_type_dict_new = {"P":[], "K":[], "PK":[]}
			EDP_sum_dict_new = {"P":0, "K":0, "PK":0}
			EDP_inter_dict_new = {"P":0, "K":0, "PK":0}
			EDP_NoC_dict = {"P":0, "K":0, "PK":0}
			energy_NoC_dict = {"P":0, "K":0, "PK":0}
			delay_NoC_dict = {"P":0, "K":0, "PK":0}
			energy_sum_dict_new = {"P":0, "K":0, "PK":0}
			delay_sum_dict_new = {"P":0, "K":0, "PK":0}

			line = "layer " + str(i) + "--------------------------------------"
			print(line, file = f)
			excel_per_layer = []
			excel_per_layer.append(i)

			for para_type_1 in para_type:
				EDP = EDP_intra_layer[i][para_type_1]
				EDP_NoC_dict[para_type_1] = EDP
				energy_NoC_dict[para_type_1] = energy_intra_layer[i][para_type_1]
				delay_NoC_dict[para_type_1] = delay_intra_layer[i][para_type_1]
				#EDP_sum_dict[para_type_1] += EDP
				line = "parallel " + str(para_type_1) + "  noc_EDP " + str(EDP)
				print(line, file = f)
				data_list[para_type_1] = [EDP]

			# 计算层间的EDP
			for para_type_1 in para_type:
				EDP_num = {}
				energy_num = {}
				delay_num = {}
				EDP_num_only = {}
				for para_type_2 in para_type:
					parallel_1 = parallel_type_1[para_type_1]
					parallel_2 = parallel_type_2[para_type_2]
					worstCommFlitNum, d2d_e , dram_e, EDP = getInterLayer(network_param_2, parallel_1, parallel_2, NOC_NODE_NUM, NoC_w)
					EDP_num[para_type_2] = EDP + EDP_sum_dict[para_type_2]
					EDP_num_only[para_type_2] = EDP
					energy_num[para_type_2] = d2d_e + dram_e + energy_sum_dict[para_type_2]
					delay_num[para_type_2] = worstCommFlitNum + delay_sum_dict[para_type_2]

					data_list[para_type_1].append(d2d_e)
					data_list[para_type_1].append(dram_e)
					data_list[para_type_1].append(worstCommFlitNum)
					data_list[para_type_1].append(EDP)
				
				line = "layer " + str(i-1) + " : parallel " + str(para_type_1)
				print(line, file = f)

				EDP_num_min = EDP_num["P"]
				para_min = "P"
				for pa in EDP_num:
					if EDP_num[pa] < EDP_num_min:
						EDP_num_min = EDP_num[pa]
						para_min = pa
				
				EDP_sum_dict_new[para_type_1] += EDP_num_min + EDP_NoC_dict[para_type_1]
				EDP_inter_dict_new[para_type_1] += EDP_num_only[pa]
				energy_sum_dict_new[para_type_1] += energy_num[pa] + energy_NoC_dict[para_type_1]
				delay_sum_dict_new[para_type_1] += delay_num[pa] + delay_NoC_dict[para_type_1]

				parallel_type_dict_new[para_type_1] = [para_min]
				parallel_type_dict_new[para_type_1] += parallel_type_dict[para_min]
				
				line2 = "---------nop_EDP = " +str(EDP_num_only)
				print(line2, file = f)
				line2 = "---------EDP_num = " +str(EDP_num)
				print(line2, file = f)

			parallel_type_dict = copy.deepcopy(parallel_type_dict_new)
			EDP_sum_dict = copy.deepcopy(EDP_sum_dict_new)
			EDP_inter_dict = copy.deepcopy(EDP_inter_dict_new)
			energy_sum_dict = copy.deepcopy(energy_sum_dict_new)
			delay_sum_dict = copy.deepcopy(delay_sum_dict_new)

			excel_per_layer = ["layer"+str(i)] + data_list["P"] +data_list["PK"] + data_list["K"] + [str(energy_sum_dict),str(delay_sum_dict), str(EDP_inter_dict), str(EDP_sum_dict), str(parallel_type_dict)]

			excel_datas.append(excel_per_layer)
			
		line3 = "----EDP_sum_dict = " + str(EDP_sum_dict)
		line1 = "----parallel_type_dict = " +str(parallel_type_dict)
		line4 = "----energy_sum_dict = " + str(energy_sum_dict)
		line5 = "----delay_sum_dict = " + str(delay_sum_dict)
		print(line3, file = f)
		print(line1, file = f)
		print(line4, file = f)
		print(line5, file = f)
	
	workbook = openpyxl.Workbook()
	sheet = workbook.get_sheet_by_name('Sheet') 
	# 写入标题
	column_tite = ["layer_id", "P_noc_edp", "P_P_d2d_e","P_P_dram_e","P_P_comm_d", "P_P_nop_edp", "P_K_d2d_e","P_K_dram_e","P_K_comm_d", "P_K_nop_edp", "P_PK_d2d_e","P_PK_dram_e","P_PK_comm_d", "P_PK_nop_edp"]
	column_tite += ["PK_noc_edp", "PK_P_d2d_e","PK_P_dram_e","PK_P_comm_d", "PK_P_nop_edp", "PK_K_d2d_e","PK_K_dram_e","PK_K_comm_d", "PK_K_nop_edp", "PK_PK_d2d_e","PK_PK_dram_e","PK_PK_comm_d", "PK_PK_nop_edp"]
	column_tite += ["K_noc_edp", "K_P_d2d_e","K_P_dram_e","K_P_comm_d", "K_P_nop_edp", "K_K_d2d_e","K_K_dram_e","K_K_comm_d", "K_K_nop_edp", "K_PK_d2d_e","K_PK_dram_e","K_PK_comm_d", "K_PK_nop_edp"]
	column_tite += ["energy_sum", "delay_sum", "EDP_inter_layer", "edp_sum", "par_index"]
	for col,column in enumerate(column_tite):
		sheet.cell(1, col+1, column)
	# 写入每一行
	for row, data in enumerate(excel_datas):
		for col, column_data in enumerate(data):
			sheet.cell(row+2, col+1, column_data)

	setname = str(network_param_2["P"]) + "_" + str(network_param_2["Q"]) + "_" + str(network_param_2["K"]) + "_" + str(network_param_2["C"]) + "_" + str(NoC_w) + "_" + str(NOC_NODE_NUM)

	workbook.save("./test/noc_nop_edp/"+app_name+"_noc_nop.xls")

def noc_test(app_name):
	PENoCNum = 16*16*16
	NOC_NODE_NUM = 20
	NoC_w = 5

	EDP_sum_dict = {"P":0, "K":0, "PK":0}

	para_type = ["P","K","PK"]
	parallel_type = {"P":{"P":16,"Q":1,"K":1},'K':{"P":1,"Q":1,"K":16},'PK':{"P":4,"Q":1,"K":4}}
	
	file_name = "./test/noc_nop_edp/" + app_name + "_noc_edp.txt"
	f = open(file_name,'w')

	# --- 获得神经网络参数
	layer_dict, layer_id_list = getLayerParam(app_name)
	# --- 获得片内EDP数值
	EDP_intra_layer = getIntraLayerEDP(app_name)

	para_list = []
	edp_sum = 0
	layer_num = 0

	for i in layer_id_list:
		line = "layer " + str(i) + "-----------------"
		print(line, file = f)
		EDP_min = 0
		for para_type_2 in para_type:
			EDP = EDP_intra_layer[i][para_type_2]
			EDP_sum_dict[para_type_2] += EDP
			line = "parallel " + str(para_type_2) + "  EDP " + str(EDP)
			print(line, file = f)
			if EDP_min == 0 or EDP_min > EDP:
				EDP_min = EDP
				EDP_min_id = para_type_2
	
		line = "layer " + str(i) + " :  EDP_min_id = " + str(EDP_min_id) + "  EDP_Min =  " + str(EDP_min)
		print(line, file = f)
		para_list.append(EDP_min_id)
		edp_sum += EDP_min

		if i > layer_id_list[0]:
			network_param_2 = layer_dict[i]
			print(network_param_2)
			worstCommFlitNum, d2d_e , dram_e, EDP = getInterLayer(network_param_2, parallel_type[para_list[layer_num - 1]], parallel_type[para_list[layer_num]], NOC_NODE_NUM, NoC_w)
			edp_sum += EDP

			line = "layer " + str(i-1) + " to " + str(i) + " : nop_EDP =  " + str(EDP)
			print(line, file = f)
		layer_num += 1
		
	line3 = "----EDP_sum = " + str(edp_sum)
	line1 = "----parallel_type_dict = " +str(para_list)
	print(line3, file = f)
	print(line1, file = f)

def get_set_comm_num_test(app_name):

	para_type = ["P","K","PK"]
	dim_seq_1 = ["K","P","Q"]
	dim_seq_2 = ["K","P","Q"]
	parallel_type_1 = {"P":{"P":16,"Q":1,"K":1},'K':{"P":1,"Q":1,"K":16},'PK':{"P":4,"Q":1,"K":4}}
	parallel_type_2 = {"P":{"P":16,"Q":1,"K":1},'K':{"P":1,"Q":1,"K":16},'PK':{"P":4,"Q":1,"K":4}}

	file_name = "./test/nop_set_comm/" + app_name + "_set_comm.txt"
	f = open(file_name,'w')
	line = "app_name " + app_name + "--------------------------------------"
	print(line, file = f)

	# --- 获得神经网络参数
	layer_dict, layer_id_list = getLayerParam(app_name)

	comm_set_layer = {}
	comm_chip_set_layer = {}
	
	for i in reversed(layer_id_list):
		if i < layer_id_list[-1]:
			network_param_2 = layer_dict[i+1]

			line = "layer " + str(i) + "-------------------------"
			print(line, file = f)

			comm_set_layer[i] = {}
			comm_chip_set_layer[i] = {}

			for para_type_1 in para_type:
				for para_type_2 in para_type:
					parallel_1 = parallel_type_1[para_type_1]
					parallel_2 = parallel_type_2[para_type_2]
					comm_num_dict, comm_type_dict, comm_type_times_dict, chiplet_num = getInterLayerComm(dim_seq_1, dim_seq_2, parallel_1, network_param_2, parallel_2, 0)
					comm_chip_set_dict, comm_set_dict = getSetComm(comm_num_dict)

					line0 = "parallel type : " + str(para_type_1) + " to " + str(para_type_2)
					line1 = "---comm_chip_set_dict: " + str(comm_chip_set_dict)
					line2 = "---comm_set_dict: " + str(comm_set_dict)
					print(line0, file = f)
					print(line1, file = f)
					print(line2, file = f)

					par_type_str = str(para_type_1) + "_" + str(para_type_2)
					comm_chip_set_layer[i][par_type_str] = comm_chip_set_dict
					comm_set_layer[i][par_type_str] = comm_set_dict
	
	print("comm_chip_set_layer ", comm_chip_set_layer, file = f)
	print("comm_set_layer ", comm_set_layer, file = f)
	f.close()
	return comm_chip_set_layer, comm_set_layer # [layer_num][parallel_type][pac_id][packet_num, recv_id_list] , parallel_type = parallel_layer_i + "_" + parallel_layer_(i+1)

if __name__ == '__main__':
	app_name = str(sys.argv[1])
	#nop_noc_test(app_name)
	#noc_test(app_name)
	comm_chip_set_layer, comm_set_layer = get_set_comm_num_test(app_name)