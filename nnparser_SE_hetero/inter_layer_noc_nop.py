import math
import os
import sys
import random
import numpy as np
import copy
from enum import Enum
from matplotlib import pyplot as plt
import openpyxl
from config import *

def getChipletID(p,q,k,dim_mapping_seq, debug_flag = 0):
	if debug_flag == 0:
		chiplet_id = p * dim_mapping_seq["P"] + q * dim_mapping_seq["Q"] + k * dim_mapping_seq["K"]
	else:
		chiplet_id = str([p,q,k])
	return chiplet_id

def extendChipletID(pre_id, k, k_offset, debug_flag = 0):
	if debug_flag == 0:
		chiplet_id = pre_id + k * k_offset
	else:
		chiplet_id = pre_id[:-2] + str(k) + "]"
	return chiplet_id

#获得包含的维度量
def getIndex(addr, offset, index_dict):
	dict_out = {}
	addr_start = addr
	addr_end = addr+offset-1
	for i in range(len(index_dict)-1):
		start = index_dict[i]
		end = index_dict[i+1] - 1
		if addr_start < start and addr_end >= start and addr_end <= end:
			dict_out[i] = addr_end - start + 1
		elif addr_start < start and addr_end > end:
			dict_out[i] = end - start + 1
		elif addr_start >= start and addr_end <= end:
			dict_out[i] = addr_end - addr_start + 1
		elif addr_start >= start and addr_start <=end and addr_end > end:
			dict_out[i] = end - addr_start + 1
	return dict_out

#获得当前chiplet所需之前chiplet的通信量
def getCommDict(addr, size_dict, dict, dim_mapping_seq, debug_flag):
	p = addr[0]
	q = addr[1]
	p_size = size_dict["P"]
	q_size = size_dict["Q"]
	k_size = size_dict["K"]

	p_dict = getIndex(p, p_size,dict["P"])
	q_dict = getIndex(q, q_size,dict["Q"])
	k_dict = getIndex(0, k_size,dict["K"])

	comm_dict = {}
	for q in q_dict:
		for p in p_dict:
			for k in k_dict:
				chiplet_id = getChipletID(p,q,k,dim_mapping_seq,debug_flag)
				comm_num = q_dict[q] * p_dict[p] * k_dict[k]
				comm_dict[chiplet_id] = comm_num
	return comm_dict

#计算层间通信
# network_param_1 = layer i : {"P":P, "Q":Q, "K":K, "C":C, "R":R, "S":S}
# network_param_2 = layer i+1 : {"P":P, "Q":Q, "K":K, "C":C, "R":R, "S":S, "stride":stride, "padding":padding}
# parallel_1 = layer i : {"P":PP,"Q":PQ,"K":PK}
# parallel_2 = layer i+1 : {"P":PP,"Q":PQ,"K":PK}
# chiplet_size = [chiplet_h, chiplet_w]
# dim_seq_1 = ["P","Q","K"]，相当于考量chiplet_id是先按哪个维度排列的
# debug_flag = 0 :正常运行，1 :方便看结果
def getInterLayerComm(dim_seq_1, dim_seq_2, parallel_1, network_param_2, parallel_2, debug_flag = 0):
	#---上一网络层的output act每个chiplet所拥有的大小

	network_param_1 = {"P":0, "Q":0, "K":0}
	network_param_1["P"] = network_param_2["H"]
	network_param_1["Q"] = network_param_2["M"]
	network_param_1["K"] = network_param_2["C"]

	P_i = math.ceil(network_param_1["P"]/parallel_1["P"])
	Q_i = math.ceil(network_param_1["Q"]/parallel_1["Q"])
	K_i = math.ceil(network_param_1["K"]/parallel_1["K"])
	single_chip_act_size_1 = {"P":P_i, "Q":Q_i, "K":K_i}
	
	#---获得有关chiplet id编号的信息
	dim_mapping_seq_1 = {"P":1,"Q":1,"K":1}
	par = 1
	for dim in dim_seq_1:
		dim_mapping_seq_1[dim] = par
		par *= parallel_1[dim]
	dim_mapping_seq_2 = {"P":1,"Q":1,"K":1}
	par = 1
	for dim in dim_seq_2:
		dim_mapping_seq_2[dim] = par
		par *= parallel_2[dim]

	#---当前层的output act与input act每个chiplet所拥有的大小
	stride = network_param_2["stride"]
	padding = network_param_2["padding"]
	P_output = math.ceil(network_param_2["P"]/parallel_2["P"])
	Q_output = math.ceil(network_param_2["Q"]/parallel_2["Q"])
	P_input = network_param_2["R"] + (P_output-1) * stride
	Q_input = network_param_2["S"] + (Q_output-1) * stride
	K_input = network_param_2["C"]
	single_chip_act_size_2 = {"P":P_input, "Q":Q_input, "K":K_input}

	#---记录每一块chiplet所需要的input act的起始块地址offset
	chiplet_act_index_dict = {}
	chiplet_act_index_dict_single = {}
	for q in range(parallel_2["Q"]):
		for p in range(parallel_2["P"]):
			p_index = P_output * p
			q_index = Q_output * q
			if p_index >= network_param_2["P"]:
				break
			chiplet_index = getChipletID(p,q,0,dim_mapping_seq_2,debug_flag)
			offset_p = P_output * p *stride - padding
			offset_q = Q_output * q *stride - padding
			offset = [offset_p, offset_q]
			chiplet_act_index_dict[chiplet_index] = offset
	
	chiplet_num = len(chiplet_act_index_dict) * parallel_2["K"]

	#---记录前一层chiplet的output act分配情况
	chiplet_act_index_dict_pre = {"P":{},"Q":{},"K":{}}
	for p in range(parallel_1["P"]):
		offset_p = P_i * p
		if offset_p < network_param_1["P"]:
			chiplet_act_index_dict_pre["P"][p] = offset_p
		else:
			#chiplet_act_index_dict_pre["P"][p] = network_param_1["P"]
			break
	if p in chiplet_act_index_dict_pre["P"]:
		chiplet_act_index_dict_pre["P"][p+1] = network_param_1["P"]
	else:
		chiplet_act_index_dict_pre["P"][p] = network_param_1["P"]

	for q in range(parallel_1["Q"]):
		offset_q = Q_i * q
		chiplet_act_index_dict_pre["Q"][q] = offset_q
		if offset_q < network_param_1["Q"]:
			chiplet_act_index_dict_pre["Q"][q] = offset_q
		else:
			break
	if q in chiplet_act_index_dict_pre["Q"]:
		chiplet_act_index_dict_pre["Q"][q+1] = network_param_1["Q"]
	else:
		chiplet_act_index_dict_pre["Q"][q] = network_param_1["Q"]

	for k in range(parallel_1["K"]):
		offset_k = K_i * k
		chiplet_act_index_dict_pre["K"][k] = offset_k
		if offset_k < network_param_1["K"]:
			chiplet_act_index_dict_pre["K"][k] = offset_k
		else:
			break
	if k in chiplet_act_index_dict_pre["K"]:
		chiplet_act_index_dict_pre["K"][k+1] = network_param_1["K"]
	else:
		chiplet_act_index_dict_pre["K"][k] = network_param_1["K"]

	#---记录当前chiplet与上一层chiplet的通信情况
	#---comm_inter_layer_dict_o_i[layer(i+1)_chiplet_id] = {[layer(i)_chiplet_id]:comm_num,...}
	#---comm_inter_layer_dict_i_o[layer(i)_chiplet_id] = {[layer(i+1)_chiplet_id]:comm_num,...}
	comm_inter_layer_dict_o_i = {}
	comm_inter_layer_dict_i_o = {}
	for chip_id in chiplet_act_index_dict:
		index = chiplet_act_index_dict[chip_id]
		comm_dict = getCommDict(index, single_chip_act_size_2, chiplet_act_index_dict_pre, dim_mapping_seq_1, debug_flag)
		comm_inter_layer_dict_o_i[chip_id] = comm_dict
		for chip_id_i in comm_dict:
			if chip_id_i not in comm_inter_layer_dict_i_o:
				comm_inter_layer_dict_i_o[chip_id_i] = {}
			comm_inter_layer_dict_i_o[chip_id_i][chip_id] = comm_dict[chip_id_i]
			
	comm_num_dict = {}
	comm_set_dict = {} # 组间通信情况
	comm_type_dict = {"uni-cast":0, "multi-cast":0, "broadcast":0}
	comm_type_times_dict = {"uni-cast":0, "multi-cast":0, "broadcast":0}
	for chip_id_i in comm_inter_layer_dict_i_o:
		comm_num_dict[chip_id_i] = {}
		num = 0
		for chip_id_o in comm_inter_layer_dict_i_o[chip_id_i]:
			chip_id_list = []
			for k in range(parallel_2["K"]):
				chip_id_o_ex = extendChipletID(chip_id_o, k, dim_mapping_seq_2["K"], debug_flag)
				if debug_flag == 1:
					chip_id_list.append(chip_id_o_ex)
				else:
					if chip_id_o_ex != chip_id_i:
						chip_id_list.append(chip_id_o_ex)

			commNum = comm_inter_layer_dict_i_o[chip_id_i][chip_id_o]
			if len(chip_id_list) == 1:
				comm_type_dict["uni-cast"] += commNum
				comm_type_times_dict["uni-cast"] += 1
			elif len(chip_id_list) + 1 < (parallel_2["P"]*parallel_2["Q"]*parallel_2["K"]) and len(chip_id_list) > 1:
				comm_type_dict["multi-cast"] += commNum
				comm_type_times_dict["multi-cast"] += 1
			elif len(chip_id_list) + 1 == (parallel_2["P"]*parallel_2["Q"]*parallel_2["K"]):
				comm_type_dict["broadcast"] += commNum
				comm_type_times_dict["broadcast"] += 1
			comm_num_dict[chip_id_i][num] = [commNum, chip_id_list]
			num += 1
	return comm_num_dict, comm_type_dict,comm_type_times_dict, chiplet_num

def convertDictIndex(dict):
	out_dict = {}
	pac_id = 0
	for id_i in dict:
		out_dict[id_i] = {}
		pac_id = 0
		for list_str_in in dict[id_i]:
			list_str = list_str_in.replace("[","start, ")
			list_str = list_str.replace("]",", end")
			id_list = list_str.split(", ")
			list = []
			for i in range(len(id_list)):
				if i > 0 and i < len(id_list)-1:
					id_o = int(id_list[i])
					list.append(id_o)
			pac_id_str = "pacID" + str(pac_id)
			out_dict[id_i][pac_id_str] = [list , dict[id_i][list_str_in]]
			pac_id += 1
	return out_dict

def getSetComm(comm_all , set_e_num = 4):
	comm_chip_set_dict = {}
	comm_set_dict = {}
	for chip_id_i in comm_all:
		set_id_i = int(chip_id_i / set_e_num)
		if set_id_i not in comm_set_dict:
			comm_set_dict[set_id_i] = {}
		for packet_id in comm_all[chip_id_i]:
			chip_id_list = comm_all[chip_id_i][packet_id][1];
			packet_num = comm_all[chip_id_i][packet_id][0];
			chip_id_list_set = []
			set_id_list = []
			for chip_id_o in chip_id_list:
				set_id_o = int(chip_id_o / set_e_num)
				if set_id_o == set_id_i:
					pass
				else:
					chip_id_list_set.append(chip_id_o)
					if set_id_o not in set_id_list:
						set_id_list.append(set_id_o)
			if len(chip_id_list_set) > 0:
				if chip_id_i not in comm_chip_set_dict:
					comm_chip_set_dict[chip_id_i] = {}
				chip_id_list_set_str = str(chip_id_list_set)
				if chip_id_list_set_str not in comm_chip_set_dict[chip_id_i]:
					comm_chip_set_dict[chip_id_i][chip_id_list_set_str] = packet_num
				else:
					comm_chip_set_dict[chip_id_i][chip_id_list_set_str] += packet_num
			if len(set_id_list) > 0:
				set_id_list = sorted(set_id_list)
				set_id_list_str = str(set_id_list)
				if set_id_list_str not in comm_set_dict[set_id_i]:
					comm_set_dict[set_id_i][set_id_list_str] = packet_num
				else:
					comm_set_dict[set_id_i][set_id_list_str] += packet_num
	
	# convert comm_chip_set
	comm_chip_set_dict_out = convertDictIndex(comm_chip_set_dict)
	comm_set_dict_out = convertDictIndex(comm_set_dict)

	return comm_chip_set_dict_out, comm_set_dict_out

def setRouteTable_Mesh(NOC_NODE_NUM, NoC_w):
	# mesh 4*4
	F = {}
	bw_scales = {}
	for src in range (NOC_NODE_NUM):
		for dst in range (NOC_NODE_NUM):
			local = src + 1000
			F[(local,src)] = 0
			F[(src,local)] = 0
			src_x = src %  NoC_w
			src_y = int(src / NoC_w)
			dst_x = dst %  NoC_w
			dst_y = int(dst / NoC_w)
			if (src_x == dst_x) :
				if (src_y - dst_y == 1) or (src_y- dst_y == -1) :
					F[(src,dst)] = 0
					bw_scales[(src,dst)] = 1
			elif (src_y == dst_y) :
				if (src_x - dst_x == 1) or (src_x - dst_x == -1):
					F[(src,dst)] = 0
					bw_scales[(src,dst)] = 1

	# print ("F",F)

	noc_route_table = {}
	hops = {}
	noc_route_ids = {}

	for src in range (0,NOC_NODE_NUM):
		for dst in range (0,NOC_NODE_NUM):
			cur_dst = src
			cur_src = src
			noc_route_table[(src,dst)] = []
			noc_route_ids[(src,dst)] = []
			while cur_dst != dst:
				src_x = cur_src %  NoC_w
				src_y = int(cur_src / NoC_w)
				dst_x = dst %  NoC_w
				dst_y = int(dst / NoC_w)
				# print ("src_x",src_x,"src_y",src_y,"dst_x",dst_x,"dst_y",dst_y)
				if (src_x > dst_x): # go west
					cur_dst = src_x-1 +  src_y * NoC_w
				elif (src_x < dst_x): # go east
					cur_dst = src_x+1 +  src_y * NoC_w
				elif (src_y < dst_y): # go north
					cur_dst = src_x + (src_y+1) * NoC_w
				elif (src_y > dst_y): # go south
					cur_dst = src_x + (src_y-1) * NoC_w
				#print ("cur_dst",cur_dst)
				noc_route_table[(src,dst)].append((cur_src,cur_dst))
				cur_src = cur_dst
				noc_route_ids[(src,dst)].append(cur_dst)
			#print ("\nnoc_route_table:",src,dst,noc_route_table[(src,dst)])
	#print ("noc_route_table",noc_route_table)

	route_table = {}
	for src in range (1000,1000+NOC_NODE_NUM):
		for dst in range (1000,1000+NOC_NODE_NUM):
			route_table[(src,dst)] = []
			noc_src = src - 1000
			noc_dst = dst -1000
			route_table[(src,dst)] = noc_route_table[(noc_src,noc_dst)].copy()
			if (src!=dst):
				route_table[(src,dst)].append((noc_dst,dst))
				route_table[(src,dst)].insert(0,(src,noc_src))
				F[(noc_dst,dst)] = 0
				F[(src,noc_src)] = 0
				bw_scales[(src,noc_src)] = 1
				bw_scales[(noc_dst,dst)] = 1




	for item in route_table:
		hops[item] = len(route_table[item])
		#print (item,route_table[item])

	#print ("hops==========",sum(hops.values())/NOC_NODE_NUM/NOC_NODE_NUM)
	#print("route_table=========")
	#print(route_table)
	return route_table, F, bw_scales

def setRouteTable_Ring(NOC_NODE_NUM):
	# mesh 4*4
	F = {}
	bw_scales = {}
	for src in range (NOC_NODE_NUM):
		local = src + 1000
		F[(local,src)] = 0
		F[(src,local)] = 0
		bw_scales[(src,local)] = 1
		bw_scales[(local,src)] = 1

		beside_node = (src + 1) % NOC_NODE_NUM
		F[(src,beside_node)] = 0
		bw_scales[(src,beside_node)] = 2

	# print ("F",F)

	noc_route_table = {}
	hops = {}
	noc_route_ids = {}

	for src in range (0,NOC_NODE_NUM):
		for dst in range (0,NOC_NODE_NUM):
			cur_dst = src
			cur_src = src
			noc_route_table[(src,dst)] = []
			noc_route_ids[(src,dst)] = []
			while cur_dst != dst:
				cur_dst += 1
				cur_dst = cur_dst % NOC_NODE_NUM
				noc_route_table[(src,dst)].append((cur_src,cur_dst))
				cur_src = cur_dst
				noc_route_ids[(src,dst)].append(cur_dst)

	route_table = {}
	for src in range (1000,1000+NOC_NODE_NUM):
		for dst in range (1000,1000+NOC_NODE_NUM):
			route_table[(src,dst)] = []
			noc_src = src - 1000
			noc_dst = dst -1000
			route_table[(src,dst)] = noc_route_table[(noc_src,noc_dst)].copy()
			if (src!=dst):
				route_table[(src,dst)].append((noc_dst,dst))
				route_table[(src,dst)].insert(0,(src,noc_src))
				F[(noc_dst,dst)] = 0
				F[(src,noc_src)] = 0
				bw_scales[(src,noc_src)] = 1
				bw_scales[(noc_dst,dst)] = 1




	for item in route_table:
		hops[item] = len(route_table[item])
		#print (item,route_table[item])

	#print ("hops==========",sum(hops.values())/NOC_NODE_NUM/NOC_NODE_NUM)
	#print("route_table=========")
	#print(route_table)
	return route_table, F, bw_scales

def calCommCycle(F,routeTable, bw_scales, commDict):
	F_cur=F.copy()

	link_comm_num_sum = 0

	for send_id in commDict:
		for packet in commDict[send_id]:
			commNum = commDict[send_id][packet][0]
			commFlitNum = math.ceil(commNum / neu_per_flit_act_nop)
			dst_list = commDict[send_id][packet][1]
			link_list = []
			for dst_id in dst_list:
				for link in routeTable[(send_id + 1000, dst_id + 1000)]:
					if link not in link_list:
						link_list.append(link)
						F_cur[link] += commFlitNum / bw_scales[link]
						link_comm_num_sum += commNum

	worstCommFlitNum = max(F_cur.values()) 
	worstlinks = []
	for item in F_cur:
		if F_cur[item] == max(F_cur.values()): 
			worstlinks.append(item)

	print(F_cur)

	return F_cur, worstCommFlitNum, worstlinks, link_comm_num_sum

def getCalCycle(chiplet_num, network_param_2, PENoCNum):
	calNum = network_param_2["P"] * network_param_2["Q"] * network_param_2["K"] * network_param_2["C"] * network_param_2["R"] * network_param_2["S"]
	CalCycle = math.ceil(calNum / PENoCNum / chiplet_num)
	return CalCycle

def calCommEnergy(link_comm_num_sum, comm_type_dict):
	d2d_comm_num = link_comm_num_sum # act num
	d2d_energy = d2d_comm_num * act_wgt_width * DIE2DIE_energy_ratio

	dram_access_num = 0
	for i in comm_type_dict:
		dram_access_num += comm_type_dict[i]
	#print(comm_type_dict)
	#print(dram_access_num)
	dram_access_energy = dram_access_num * act_wgt_width * DRAM_energy_ratio
	return d2d_energy, dram_access_energy

def getInterLayer(network_param_2, parallel_1, parallel_2, NOC_NODE_NUM, NoC_w, topology="mesh"):
	dim_seq_1 = ["K","P","Q"]
	dim_seq_2 = ["K","P","Q"]
	# 通信量计算
	comm_num_dict, comm_type_dict, comm_type_times_dict, chiplet_num = getInterLayerComm(dim_seq_1, dim_seq_2, parallel_1, network_param_2, parallel_2, 0)
	
	# 拓扑构建routing table
	if topology == "mesh":
		routeTable, F, bw_scales = setRouteTable_Mesh(NOC_NODE_NUM, NoC_w)
	elif topology == "ring":
		routeTable, F, bw_scales = setRouteTable_Ring(NOC_NODE_NUM)

	# 计算链路通信量-delay
	F_final, worstCommFlitNum, worstlinks, link_comm_num_sum = calCommCycle(F,routeTable, bw_scales, comm_num_dict)

	# 计算energy
	d2d_e , dram_e = calCommEnergy(link_comm_num_sum, comm_type_dict)

	edp = 2 * (d2d_e + dram_e) * worstCommFlitNum / freq_1G / PE_freq

	return worstCommFlitNum, d2d_e , dram_e, edp

def calInterComm_simba(par_C_pre, par_K_pre, par_C_cur, par_K_cur, network_param_2):
	P = network_param_2["P"]
	Q = network_param_2["Q"]
	H = network_param_2["H"]
	M = network_param_2["M"]
	C = network_param_2["C"]
	chiplet_pre = par_C_pre * par_K_pre
	chiplet_cur = par_C_cur * par_K_cur
	chiplet_ratio = (chiplet_cur / chiplet_pre)

	if par_C_cur >= par_K_pre:
		if chiplet_ratio <= 1:
			worstCommFlitNum = 0
			d2d_energy = 0
			dram_energy = 0
			EDP = 0
		elif chiplet_ratio > 1:
			comm_num = H * M * C
			comm_num_flit = math.ceil(comm_num / neu_per_flit_act_nop)
			worstCommFlitNum = comm_num_flit / par_C_cur
			d2d_energy = comm_num / par_C_cur * (chiplet_cur - chiplet_pre) * act_wgt_width * DIE2DIE_energy_ratio
			dram_energy = chiplet_cur * (comm_num / par_C_cur) * act_wgt_width * DRAM_energy_ratio
			EDP = worstCommFlitNum * (d2d_energy + dram_energy) / freq_1G / PE_freq
	elif par_C_cur < par_K_pre:
		ratio = math.ceil(par_K_pre/par_C_cur)
		comm_num = H * M * C
		comm_num_flit = math.ceil(comm_num / neu_per_flit_act_nop)
		num_pre_per_chip = comm_num / par_K_pre
		worstCommFlitNum = math.ceil(num_pre_per_chip / neu_per_flit_act_nop) * (ratio - 1)
		d2d_energy = num_pre_per_chip * (ratio - 1) * chiplet_pre * act_wgt_width * DIE2DIE_energy_ratio
		dram_energy =  comm_num * act_wgt_width * DRAM_energy_ratio * par_K_cur
		if par_K_cur > par_C_pre * ratio:
			worstCommFlitNum += comm_num_flit / par_C_cur
			d2d_energy += comm_num / par_C_cur * (chiplet_cur - chiplet_pre) * act_wgt_width * DIE2DIE_energy_ratio
		EDP = worstCommFlitNum * (d2d_energy + dram_energy) / freq_1G / PE_freq
	return worstCommFlitNum, d2d_energy , dram_energy, EDP


if __name__ == '__main__':

	excel_datas = []

	NOC_NODE_NUM = 20
	NoC_w = 5

	# 1表示前一层，2表示当前层
	#network_param_1 = {"P":224,"Q":224,"C":3,"K":64,"R":3,"S":3}
	network_param_2 = {"H":56,"M":56, "P":56,"Q":56,"C":64,"K":64,"R":1,"S":1, "stride":1, "padding":0}
	#network_param_2 = {"H":7,"M":7, "P":7,"Q":7,"C":512,"K":2048,"R":1,"S":1, "stride":1, "padding":0}
	

	#parallel_type_1 = {"P":{"P":16,"Q":1,"K":1},'Q':{"P":1,"Q":16,"K":1},'K':{"P":1,"Q":1,"K":16},"PQ":{"P":4,"Q":4,"K":1},'PK':{"P":4,"Q":1,"K":4},'QK':{"P":1,"Q":4,"K":4}}
	#parallel_type_2 = {"P":{"P":16,"Q":1,"K":1},'Q':{"P":1,"Q":16,"K":1},'K':{"P":1,"Q":1,"K":16},"PQ":{"P":4,"Q":4,"K":1},'PK':{"P":4,"Q":1,"K":4},'QK':{"P":1,"Q":4,"K":4}}
	#parallel_type_1 = {'K':{"P":1,"Q":1,"K":16},"P":{"P":16,"Q":1,"K":1},'PK':{"P":4,"Q":1,"K":4}}
	#parallel_type_2 = {'K':{"P":1,"Q":1,"K":16},"P":{"P":16,"Q":1,"K":1},'PK':{"P":4,"Q":1,"K":4}}

	parallel_type_2 = {"P":{"P":16,"Q":1,"K":1}}
	parallel_type_1 = {'P':{"P":16,"Q":1,"K":1}}

	times = 0
	for type2 in parallel_type_2:
		for type1 in parallel_type_1:
			parallel_1 = parallel_type_1[type1]
			parallel_2 = parallel_type_2[type2]
			dim_seq_1 = ["K","P","Q"]
			dim_seq_2 = ["K","P","Q"]
			print("----------- times = ", times)
			print("parallel_1")
			print(parallel_1)
			print("parallel_2")
			print(parallel_2)

			# 通信量计算
			comm_num_dict, comm_type_dict, comm_type_times_dict, chiplet_num = getInterLayerComm(dim_seq_1, dim_seq_2, parallel_1, network_param_2, parallel_2, 0)
			print("comm_num_dict", comm_num_dict)
			print("comm_type_dict", comm_type_dict)
			print("comm_type_times_dict", comm_type_times_dict)
			# 拓扑构建routing table
			routeTable, F, bw_scales = setRouteTable_Mesh(NOC_NODE_NUM, NoC_w)

			# 计算链路通信量
			F_cur, worstCommFlitNum, worstlinks, link_comm_num_sum = calCommCycle(F,routeTable,bw_scales, comm_num_dict)
			
			print("F_cur", F_cur)
			print("worstCommFlitNum", worstCommFlitNum)
			print("worstlinks", worstlinks)
			print("link_comm_num_sum", link_comm_num_sum)
			# 计算片上计算时间
			PENoCNum = 16*8*16
			calCycle_noc = getCalCycle(chiplet_num, network_param_2, PENoCNum)
			d2d_energy, dram_access_energy = calCommEnergy(link_comm_num_sum, comm_type_dict)

			edp = 2 * (d2d_energy + dram_access_energy) * worstCommFlitNum / freq_1G / PE_freq
			
			print("d2d_energy", d2d_energy)
			print("dram_access_energy", dram_access_energy)
			print("edp ",edp)
			#excel_datas.append([chiplet_num, calCycle_noc, type1+"-"+type2, type1, type2, comm_type_times_dict["uni-cast"], comm_type_times_dict["multi-cast"], comm_type_times_dict["broadcast"], comm_type_dict["uni-cast"], comm_type_dict["multi-cast"], comm_type_dict["broadcast"] , worstCommFlitNum, str(worstlinks), str(F_final)])
			#print(comm_num_dict)
			times += 1
	
	workbook = openpyxl.Workbook()
	sheet = workbook.get_sheet_by_name('Sheet') 
	# 写入标题
	column_tite = ["chiplet_num", "calCycle_noc", "parallel-type","parallel-type-i","parallel-type-i+1", "times-uni-cast","times-multi-cast","times-broadcast","num-uni-cast","num-multi-cast","num-broadcast", "worstCommFlitNum", "worstlinks", "F"]
	for col,column in enumerate(column_tite):
		sheet.cell(1, col+1, column)
	# 写入每一行
	for row, data in enumerate(excel_datas):
		for col, column_data in enumerate(data):
			sheet.cell(row+2, col+1, column_data)

	#setname = str(network_param_2["P"]) + "_" + str(network_param_2["Q"]) + "_" + str(network_param_2["K"]) + "_" + str(network_param_2["C"]) + "_" + str(NoC_w) + "_" + str(NOC_NODE_NUM)

	#workbook.save("./test/1116/inter_layer_output_"+setname+"_topo.xls")