import math
import os
import sys
import random
import numpy as np
import copy
from enum import Enum
from single_engine_predict_granularity import *
from mesh_hetero import *
from matplotlib import pyplot as plt
from config import *
import openpyxl


def randomTest(GATest,iterTime, HW_param, memory_param , if_multicast, filename):

	degrade_ratio_list = []
	excel_datas = []
	edp_res_min = 0

	fitness_min_ran = 0
	fitness_list = []
	fitness_min_ran_list = []
	for i in range(iterTime):
		#---生成个代---
		for_list, act_wgt_dict, out_dict, parallel_dim_list, partition_list, code = GATest.GaGetChild()
		#---计算适应度---
		compuation_cycles, runtime_list,cp_list,utilization_ratio_list, energy_dram_list, energy_L2_list, energy_L1_list, energy_MAC, energy_die2die_in_layer , L2_traffic,compuation_redundance,degrade_list = \
			calFitness_granu(for_list, act_wgt_dict, out_dict, parallel_dim_list, partition_list, GATest.network_param, HW_param, memory_param, if_multicast)
		# energy_die2die_in_layer 是粗略计算一下层内运算过程中的die间通信
		
		#---比较适应度，并记录相关变量---
		
		#degrade_list =  [core_degrade_rd_opt, core_degrade_wr_opt, core_degrade_rd_act, core_degrade_rd_wgt, chip_degrade_rd_opt, chip_degrade_wr_opt, chip_degrade_rd_act, chip_degrade_rd_wgt]
	
		degrade = max(degrade_list[0], degrade_list[1],  degrade_list[2], degrade_list[3], \
			degrade_list[4]+degrade_list[6]+degrade_list[7] , degrade_list[5])

		if degrade<1:
			degrade = 1
		e_mem = sum(energy_dram_list)+sum(energy_L2_list)+sum(energy_L1_list)
		energy_sum = e_mem + energy_MAC +(energy_die2die_in_layer)
		delay = degrade * compuation_cycles   # cycles
		edp_res = delay * energy_sum  /(PE_freq * freq_1G) # pJ*s
		
		if edp_res_min == 0 or edp_res < edp_res_min:
			edp_res_min = edp_res
			min_iter_id = i

		excel_datas.append([i, compuation_cycles, str(for_list[0]), \
			parallel_dim_list[0][0],parallel_dim_list[0][1],parallel_dim_list[0][2], \
			parallel_dim_list[1][0],parallel_dim_list[1][1],parallel_dim_list[1][2], \
			parallel_dim_list[0][0]*parallel_dim_list[1][0], \
			parallel_dim_list[0][1]*parallel_dim_list[1][1], \
			parallel_dim_list[0][2]*parallel_dim_list[1][2], \
			parallel_dim_list[0][0]*parallel_dim_list[1][0]*parallel_dim_list[0][1]*parallel_dim_list[1][1], \
			str(partition_list), runtime_list[0], runtime_list[1], runtime_list[2],  \
			runtime_list[3], runtime_list[4], runtime_list[5], runtime_list[6],\
			cp_list[0], cp_list[1], cp_list[2],cp_list[3], cp_list[4], cp_list[5] ,\
		    utilization_ratio_list[0], utilization_ratio_list[1], utilization_ratio_list[2],utilization_ratio_list[3], utilization_ratio_list[4], utilization_ratio_list[5], \
			energy_dram_list[0], energy_dram_list[1], energy_dram_list[2], energy_dram_list[3], \
			energy_L2_list[0], energy_L2_list[1], energy_L2_list[2], energy_L2_list[3], \
			energy_L1_list[0], energy_L1_list[1], \
			sum(energy_dram_list), sum(energy_L2_list), sum(energy_L1_list), energy_MAC, energy_die2die_in_layer, e_mem, \
		    str(code),L2_traffic[0],L2_traffic[1],L2_traffic[2], compuation_redundance, \
			degrade_list[0],degrade_list[1],degrade_list[2], degrade_list[3], \
			degrade_list[4],degrade_list[5],degrade_list[6], degrade_list[7], \
			degrade, 1/degrade, delay,energy_sum, edp_res_min ])
		# print("######---------Times = ", i)
		

	# 写入每一行
	return excel_datas[min_iter_id], edp_res_min

def run_randomTest_granu_model(model_name,the_layer_name):
	workbook = openpyxl.Workbook()
	sheet = workbook.get_sheet_by_name('Sheet') 
	# 写入标题
	column_tite = ["index", "compuation_cycles", "dataflow", \
		"PP2","PQ2","PK2","PP3","PQ3","PK3","PP","PQ","PKtotal","PPPQtotal", \
		"partition_list",\
		"runtimeP","runtimeQ", "runtimeC", "runtimeK", "runtimeChipNum", "runtimeCoreNum", "runtime_calNum",\
		"ol1_cp_id","al1_cp_id","wl1_cp_id","ol2_cp_id","al2_cp_id","wl2_cp_id", \
		"ol1_util","al1_util","wl1_util","ol2_util","al2_util","wl2_util", \
		"e_wr_opt_dram", "e_rd_opt_dram", "e_rd_wgt_dram", "e_rd_act_dram", \
		"e_wr_opt_L2", "e_rd_opt_L2", "e_rd_wgt_L2", "e_rd_act_L2", \
		"e_rd_wgt_L1", "e_rd_act_L1", \
		"e_dram", "e_L2", "e_L1", "e_MAC", "e_die2die", "e_mem", "code",\
		"rd_opt_ratio","rd_act_ratio","rd_wgt_ratio","compuation_redundance", \
		"core_degrade_rd_opt", "core_degrade_wr_opt", "core_degrade_rd_act", "core_degrade_rd_wgt", \
		"chip_degrade_rd_opt", "chip_degrade_wr_opt", "chip_degrade_rd_act", "chip_degrade_rd_wgt", \
		"degrade", "mac_util.", "delay","energy_sum", "edp_res", \
		
		"num_chip","num_core","PE_C","PE_K",\
		"WL1","AL1","OL1",\
		"WL2","AL2","OL2", "l1_base","WAO_L1_type","WAO_L2", "area_chip_mm2"]
	for col,column in enumerate(column_tite):
		sheet.cell(1, col+1, column)

	DNN_model = model_name
	

	f = open("./nn_input_pqck/" + DNN_model + ".txt")
	lines = f.readlines()
	line_num = 0
	for line in lines:
		if line.startswith("#"):
			pass
		else:
			line_item = line.split(" ")
			layer_name = line_item[0]
			if layer_name == the_layer_name and line_item[1] == "CONV":
				network_param = {}
				network_param['P'] = int(line_item[9])  #按照output计算
				network_param['Q'] = int(line_item[10])	
				network_param['R'] = network_param['S'] = int(line_item[5])
				network_param['C'] = int(line_item[4])	
				network_param['stride'] = int(line_item[6])
				network_param['K'] = int(line_item[8])
	
				print (layer_name,network_param)
				
				filename = './randomTest_result_'+DNN_model+'/'+layer_name+'_granularity_128T'+'.xlsx'
				
				PE_C_list = [16]
				PE_K_list = [16]
				core_num_list = [4,8,16,32,64,128,256] # 1-256 2-128 4-64 8-32 16-16 
				chiplet_num_list = [1,2,4,8,16,32,64]
				l1_list = [0.5,1,2,4,8,16]
				# W A O ratio 探索
				WAO_L1_list = { 1:[8,2,1], 2:[2,8,1] , 3:[5,5,1] , 4:[9,1.5,0.5], 5:[1.5,9,0.5]  }
				WAO_L2_list = [1,2,4]

				def valid(PE_C,PE_K,core_num,chiplet_num,WL1,AL1,OL1,WL2,AL2,OL2):
					# 16TOPs
					if PE_C*PE_K*core_num*chiplet_num == MAC_NUM and \
						WL1 > 0 and \
						AL1 > 0 and \
						OL1 > 0 :
						return True
					else:
						return False
				
				num2list ={}
				num2list[1] = [1,1]; 	num2list[2] = [2,1]; 		num2list[4] = [2,2]
				num2list[8] = [4,2];  	num2list[16] = [4,4]; 		num2list[32] = [4,8]
				num2list[64] = [8,8];	num2list[128] = [16,8];    	num2list[256] = [16,16]
				all_valid_param = 0
				all_results = []
				for l1 in l1_list:
					for WAO_L1_type in WAO_L1_list:
						for WAO_L2 in WAO_L2_list:
							for PE_C in PE_C_list:
								for PE_K in PE_K_list:
									for core_num in core_num_list:
										for chiplet_num in chiplet_num_list:
											WAO_L1 = WAO_L1_list[WAO_L1_type]
											WL1 = l1 * WAO_L1[0]; AL1 = l1 * WAO_L1[1]; OL1 = l1 * WAO_L1[2]
											WL2 =  WL1*core_num*WAO_L2;  AL2 = AL1 * core_num*WAO_L2; OL2 = OL1*core_num*WAO_L2

											if valid(PE_C,PE_K,core_num,chiplet_num,WL1,AL1,OL1,WL2,AL2,OL2):
												print ("--",all_valid_param)
												print (PE_C,PE_K,core_num,chiplet_num,WL1,AL1,OL1,WL2,AL2,OL2)
												all_valid_param += 1

												area_L1 = SRAM_area(WL1) + SRAM_area(AL1) + SRAM_area(OL1) # 不准确 ol1是reg file
												area_L2 = SRAM_area(WL2) + SRAM_area(AL2) + SRAM_area(OL2)
												area_pe_mac= PE_K * PE_C * area_MAC
												area_chip = area_L2 + (area_L1 + area_pe_mac + area_noc_router ) * core_num + 3* buffer_noc_router
												area_chip_mm2 = area_chip / 1000 / 1000
												
												HW_param = {"Chiplet":num2list[chiplet_num],"PE":num2list[core_num],"intra_PE":{"C":PE_C,"K":PE_K}}
												memory_param = {"OL1":OL1,"OL2":OL2,"AL1":AL1,"AL2":AL2,"WL1":WL1,"WL2":WL2}
													
												NoC_w = HW_param["PE"][1] + 1
												NOC_NODE_NUM = NoC_w * HW_param["PE"][0]
												NoP_w = HW_param["Chiplet"][1] + 1
												NOP_SIZE = NoP_w * HW_param["Chiplet"][0]
												
												TOPO_param = {"NoC_w":NoC_w, "NOC_NODE_NUM": NOC_NODE_NUM, "NoP_w": NoP_w, "NOP_SIZE": NOP_SIZE,"nop_scale_ratio": nop_bandwidth/noc_bandwidth }
												# --- 生成noc-nop结构图
												# NoC_param, all_sim_node_num = construct_noc_nop_topo(TOPO_param["NOC_NODE_NUM"],TOPO_param["NoC_w"], TOPO_param["NOP_SIZE"],TOPO_param["NoP_w"], TOPO_param["nop_scale_ratio"])
												debug=0
												if_multicast = 1
												assert (if_multicast ==1 )
												chiplet_parallel = "P_K_PK"			# choices: "Pq" "All" "Channel" "Hybrid" "P_K_PK"
												assert (chiplet_parallel == "P_K_PK")
												core_parallel = "All"
												GATest = GaEncode(network_param, HW_param, debug, chiplet_parallel = chiplet_parallel, core_parallel = core_parallel )
												iterTime = 5
												result_list,edp_min = randomTest(GATest, iterTime, HW_param, memory_param, if_multicast, filename)
												result_list.append(chiplet_num); result_list.append(core_num); result_list.append(PE_C); result_list.append(PE_K)
												result_list.append(WL1); result_list.append(AL1); result_list.append(OL1)
												result_list.append(WL2); result_list.append(AL2); result_list.append(OL2)
												result_list.append(l1); result_list.append(WAO_L1_type);result_list.append(WAO_L2) #new added
												result_list.append(area_chip_mm2)
												all_results.append(result_list)
																	

				for row, data in enumerate(all_results):
					for col, column_data in enumerate(data):
						sheet.cell(row+2, col+1, column_data)
				workbook.save(filename)
				print ("all_valid_param",all_valid_param)

if __name__ == '__main__':
	run_randomTest_granu_model( str(sys.argv[1]) ,  str(sys.argv[2]))

