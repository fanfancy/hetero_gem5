# -*- coding: utf-8 -*-
import openpyxl
from config import *
from inter_layer_noc_nop import setRouteTable_Mesh, setRouteTable_Ring, getInterLayerComm, calCommEnergy
import math

def calCommCycle(F, routeTable, bw_scales, commDict, topology = 'mesh'):
    F_cur=F.copy()
    link_comm_num_sum = 0
    # print('---------')
    for send_id in commDict:
        # print('send_id = ', send_id)
        for packet in commDict[send_id]:
            commNum = commDict[send_id][packet][0]
            commFlitNum = math.ceil(commNum / neu_per_flit_act_nop)
            dst_list = commDict[send_id][packet][1]
            # print('dst_list = ', dst_list)
            link_list = []
            for dst_id in dst_list:
                if topology == 'IODie' and (send_id // 4 == dst_id // 4):
                    continue
                # print('routeTable[(send_id + 1000, dst_id + 1000)] = ', routeTable[(send_id + 1000, dst_id + 1000)])
                for link in routeTable[(send_id + 1000, dst_id + 1000)]:
                    # print('link = ', link)
                    if link not in link_list:
                        link_list.append(link)
                        F_cur[link] += commFlitNum / bw_scales[link]
                        link_comm_num_sum += commNum
            # print('F_cur = ', F_cur)

    worstCommFlitNum = max(F_cur.values()) 
    worstlinks = []
    for item in F_cur:
        if F_cur[item] == max(F_cur.values()): 
            worstlinks.append(item)

    # print(F_cur)
    return F_cur, worstCommFlitNum, worstlinks, link_comm_num_sum

def setRouteTable_IODie(NOC_NODE_NUM):
    # mesh 4*4
    IODie_dict = {0:[0, 1, 2, 3], 4:[4, 5, 6, 7], 8:[8, 9, 10, 11], 12:[12, 13, 14, 15]}
    assert NOC_NODE_NUM == len(IODie_dict.keys()) * 4, "NOC_NODE_NUM != 16！"

    F = {}
    bw_scales = {}
    # local link
    for src in range (NOC_NODE_NUM):
        local = src + 1000
        F[(local, src)] = 0
        F[(src, local)] = 0
        bw_scales[(src, local)] = ddr_bandwidth / nop_bandwidth
        bw_scales[(local, src)] = ddr_bandwidth / nop_bandwidth

    # IODie link
    for src in IODie_dict.keys():
        src_x = (src / 4) % 2
        src_y = (src / 4) // 2
        for dst in IODie_dict[src]:
            F[(src, dst)] = 0
            bw_scales[(src, dst)] = 1

            F[(dst, src)] = 0
            bw_scales[(dst, src)] = 1

        for dst in IODie_dict.keys():
            dst_x = (dst / 4) % 2
            dst_y = (dst / 4) // 2
            if (src_x == dst_x) :
                if abs(dst_y - src_y) == 1:
                    F[(src, dst)] = 0
                    bw_scales[(src, dst)] = 1

                    F[(dst, src)] = 0
                    bw_scales[(dst, src)] = 1
            elif (src_y == dst_y) :
                if abs(src_x - dst_x) == 1:
                    F[(src, dst)] = 0
                    bw_scales[(src, dst)] = 1

                    F[(dst, src)] = 0
                    bw_scales[(dst, src)] = 1

    
    def cmesh_route(src, dst):
        # print(src, dst)
        if src in IODie_dict.keys():
            src_router = src
        else:
            for key in IODie_dict.keys():
                if src in IODie_dict[key]:
                    src_router = key
                    # print('IODie_dict[key] = ', IODie_dict[key])
                    break
        
        if dst in IODie_dict.keys():
            dst_router = dst
        else:
            for key in IODie_dict.keys():
                if dst in IODie_dict[key]:
                    dst_router = key
                    break
        
        route = []
        if src_router == dst_router:
            route.append((src, src_router))
            route.append((src_router, dst))
            return route
        
        if src != src_router:
            route.append((src, src_router))
        cur_src = src_router

        while cur_src != dst_router: # go to the first noc node
            src_noc_x = (cur_src / 4) %  2
            src_noc_y = (cur_src / 4) // 2
            dst_noc_x = (dst_router / 4) % 2
            dst_noc_y = (dst_router / 4) // 2

            if (src_noc_x > dst_noc_x):  # go west
                cur_noc_dst = (src_noc_x - 1 + src_noc_y * 2) * 4

            elif (src_noc_x < dst_noc_x): # go east
                cur_noc_dst = (src_noc_x + 1 + src_noc_y * 2) * 4

            elif (src_noc_y < dst_noc_y): # go north
                cur_noc_dst = (src_noc_x + (src_noc_y + 1) * 2) * 4

            elif (src_noc_y > dst_noc_y): # go south
                cur_noc_dst = (src_noc_x + (src_noc_y - 1) * 2) * 4

            route.append((cur_src, cur_noc_dst))
            cur_src = cur_noc_dst

        if dst != dst_router:     
            route.append((dst_router, dst))
        return route

    route_table = {}
    hops = {}
    for src in range (1000, 1000 + NOC_NODE_NUM):
        for dst in range (1000, 1000 + NOC_NODE_NUM):
            route_table[(src, dst)] = []
            noc_src = src - 1000
            noc_dst = dst - 1000
            if noc_src != noc_dst:
                route_table[(src, dst)] = cmesh_route(noc_src, noc_dst)
                route_table[(src, dst)].append((noc_dst, dst))
                route_table[(src, dst)].insert(0, (src, noc_src))
                F[(noc_dst, dst)] = 0
                F[(src, noc_src)] = 0
                bw_scales[(src, noc_src)] = 1
                bw_scales[(noc_dst, dst)] = 1
    
    print("F", F)
    print("route_table=========")
    for item in route_table:
        print(item, route_table[item])
    return route_table, F, bw_scales

if __name__ == '__main__':
    NOC_NODE_NUM = 16
    NoC_w = 4
    topology = 'ring'

    # 拓扑构建routing table
    if topology == "mesh":
        routeTable, F, bw_scales = setRouteTable_Mesh(NOC_NODE_NUM, NoC_w)
    elif topology == "ring":
        routeTable, F, bw_scales = setRouteTable_Ring(NOC_NODE_NUM)
    elif topology == 'IODie':
        routeTable, F, bw_scales = setRouteTable_IODie(NOC_NODE_NUM)
    else:
        raise NotImplementedError

    print('bw_scales = ', bw_scales)

    # 1表示前一层，2表示当前层
    # network_param_2 = {"P":224,"Q":224,"C":3,"K":64,"R":3,"S":3}
    network_param_2 = {"H":64,"M":64, "P":64,"Q":64,"C":64,"K":64,"R":3,"S":3, "stride":1, "padding":1}
    # network_param_2 = {"H":7,"M":7, "P":7,"Q":7,"C":512,"K":2048,"R":1,"S":1, "stride":1, "padding":0}
    
    parallel_type_1 = {'K':{"P":1,"Q":1,"K":16},"P":{"P":16,"Q":1,"K":1},'PK':{"P":4,"Q":1,"K":4}}
    parallel_type_2 = {'K':{"P":1,"Q":1,"K":16},"P":{"P":16,"Q":1,"K":1},'PK':{"P":4,"Q":1,"K":4}}
    # parallel_type_1 = {'P':{"P":16,"Q":1,"K":1}}
    # parallel_type_1 = {'K':{"P":1,"Q":1,"K":16}}
    # parallel_type_1 = {'P':{"P":16,"Q":1,"K":1}}
    # parallel_type_2 = {"PK":{"P":4,"Q":1,"K":4}}
    # parallel_type_2 = {'P':{"P":16,"Q":1,"K":1}}

    excel_datas = []

    times = 0
    for type2 in parallel_type_2:
        for type1 in parallel_type_1:
            parallel_1 = parallel_type_1[type1]
            parallel_2 = parallel_type_2[type2]
            dim_seq_1 = ["K","P","Q"]
            dim_seq_2 = ["K","P","Q"]
            print("----------- times = ", times)
            print("parallel_1 = ", parallel_1)
            print("parallel_2 = ", parallel_2)

            # 通信量计算
            comm_num_dict, comm_type_dict, comm_type_times_dict, chiplet_num = getInterLayerComm(dim_seq_1, dim_seq_2, parallel_1, network_param_2, parallel_2, 0)
            print("comm_num_dict", comm_num_dict)
            print("comm_type_dict", comm_type_dict)
            print("comm_type_times_dict", comm_type_times_dict)

            # 计算链路通信量
            F_cur, worstCommFlitNum, worstlinks, link_comm_num_sum = calCommCycle(F, routeTable, bw_scales, comm_num_dict, topology)
            
            print("F_cur", F_cur)
            print("worstCommFlitNum", worstCommFlitNum)
            print("worstlinks", worstlinks)
            print("link_comm_num_sum", link_comm_num_sum)
            # 计算片上计算时间
            d2d_energy, dram_access_energy = calCommEnergy(link_comm_num_sum, comm_type_dict)

            edp = 2 * (d2d_energy + dram_access_energy) * worstCommFlitNum / freq_1G / PE_freq
            
            print("d2d_energy", d2d_energy)
            print("dram_access_energy", dram_access_energy)
            print("edp ",edp)
            excel_datas.append([chiplet_num, type1+"-"+type2, comm_type_times_dict["uni-cast"], comm_type_times_dict["multi-cast"], comm_type_times_dict["broadcast"], comm_type_dict["uni-cast"], comm_type_dict["multi-cast"], comm_type_dict["broadcast"] , worstCommFlitNum, str(worstlinks), str(F_cur)])
            #print(comm_num_dict)
            times += 1
    
    workbook = openpyxl.Workbook()
    sheet = workbook.get_sheet_by_name('Sheet') 
    # 写入标题
    column_tite = ["chiplet_num", "parallel-type", "times-uni-cast","times-multi-cast","times-broadcast","num-uni-cast","num-multi-cast","num-broadcast", "worstCommFlitNum", "worstlinks", "F"]
    for col,column in enumerate(column_tite):
        sheet.cell(1, col+1, column)
    # 写入每一行
    for row, data in enumerate(excel_datas):
        for col, column_data in enumerate(data):
            sheet.cell(row+2, col+1, column_data)

    workbook.save("nop_topology_explore.xlsx")
