import math
import os
import sys
import random
import numpy as np
import copy
from enum import Enum
from matplotlib import pyplot as plt
import openpyxl
import xlrd
from inter_layer_noc_nop import *

def getLayerParam(app_name):
	layer_num = 0
	layer_dict = {}
	layer_id_list = []
	f = open("./nn_input_noc_nop/" + app_name + ".txt")

	print("network model ----- " + app_name + " -------------")

	lines = f.readlines()
	for line in lines:
		if line.startswith("#") or line.startswith("*"):
			pass
		elif line != "\n":
			line_item = line.split(" ")
			layer_num += 1
			H = int(line_item[1])
			M = int(line_item[2])
			C = int(line_item[3])
			R = int(line_item[4])
			S = int(line_item[4])
			stride = int(line_item[5])
			padding = int(line_item[6])
			K = int(line_item[7])
			P = int(line_item[8])
			Q = int(line_item[9])

			layer_id_list.append(layer_num)
			layer_dict[layer_num] = {"H":H,"M":M,"P":P,"Q":Q,"C":C,"K":K,"R":R,"S":S, "stride":stride, "padding":padding}
			print("layer" + str(layer_num) + " : " + str(layer_dict[layer_num]))
	f.close()
	return layer_dict, layer_id_list

def getIntraLayerEDP(app_name, file_type = "intra_layer", struct_type = ""):
	EDP_intra_layer = {}

	file_name = "./final_test/intra_layer_edp_parse/" + struct_type + app_name + "_" + file_type +".txt"
	f = open(file_name)
	lines = f.readlines()
	layer_num = 0

	for line in lines:
		if line.startswith("layer"):
			line_item = line.split("\t")
			layer_num += 1
			EDP_intra_layer[layer_num] = {}
			EDP_intra_layer[layer_num]["P"] = float(line_item[1])
			EDP_intra_layer[layer_num]["PK"] = float(line_item[2])
			EDP_intra_layer[layer_num]["K"] = float(line_item[3])
	print(file_type," : ", EDP_intra_layer)
	return EDP_intra_layer

def getIntraLayerEDP_simba(app_name, file_type = "intra_layer", struct_type = "simba"):
	EDP_intra_layer = {}

	file_name = "./final_test/intra_layer_edp/" + struct_type + "_" + app_name + "_" + file_type +".txt"
	f = open(file_name)
	lines = f.readlines()
	layer_num = 0

	for line in lines:
		if line.startswith("layer"):
			line_item = line.split("\t")
			layer_num += 1
			if file_type == "intra_layer_parallel":
				line_k = line_item[1].split(",")
				line_c = line_item[2].split(",")
				line_kc = line_item[3].split(",")
				EDP_intra_layer[layer_num] = {}
				EDP_intra_layer[layer_num]["K"] = {"K":int(line_k[0]),"C":int(line_k[1])}
				EDP_intra_layer[layer_num]["C"] = {"K":int(line_c[0]),"C":int(line_c[1])}
				EDP_intra_layer[layer_num]["KC"] = {"K":int(line_kc[0]),"C":int(line_kc[1])}
			else:
				EDP_intra_layer[layer_num] = {}
				EDP_intra_layer[layer_num]["K"] = float(line_item[1])
				EDP_intra_layer[layer_num]["C"] = float(line_item[2])
				EDP_intra_layer[layer_num]["KC"] = float(line_item[3])
	print(file_type," : ", EDP_intra_layer)
	return EDP_intra_layer

def getIntraLayerEDP_nnbaton(app_name, file_type = "intra_layer", struct_type = ""):
	EDP_intra_layer = {}

	file_name = "./final_test/intra_layer_edp/" + struct_type + app_name + "_" + file_type +".txt"
	f = open(file_name)
	lines = f.readlines()
	layer_num = 0

	for line in lines:
		if line.startswith("layer"):
			line_item = line.split("\t")
			layer_num += 1
			EDP_intra_layer[layer_num] = {}
			EDP_intra_layer[layer_num]["P"] = float(line_item[1])
			EDP_intra_layer[layer_num]["PK_1"] = float(line_item[2])
			EDP_intra_layer[layer_num]["PK_2"] = float(line_item[3])
			EDP_intra_layer[layer_num]["K"] = float(line_item[4])
	print(file_type," : ", EDP_intra_layer)
	return EDP_intra_layer

def nop_noc_test_ours(app_name):
	NOC_NODE_NUM = 16
	NoC_w = 4

	EDP_sum_dict = {"P":0, "K":0, "PK":0}
	EDP_inter_dict = {"P":0, "K":0, "PK":0}
	energy_sum_dict = {"P":0, "K":0, "PK":0}
	delay_sum_dict = {"P":0, "K":0, "PK":0}
	energy_inter_dict = {"P":0, "K":0, "PK":0}
	delay_inter_dict = {"P":0, "K":0, "PK":0}
	inter_layer_edp_dict = {}
	parallel_type_dict = {"P":[],"K":[],"PK":[]}

	para_type = ["P","K","PK"]
	parallel_type_1 = {"P":{"P":16,"Q":1,"K":1},'K':{"P":1,"Q":1,"K":16},'PK':{"P":4,"Q":1,"K":4}}
	parallel_type_2 = {"P":{"P":16,"Q":1,"K":1},'K':{"P":1,"Q":1,"K":16},'PK':{"P":4,"Q":1,"K":4}}

	file_name = "./final_test/noc_nop/" + app_name + "_noc_nop.txt"
	f = open(file_name,'w')

	# --- 获得神经网络参数
	layer_dict, layer_id_list = getLayerParam(app_name)
	# --- 获得片内EDP数值
	EDP_intra_layer = getIntraLayerEDP(app_name)
	energy_intra_layer = getIntraLayerEDP(app_name, "intra_layer_energy")
	delay_intra_layer = getIntraLayerEDP(app_name, "intra_layer_delay")

	excel_datas = []
	data_list = {"P":[],"PK":[],"K":[]}

	# ---初始（最下面一层的）NoC edp
	for para_type_2 in para_type:
		layer_id = layer_id_list[-1]
		EDP = EDP_intra_layer[layer_id][para_type_2]
		energy = energy_intra_layer[layer_id][para_type_2]
		delay = delay_intra_layer[layer_id][para_type_2]
		EDP_sum_dict[para_type_2] += EDP
		energy_sum_dict[para_type_2] += energy
		delay_sum_dict[para_type_2] += delay
		line = "parallel " + str(para_type_2) + "  noc_EDP " + str(EDP)
		print(line, file = f)
		data_list[para_type_2] = [EDP]
	print("energy_sum_dict",energy_sum_dict)
	print("delay_sum_dict",delay_sum_dict)

	for ii in range(12):
			data_list["P"].append(0)
			data_list["PK"].append(0)
			data_list["K"].append(0)
	
	excel_per_layer = ["layer"+str(layer_id_list[-1])] + data_list["P"] +data_list["PK"] + data_list["K"] + [str(energy_sum_dict), str(delay_sum_dict), str(EDP_inter_dict), str(EDP_sum_dict), str(parallel_type_dict)]

	excel_datas.append(excel_per_layer)
	

	for i in reversed(layer_id_list):
		if i < layer_id_list[-1]:
			inter_layer_edp_dict[i] = {}
			network_param_2 = layer_dict[i+1]
			parallel_type_dict_new = {"P":[], "K":[], "PK":[]}
			EDP_sum_dict_new = {"P":0, "K":0, "PK":0}
			EDP_inter_dict_new = {"P":0, "K":0, "PK":0}
			EDP_NoC_dict = {"P":0, "K":0, "PK":0}
			energy_NoC_dict = {"P":0, "K":0, "PK":0}
			delay_NoC_dict = {"P":0, "K":0, "PK":0}
			energy_sum_dict_new = {"P":0, "K":0, "PK":0}
			delay_sum_dict_new = {"P":0, "K":0, "PK":0}
			energy_inter_dict_new = {"P":0, "K":0, "PK":0}
			delay_inter_dict_new = {"P":0, "K":0, "PK":0}

			line = "layer " + str(i) + "--------------------------------------"
			print(line, file = f)
			excel_per_layer = []
			excel_per_layer.append(i)

			for para_type_1 in para_type:
				EDP = EDP_intra_layer[i][para_type_1]
				EDP_NoC_dict[para_type_1] = EDP
				energy_NoC_dict[para_type_1] = energy_intra_layer[i][para_type_1]
				delay_NoC_dict[para_type_1] = delay_intra_layer[i][para_type_1]
				#EDP_sum_dict[para_type_1] += EDP
				line = "parallel " + str(para_type_1) + "  noc_EDP " + str(EDP)
				print(line, file = f)
				data_list[para_type_1] = [EDP]

			# 计算层间的EDP
			for para_type_1 in para_type:
				EDP_num = {}
				energy_num = {}
				delay_num = {}
				energy_inter_num = {}
				delay_inter_num = {}
				EDP_num_only = {}
				inter_layer_edp_dict[i][para_type_1] = {}
				for para_type_2 in para_type:
					parallel_1 = parallel_type_1[para_type_1]
					parallel_2 = parallel_type_2[para_type_2]
					worstCommFlitNum, d2d_e , dram_e, EDP = getInterLayer(network_param_2, parallel_1, parallel_2, NOC_NODE_NUM, NoC_w)
					EDP_num[para_type_2] = EDP + EDP_sum_dict[para_type_2]
					EDP_num_only[para_type_2] = EDP
					energy_num[para_type_2] = d2d_e + dram_e + energy_sum_dict[para_type_2]
					delay_num[para_type_2] = worstCommFlitNum + delay_sum_dict[para_type_2]
					energy_inter_num[para_type_2] = d2d_e + dram_e + energy_inter_dict[para_type_2]
					delay_inter_num[para_type_2] = worstCommFlitNum + delay_inter_dict[para_type_2]

					data_list[para_type_1].append(d2d_e)
					data_list[para_type_1].append(dram_e)
					data_list[para_type_1].append(worstCommFlitNum)
					data_list[para_type_1].append(EDP)

					inter_layer_edp_dict[i][para_type_1][para_type_2] = [EDP, d2d_e, dram_e, worstCommFlitNum]
				
				line = "layer " + str(i) + " : parallel " + str(para_type_1)
				print(line, file = f)

				EDP_num_min = EDP_num["P"]
				para_min = "P"
				for pa in EDP_num:
					if EDP_num[pa] < EDP_num_min:
						EDP_num_min = EDP_num[pa]
						para_min = pa
				
				EDP_sum_dict_new[para_type_1] = EDP_num_min + EDP_NoC_dict[para_type_1]
				EDP_inter_dict_new[para_type_1] = EDP_inter_dict[para_min] + EDP_num_only[para_min]
				energy_sum_dict_new[para_type_1] = energy_num[para_min] + energy_NoC_dict[para_type_1]
				delay_sum_dict_new[para_type_1] = delay_num[para_min] + delay_NoC_dict[para_type_1]
				energy_inter_dict_new[para_type_1] = energy_inter_num[para_min]
				delay_inter_dict_new[para_type_1] = delay_inter_num[para_min]

				parallel_type_dict_new[para_type_1] = [para_min]
				parallel_type_dict_new[para_type_1] += parallel_type_dict[para_min]
				
				line2 = "---------nop_EDP = " +str(EDP_num_only)
				print(line2, file = f)
				line2 = "---------EDP_num = " +str(EDP_num)
				print(line2, file = f)

			parallel_type_dict = copy.deepcopy(parallel_type_dict_new)
			EDP_sum_dict = copy.deepcopy(EDP_sum_dict_new)
			EDP_inter_dict = copy.deepcopy(EDP_inter_dict_new)
			energy_sum_dict = copy.deepcopy(energy_sum_dict_new)
			delay_sum_dict = copy.deepcopy(delay_sum_dict_new)
			energy_inter_dict = copy.deepcopy(energy_inter_dict_new)
			delay_inter_dict = copy.deepcopy(delay_inter_dict_new)

			excel_per_layer = ["layer"+str(i)] + data_list["P"] +data_list["PK"] + data_list["K"] + [str(energy_sum_dict),str(delay_sum_dict), str(EDP_inter_dict), str(EDP_sum_dict), str(parallel_type_dict)]

			excel_datas.append(excel_per_layer)
			
		line3 = "----EDP_sum_dict = " + str(EDP_sum_dict)
		line1 = "----parallel_type_dict = " +str(parallel_type_dict)
		line4 = "----energy_sum_dict = " + str(energy_sum_dict)
		line5 = "----delay_sum_dict = " + str(delay_sum_dict)
		line6 = "----edp_inter_layer = " + str(EDP_inter_dict)
		line7 = "----energy_inter_layer = " + str(energy_inter_dict)
		line8 = "----delay_inter_layer = " + str(delay_inter_dict)
		print(line3, file = f)
		print(line1, file = f)
		print(line4, file = f)
		print(line5, file = f)
		print(line6, file = f)
		print(line7, file = f)
		print(line8, file = f)

	
	# 计算全K, 全P, 全PK
	print(inter_layer_edp_dict)
	edp_K, edp_K_inter = oneParallelTest(inter_layer_edp_dict, EDP_intra_layer, "K")
	edp_P, edp_P_inter = oneParallelTest(inter_layer_edp_dict, EDP_intra_layer, "P")
	edp_PK, edp_PK_inter = oneParallelTest(inter_layer_edp_dict, EDP_intra_layer, "PK")
	lineK = "EDP_ALL_K = " + str(edp_K) + "  ;  K_inter_layer = " + str(edp_K_inter)
	lineP = "EDP_ALL_P = " + str(edp_P) + "  ;  P_inter_layer = " + str(edp_P_inter)
	linePK = "EDP_ALL_PK = " + str(edp_PK) + "  ;  PK_inter_layer = " + str(edp_PK_inter)
	print(lineP, file = f)
	print(lineK, file = f)
	print(linePK, file = f)


	workbook = openpyxl.Workbook()
	sheet = workbook.get_sheet_by_name('Sheet') 
	# 写入标题
	column_tite = ["layer_id", "P_noc_edp", "P_P_d2d_e","P_P_dram_e","P_P_comm_d", "P_P_nop_edp", "P_K_d2d_e","P_K_dram_e","P_K_comm_d", "P_K_nop_edp", "P_PK_d2d_e","P_PK_dram_e","P_PK_comm_d", "P_PK_nop_edp"]
	column_tite += ["PK_noc_edp", "PK_P_d2d_e","PK_P_dram_e","PK_P_comm_d", "PK_P_nop_edp", "PK_K_d2d_e","PK_K_dram_e","PK_K_comm_d", "PK_K_nop_edp", "PK_PK_d2d_e","PK_PK_dram_e","PK_PK_comm_d", "PK_PK_nop_edp"]
	column_tite += ["K_noc_edp", "K_P_d2d_e","K_P_dram_e","K_P_comm_d", "K_P_nop_edp", "K_K_d2d_e","K_K_dram_e","K_K_comm_d", "K_K_nop_edp", "K_PK_d2d_e","K_PK_dram_e","K_PK_comm_d", "K_PK_nop_edp"]
	column_tite += ["energy_sum", "delay_sum", "EDP_inter_layer", "edp_sum", "par_index"]
	for col,column in enumerate(column_tite):
		sheet.cell(1, col+1, column)
	# 写入每一行
	for row, data in enumerate(excel_datas):
		for col, column_data in enumerate(data):
			sheet.cell(row+2, col+1, column_data)

	setname = str(network_param_2["P"]) + "_" + str(network_param_2["Q"]) + "_" + str(network_param_2["K"]) + "_" + str(network_param_2["C"]) + "_" + str(NoC_w) + "_" + str(NOC_NODE_NUM)

	workbook.save("./final_test/noc_nop_xls/"+app_name+"_noc_nop.xls")

def nop_noc_test_ours_iodie(app_name):
	NOC_NODE_NUM = 4
	NoC_w = 2

	EDP_sum_dict = {"P":0, "K":0, "PK":0}
	EDP_inter_dict = {"P":0, "K":0, "PK":0}
	energy_sum_dict = {"P":0, "K":0, "PK":0}
	delay_sum_dict = {"P":0, "K":0, "PK":0}
	energy_inter_dict = {"P":0, "K":0, "PK":0}
	delay_inter_dict = {"P":0, "K":0, "PK":0}
	inter_layer_edp_dict = {}
	parallel_type_dict = {"P":[],"K":[],"PK":[]}

	para_type = ["P","K","PK"]
	parallel_type_1 = {"P":{"P":4,"Q":1,"K":1},'K':{"P":1,"Q":1,"K":4},'PK':{"P":4,"Q":1,"K":1}}
	parallel_type_2 = {"P":{"P":4,"Q":1,"K":1},'K':{"P":1,"Q":1,"K":4},'PK':{"P":4,"Q":1,"K":1}}

	file_name = "./final_test/noc_nop/" + app_name + "_noc_nop.txt"
	f = open(file_name,'w')

	# --- 获得神经网络参数
	layer_dict, layer_id_list = getLayerParam(app_name)
	# --- 获得片内EDP数值
	EDP_intra_layer = getIntraLayerEDP(app_name,struct_type="ours_ours_")
	energy_intra_layer = getIntraLayerEDP(app_name, "intra_layer_energy","ours_ours_")
	delay_intra_layer = getIntraLayerEDP(app_name, "intra_layer_delay","ours_ours_")

	excel_datas = []
	data_list = {"P":[],"PK":[],"K":[]}

	# ---初始（最下面一层的）NoC edp
	for para_type_2 in para_type:
		layer_id = layer_id_list[-1]
		EDP = EDP_intra_layer[layer_id][para_type_2]
		energy = energy_intra_layer[layer_id][para_type_2]
		delay = delay_intra_layer[layer_id][para_type_2]
		EDP_sum_dict[para_type_2] += EDP
		energy_sum_dict[para_type_2] += energy
		delay_sum_dict[para_type_2] += delay
		line = "parallel " + str(para_type_2) + "  noc_EDP " + str(EDP)
		print(line, file = f)
		data_list[para_type_2] = [EDP]
	print("energy_sum_dict",energy_sum_dict)
	print("delay_sum_dict",delay_sum_dict)

	for ii in range(12):
			data_list["P"].append(0)
			data_list["PK"].append(0)
			data_list["K"].append(0)
	
	excel_per_layer = ["layer"+str(layer_id_list[-1])] + data_list["P"] +data_list["PK"] + data_list["K"] + [str(energy_sum_dict), str(delay_sum_dict), str(EDP_inter_dict), str(EDP_sum_dict), str(parallel_type_dict)]

	excel_datas.append(excel_per_layer)
	

	for i in reversed(layer_id_list):
		if i < layer_id_list[-1]:
			inter_layer_edp_dict[i] = {}
			network_param_2 = layer_dict[i+1]
			parallel_type_dict_new = {"P":[], "K":[], "PK":[]}
			EDP_sum_dict_new = {"P":0, "K":0, "PK":0}
			EDP_inter_dict_new = {"P":0, "K":0, "PK":0}
			EDP_NoC_dict = {"P":0, "K":0, "PK":0}
			energy_NoC_dict = {"P":0, "K":0, "PK":0}
			delay_NoC_dict = {"P":0, "K":0, "PK":0}
			energy_sum_dict_new = {"P":0, "K":0, "PK":0}
			delay_sum_dict_new = {"P":0, "K":0, "PK":0}
			energy_inter_dict_new = {"P":0, "K":0, "PK":0}
			delay_inter_dict_new = {"P":0, "K":0, "PK":0}

			line = "layer " + str(i) + "--------------------------------------"
			print(line, file = f)
			excel_per_layer = []
			excel_per_layer.append(i)

			for para_type_1 in para_type:
				EDP = EDP_intra_layer[i][para_type_1]
				EDP_NoC_dict[para_type_1] = EDP
				energy_NoC_dict[para_type_1] = energy_intra_layer[i][para_type_1]
				delay_NoC_dict[para_type_1] = delay_intra_layer[i][para_type_1]
				#EDP_sum_dict[para_type_1] += EDP
				line = "parallel " + str(para_type_1) + "  noc_EDP " + str(EDP)
				print(line, file = f)
				data_list[para_type_1] = [EDP]

			# 计算层间的EDP
			for para_type_1 in para_type:
				EDP_num = {}
				energy_num = {}
				delay_num = {}
				energy_inter_num = {}
				delay_inter_num = {}
				EDP_num_only = {}
				inter_layer_edp_dict[i][para_type_1] = {}
				for para_type_2 in para_type:
					parallel_1 = parallel_type_1[para_type_1]
					parallel_2 = parallel_type_2[para_type_2]
					worstCommFlitNum, d2d_e , dram_e, EDP = getInterLayer(network_param_2, parallel_1, parallel_2, NOC_NODE_NUM, NoC_w)
					EDP_num[para_type_2] = EDP + EDP_sum_dict[para_type_2]
					EDP_num_only[para_type_2] = EDP
					energy_num[para_type_2] = d2d_e + dram_e + energy_sum_dict[para_type_2]
					delay_num[para_type_2] = worstCommFlitNum + delay_sum_dict[para_type_2]
					energy_inter_num[para_type_2] = d2d_e + dram_e + energy_inter_dict[para_type_2]
					delay_inter_num[para_type_2] = worstCommFlitNum + delay_inter_dict[para_type_2]

					data_list[para_type_1].append(d2d_e)
					data_list[para_type_1].append(dram_e)
					data_list[para_type_1].append(worstCommFlitNum)
					data_list[para_type_1].append(EDP)

					inter_layer_edp_dict[i][para_type_1][para_type_2] = [EDP, d2d_e, dram_e, worstCommFlitNum]
				
				line = "layer " + str(i) + " : parallel " + str(para_type_1)
				print(line, file = f)

				EDP_num_min = EDP_num["P"]
				para_min = "P"
				for pa in EDP_num:
					if EDP_num[pa] < EDP_num_min:
						EDP_num_min = EDP_num[pa]
						para_min = pa
				
				EDP_sum_dict_new[para_type_1] = EDP_num_min + EDP_NoC_dict[para_type_1]
				EDP_inter_dict_new[para_type_1] = EDP_inter_dict[para_min] + EDP_num_only[para_min]
				energy_sum_dict_new[para_type_1] = energy_num[para_min] + energy_NoC_dict[para_type_1]
				delay_sum_dict_new[para_type_1] = delay_num[para_min] + delay_NoC_dict[para_type_1]
				energy_inter_dict_new[para_type_1] = energy_inter_num[para_min]
				delay_inter_dict_new[para_type_1] = delay_inter_num[para_min]

				parallel_type_dict_new[para_type_1] = [para_min]
				parallel_type_dict_new[para_type_1] += parallel_type_dict[para_min]
				
				line2 = "---------nop_EDP = " +str(EDP_num_only)
				print(line2, file = f)
				line2 = "---------EDP_num = " +str(EDP_num)
				print(line2, file = f)

			parallel_type_dict = copy.deepcopy(parallel_type_dict_new)
			EDP_sum_dict = copy.deepcopy(EDP_sum_dict_new)
			EDP_inter_dict = copy.deepcopy(EDP_inter_dict_new)
			energy_sum_dict = copy.deepcopy(energy_sum_dict_new)
			delay_sum_dict = copy.deepcopy(delay_sum_dict_new)
			energy_inter_dict = copy.deepcopy(energy_inter_dict_new)
			delay_inter_dict = copy.deepcopy(delay_inter_dict_new)

			excel_per_layer = ["layer"+str(i)] + data_list["P"] +data_list["PK"] + data_list["K"] + [str(energy_sum_dict),str(delay_sum_dict), str(EDP_inter_dict), str(EDP_sum_dict), str(parallel_type_dict)]

			excel_datas.append(excel_per_layer)
			
		line3 = "----EDP_sum_dict = " + str(EDP_sum_dict)
		line1 = "----parallel_type_dict = " +str(parallel_type_dict)
		line4 = "----energy_sum_dict = " + str(energy_sum_dict)
		line5 = "----delay_sum_dict = " + str(delay_sum_dict)
		line6 = "----edp_inter_layer = " + str(EDP_inter_dict)
		line7 = "----energy_inter_layer = " + str(energy_inter_dict)
		line8 = "----delay_inter_layer = " + str(delay_inter_dict)
		print(line3, file = f)
		print(line1, file = f)
		print(line4, file = f)
		print(line5, file = f)
		print(line6, file = f)
		print(line7, file = f)
		print(line8, file = f)

	
	# 计算全K, 全P, 全PK
	print(inter_layer_edp_dict)
	edp_K, edp_K_inter = oneParallelTest(inter_layer_edp_dict, EDP_intra_layer, "K")
	edp_P, edp_P_inter = oneParallelTest(inter_layer_edp_dict, EDP_intra_layer, "P")
	edp_PK, edp_PK_inter = oneParallelTest(inter_layer_edp_dict, EDP_intra_layer, "PK")
	lineK = "EDP_ALL_K = " + str(edp_K) + "  ;  K_inter_layer = " + str(edp_K_inter)
	lineP = "EDP_ALL_P = " + str(edp_P) + "  ;  P_inter_layer = " + str(edp_P_inter)
	linePK = "EDP_ALL_PK = " + str(edp_PK) + "  ;  PK_inter_layer = " + str(edp_PK_inter)
	print(lineP, file = f)
	print(lineK, file = f)
	print(linePK, file = f)


	workbook = openpyxl.Workbook()
	sheet = workbook.get_sheet_by_name('Sheet') 
	# 写入标题
	column_tite = ["layer_id", "P_noc_edp", "P_P_d2d_e","P_P_dram_e","P_P_comm_d", "P_P_nop_edp", "P_K_d2d_e","P_K_dram_e","P_K_comm_d", "P_K_nop_edp", "P_PK_d2d_e","P_PK_dram_e","P_PK_comm_d", "P_PK_nop_edp"]
	column_tite += ["PK_noc_edp", "PK_P_d2d_e","PK_P_dram_e","PK_P_comm_d", "PK_P_nop_edp", "PK_K_d2d_e","PK_K_dram_e","PK_K_comm_d", "PK_K_nop_edp", "PK_PK_d2d_e","PK_PK_dram_e","PK_PK_comm_d", "PK_PK_nop_edp"]
	column_tite += ["K_noc_edp", "K_P_d2d_e","K_P_dram_e","K_P_comm_d", "K_P_nop_edp", "K_K_d2d_e","K_K_dram_e","K_K_comm_d", "K_K_nop_edp", "K_PK_d2d_e","K_PK_dram_e","K_PK_comm_d", "K_PK_nop_edp"]
	column_tite += ["energy_sum", "delay_sum", "EDP_inter_layer", "edp_sum", "par_index"]
	for col,column in enumerate(column_tite):
		sheet.cell(1, col+1, column)
	# 写入每一行
	for row, data in enumerate(excel_datas):
		for col, column_data in enumerate(data):
			sheet.cell(row+2, col+1, column_data)

	setname = str(network_param_2["P"]) + "_" + str(network_param_2["Q"]) + "_" + str(network_param_2["K"]) + "_" + str(network_param_2["C"]) + "_" + str(NoC_w) + "_" + str(NOC_NODE_NUM)

	workbook.save("./final_test/noc_nop_xls/"+app_name+"_noc_nop.xls")


def nop_noc_test_simba_ours(app_name):
	NOC_NODE_NUM = 64
	NoC_w = 8

	EDP_sum_dict = {"P":0, "K":0, "PK":0}
	EDP_inter_dict = {"P":0, "K":0, "PK":0}
	energy_sum_dict = {"P":0, "K":0, "PK":0}
	delay_sum_dict = {"P":0, "K":0, "PK":0}
	energy_inter_dict = {"P":0, "K":0, "PK":0}
	delay_inter_dict = {"P":0, "K":0, "PK":0}
	inter_layer_edp_dict = {}
	parallel_type_dict = {"P":[],"K":[],"PK":[]}

	para_type = ["P","K","PK"]
	parallel_type_1 = {"P":{"P":64,"Q":1,"K":1},'K':{"P":1,"Q":1,"K":64},'PK':{"P":8,"Q":1,"K":8}}
	parallel_type_2 = {"P":{"P":64,"Q":1,"K":1},'K':{"P":1,"Q":1,"K":64},'PK':{"P":8,"Q":1,"K":8}}

	file_name = "./final_test/noc_nop/simba_ours_" + app_name + "_noc_nop.txt"
	f = open(file_name,'w')

	# --- 获得神经网络参数
	layer_dict, layer_id_list = getLayerParam(app_name)
	# --- 获得片内EDP数值
	EDP_intra_layer = getIntraLayerEDP(app_name, struct_type="simba_ours_")
	energy_intra_layer = getIntraLayerEDP(app_name, "intra_layer_energy", struct_type="simba_ours_")
	delay_intra_layer = getIntraLayerEDP(app_name, "intra_layer_delay", struct_type="simba_ours_")

	excel_datas = []
	data_list = {"P":[],"PK":[],"K":[]}

	# ---初始（最下面一层的）NoC edp
	for para_type_2 in para_type:
		layer_id = layer_id_list[-1]
		EDP = EDP_intra_layer[layer_id][para_type_2]
		energy = energy_intra_layer[layer_id][para_type_2]
		delay = delay_intra_layer[layer_id][para_type_2]
		EDP_sum_dict[para_type_2] += EDP
		energy_sum_dict[para_type_2] += energy
		delay_sum_dict[para_type_2] += delay
		line = "parallel " + str(para_type_2) + "  noc_EDP " + str(EDP)
		print(line, file = f)
		data_list[para_type_2] = [EDP]
	print("energy_sum_dict",energy_sum_dict)
	print("delay_sum_dict",delay_sum_dict)

	for ii in range(12):
			data_list["P"].append(0)
			data_list["PK"].append(0)
			data_list["K"].append(0)
	
	excel_per_layer = ["layer"+str(layer_id_list[-1])] + data_list["P"] +data_list["PK"] + data_list["K"] + [str(energy_sum_dict), str(delay_sum_dict), str(EDP_inter_dict), str(EDP_sum_dict), str(parallel_type_dict)]

	excel_datas.append(excel_per_layer)
	

	for i in reversed(layer_id_list):
		if i < layer_id_list[-1]:
			inter_layer_edp_dict[i] = {}
			network_param_2 = layer_dict[i+1]
			parallel_type_dict_new = {"P":[], "K":[], "PK":[]}
			EDP_sum_dict_new = {"P":0, "K":0, "PK":0}
			EDP_inter_dict_new = {"P":0, "K":0, "PK":0}
			EDP_NoC_dict = {"P":0, "K":0, "PK":0}
			energy_NoC_dict = {"P":0, "K":0, "PK":0}
			delay_NoC_dict = {"P":0, "K":0, "PK":0}
			energy_sum_dict_new = {"P":0, "K":0, "PK":0}
			delay_sum_dict_new = {"P":0, "K":0, "PK":0}
			energy_inter_dict_new = {"P":0, "K":0, "PK":0}
			delay_inter_dict_new = {"P":0, "K":0, "PK":0}

			line = "layer " + str(i) + "--------------------------------------"
			print(line, file = f)
			excel_per_layer = []
			excel_per_layer.append(i)

			for para_type_1 in para_type:
				EDP = EDP_intra_layer[i][para_type_1]
				EDP_NoC_dict[para_type_1] = EDP
				energy_NoC_dict[para_type_1] = energy_intra_layer[i][para_type_1]
				delay_NoC_dict[para_type_1] = delay_intra_layer[i][para_type_1]
				#EDP_sum_dict[para_type_1] += EDP
				line = "parallel " + str(para_type_1) + "  noc_EDP " + str(EDP)
				print(line, file = f)
				data_list[para_type_1] = [EDP]

			# 计算层间的EDP
			for para_type_1 in para_type:
				EDP_num = {}
				energy_num = {}
				delay_num = {}
				energy_inter_num = {}
				delay_inter_num = {}
				EDP_num_only = {}
				inter_layer_edp_dict[i][para_type_1] = {}
				for para_type_2 in para_type:
					parallel_1 = parallel_type_1[para_type_1]
					parallel_2 = parallel_type_2[para_type_2]
					worstCommFlitNum, d2d_e , dram_e, EDP = getInterLayer(network_param_2, parallel_1, parallel_2, NOC_NODE_NUM, NoC_w)
					EDP_num[para_type_2] = EDP + EDP_sum_dict[para_type_2]
					EDP_num_only[para_type_2] = EDP
					energy_num[para_type_2] = d2d_e + dram_e + energy_sum_dict[para_type_2]
					delay_num[para_type_2] = worstCommFlitNum + delay_sum_dict[para_type_2]
					energy_inter_num[para_type_2] = d2d_e + dram_e + energy_inter_dict[para_type_2]
					delay_inter_num[para_type_2] = worstCommFlitNum + delay_inter_dict[para_type_2]

					data_list[para_type_1].append(d2d_e)
					data_list[para_type_1].append(dram_e)
					data_list[para_type_1].append(worstCommFlitNum)
					data_list[para_type_1].append(EDP)

					inter_layer_edp_dict[i][para_type_1][para_type_2] = [EDP, d2d_e, dram_e, worstCommFlitNum]
				
				line = "layer " + str(i) + " : parallel " + str(para_type_1)
				print(line, file = f)

				EDP_num_min = EDP_num["P"]
				para_min = "P"
				for pa in EDP_num:
					if EDP_num[pa] < EDP_num_min:
						EDP_num_min = EDP_num[pa]
						para_min = pa
				
				EDP_sum_dict_new[para_type_1] = EDP_num_min + EDP_NoC_dict[para_type_1]
				EDP_inter_dict_new[para_type_1] = EDP_inter_dict[para_min] + EDP_num_only[para_min]
				energy_sum_dict_new[para_type_1] = energy_num[para_min] + energy_NoC_dict[para_type_1]
				delay_sum_dict_new[para_type_1] = delay_num[para_min] + delay_NoC_dict[para_type_1]
				energy_inter_dict_new[para_type_1] = energy_inter_num[para_min]
				delay_inter_dict_new[para_type_1] = delay_inter_num[para_min]

				parallel_type_dict_new[para_type_1] = [para_min]
				parallel_type_dict_new[para_type_1] += parallel_type_dict[para_min]
				
				line2 = "---------nop_EDP = " +str(EDP_num_only)
				print(line2, file = f)
				line2 = "---------EDP_num = " +str(EDP_num)
				print(line2, file = f)

			parallel_type_dict = copy.deepcopy(parallel_type_dict_new)
			EDP_sum_dict = copy.deepcopy(EDP_sum_dict_new)
			EDP_inter_dict = copy.deepcopy(EDP_inter_dict_new)
			energy_sum_dict = copy.deepcopy(energy_sum_dict_new)
			delay_sum_dict = copy.deepcopy(delay_sum_dict_new)
			energy_inter_dict = copy.deepcopy(energy_inter_dict_new)
			delay_inter_dict = copy.deepcopy(delay_inter_dict_new)

			excel_per_layer = ["layer"+str(i)] + data_list["P"] +data_list["PK"] + data_list["K"] + [str(energy_sum_dict),str(delay_sum_dict), str(EDP_inter_dict), str(EDP_sum_dict), str(parallel_type_dict)]

			excel_datas.append(excel_per_layer)
			
		line3 = "----EDP_sum_dict = " + str(EDP_sum_dict)
		line1 = "----parallel_type_dict = " +str(parallel_type_dict)
		line4 = "----energy_sum_dict = " + str(energy_sum_dict)
		line5 = "----delay_sum_dict = " + str(delay_sum_dict)
		line6 = "----edp_inter_layer = " + str(EDP_inter_dict)
		line7 = "----energy_inter_layer = " + str(energy_inter_dict)
		line8 = "----delay_inter_layer = " + str(delay_inter_dict)
		print(line3, file = f)
		print(line1, file = f)
		print(line4, file = f)
		print(line5, file = f)
		print(line6, file = f)
		print(line7, file = f)
		print(line8, file = f)

	
	# 计算全K, 全P, 全PK
	print(inter_layer_edp_dict)
	edp_K, edp_K_inter = oneParallelTest(inter_layer_edp_dict, EDP_intra_layer, "K")
	edp_P, edp_P_inter = oneParallelTest(inter_layer_edp_dict, EDP_intra_layer, "P")
	edp_PK, edp_PK_inter = oneParallelTest(inter_layer_edp_dict, EDP_intra_layer, "PK")
	lineK = "EDP_ALL_K = " + str(edp_K) + "  ;  K_inter_layer = " + str(edp_K_inter)
	lineP = "EDP_ALL_P = " + str(edp_P) + "  ;  P_inter_layer = " + str(edp_P_inter)
	linePK = "EDP_ALL_PK = " + str(edp_PK) + "  ;  PK_inter_layer = " + str(edp_PK_inter)
	print(lineP, file = f)
	print(lineK, file = f)
	print(linePK, file = f)


	workbook = openpyxl.Workbook()
	sheet = workbook.get_sheet_by_name('Sheet') 
	# 写入标题
	column_tite = ["layer_id", "P_noc_edp", "P_P_d2d_e","P_P_dram_e","P_P_comm_d", "P_P_nop_edp", "P_K_d2d_e","P_K_dram_e","P_K_comm_d", "P_K_nop_edp", "P_PK_d2d_e","P_PK_dram_e","P_PK_comm_d", "P_PK_nop_edp"]
	column_tite += ["PK_noc_edp", "PK_P_d2d_e","PK_P_dram_e","PK_P_comm_d", "PK_P_nop_edp", "PK_K_d2d_e","PK_K_dram_e","PK_K_comm_d", "PK_K_nop_edp", "PK_PK_d2d_e","PK_PK_dram_e","PK_PK_comm_d", "PK_PK_nop_edp"]
	column_tite += ["K_noc_edp", "K_P_d2d_e","K_P_dram_e","K_P_comm_d", "K_P_nop_edp", "K_K_d2d_e","K_K_dram_e","K_K_comm_d", "K_K_nop_edp", "K_PK_d2d_e","K_PK_dram_e","K_PK_comm_d", "K_PK_nop_edp"]
	column_tite += ["energy_sum", "delay_sum", "EDP_inter_layer", "edp_sum", "par_index"]
	for col,column in enumerate(column_tite):
		sheet.cell(1, col+1, column)
	# 写入每一行
	for row, data in enumerate(excel_datas):
		for col, column_data in enumerate(data):
			sheet.cell(row+2, col+1, column_data)

	workbook.save("./final_test/noc_nop_xls/simba_ours_"+app_name+"_noc_nop.xls")

def calPSumAllReduce(output_num, chiplet_num, PC3):
	output_flit_num = int(output_num / neu_per_flit_psum_nop)
	act_flit_num = int(output_num / neu_per_flit_act_nop)
	delay = (output_flit_num / chiplet_num) * (PC3-1) + (act_flit_num / chiplet_num) * (PC3-1)
	d2d_energy = (output_num * psum_width / chiplet_num) * (PC3-1) * chiplet_num * DIE2DIE_energy_ratio + (output_num * act_wgt_width / chiplet_num) * (PC3-1) * chiplet_num * DIE2DIE_energy_ratio
	if PC3 == 1:
		dram_energy = 0
	else:
		dram_energy = (output_num * psum_width / chiplet_num) * PC3 * chiplet_num * SRAM_energy(1) + (output_num * act_wgt_width / chiplet_num) * PC3 * chiplet_num * SRAM_energy(1)
	energy_list = [d2d_energy, dram_energy, d2d_energy+dram_energy]
	return delay, energy_list

def nop_noc_test_simba(app_name, struct = "simba"):
	NOC_NODE_NUM = 16
	NoC_w = 4

	EDP_sum_dict = {"K":0, "C":0, "KC":0}
	EDP_inter_dict = {"K":0, "C":0, "KC":0}
	energy_sum_dict = {"K":0, "C":0, "KC":0}
	delay_sum_dict = {"K":0, "C":0, "KC":0}
	energy_inter_dict = {"K":0, "C":0, "KC":0}
	delay_inter_dict = {"K":0, "C":0, "KC":0}
	inter_layer_edp_dict = {}
	delay_psum_dict = {"K":0, "C":0, "KC":0}
	energy_psum_dict = {"K":0, "C":0, "KC":0}
	edp_psum_dict = {"K":0, "C":0, "KC":0}
	parallel_type_dict = {"K":[],"C":[],"KC":[]}

	para_type = ["K","KC","C"]
	parallel_type_1 = {"K":{"K":64,"C":1}, "KC":{"K":8,"C":8}, "C":{"K":1,"C":64}}
	parallel_type_2 = {"K":{"K":64,"C":1}, "KC":{"K":8,"C":8}, "C":{"K":1,"C":64}}
	
	file_name = "./final_test/noc_nop/" + struct +"_"+ app_name + "_noc_nop.txt"
	f = open(file_name,'w')

	# --- 获得神经网络参数
	layer_dict, layer_id_list = getLayerParam(app_name)
	# --- 获得片内EDP数值
	EDP_intra_layer = getIntraLayerEDP_simba(app_name, struct_type = struct)
	energy_intra_layer = getIntraLayerEDP_simba(app_name, "intra_layer_energy", struct_type = struct)
	delay_intra_layer = getIntraLayerEDP_simba(app_name, "intra_layer_delay", struct_type = struct)
	para_each_layer = getIntraLayerEDP_simba(app_name, "intra_layer_parallel", struct_type = struct)

	excel_datas = []
	data_list = {"P":[],"PK":[],"K":[]}

	# 计算psum
	Psum_edp_dict = {}
	Psum_energy_dict = {}
	Psum_delay_dict = {}
	for i in layer_id_list:
		network_param = layer_dict[i]
		output_num = network_param["P"]*network_param["Q"] *network_param["K"]
		Psum_edp_dict[i] = {}
		Psum_energy_dict[i] = {}
		Psum_delay_dict[i] = {}
		for par_type in para_type:
			PC3 = para_each_layer[i][par_type]["C"]
			chiplet_num = para_each_layer[i][par_type]["C"] * para_each_layer[i][par_type]["K"]
			psum_delay, psum_energy_list = calPSumAllReduce(output_num, chiplet_num, PC3)
			Psum_edp_dict[i][par_type] = psum_delay * psum_energy_list[2] / freq_1G / PE_freq
			Psum_energy_dict[i][par_type] = psum_energy_list[2]
			Psum_delay_dict[i][par_type] = psum_delay

	# ---初始（最下面一层的）NoC edp + psum
	for para_type_2 in para_type:
		layer_id = layer_id_list[-1]
		EDP = EDP_intra_layer[layer_id][para_type_2]
		energy = energy_intra_layer[layer_id][para_type_2]
		delay = delay_intra_layer[layer_id][para_type_2]
		EDP_sum_dict[para_type_2] += EDP + Psum_edp_dict[layer_id][para_type_2]
		energy_sum_dict[para_type_2] += energy + Psum_energy_dict[layer_id][para_type_2]
		delay_sum_dict[para_type_2] += delay + Psum_delay_dict[layer_id][para_type_2]

		delay_psum_dict[para_type_2] = Psum_delay_dict[layer_id][para_type_2]
		energy_psum_dict[para_type_2] = Psum_energy_dict[layer_id][para_type_2]
		edp_psum_dict[para_type_2] = Psum_edp_dict[layer_id][para_type_2]

		line = "parallel " + str(para_type_2) + "  noc_EDP " + str(EDP) + "  psum_EDP " + str(Psum_edp_dict[layer_id][para_type_2]) + "  psum_energy " + str(Psum_energy_dict[layer_id][para_type_2]) + "  psum_delay " + str(Psum_delay_dict[layer_id][para_type_2])
		print(line, file = f)
		data_list[para_type_2] = [EDP]
	print("energy_sum_dict",energy_sum_dict)
	print("delay_sum_dict",delay_sum_dict)

	for ii in range(12):
			data_list["P"].append(0)
			data_list["PK"].append(0)
			data_list["K"].append(0)
	
	excel_per_layer = ["layer"+str(layer_id_list[-1])] + data_list["P"] +data_list["PK"] + data_list["K"] + [str(energy_sum_dict), str(delay_sum_dict), str(EDP_inter_dict), str(EDP_sum_dict), str(parallel_type_dict)]

	excel_datas.append(excel_per_layer)
	
	for i in reversed(layer_id_list):
		if i < layer_id_list[-1]:
			inter_layer_edp_dict[i] = {}
			network_param_2 = layer_dict[i+1]
			parallel_type_dict_new = {"K":[], "C":[], "KC":[]}
			EDP_sum_dict_new = {"K":0, "C":0, "KC":0}
			EDP_inter_dict_new = {"K":0, "C":0, "KC":0}
			EDP_NoC_dict = {"K":0, "C":0, "KC":0}
			energy_NoC_dict = {"K":0, "C":0, "KC":0}
			delay_NoC_dict = {"K":0, "C":0, "KC":0}
			energy_sum_dict_new = {"K":0, "C":0, "KC":0}
			delay_sum_dict_new = {"K":0, "C":0, "KC":0}
			energy_inter_dict_new = {"K":0, "C":0, "KC":0}
			delay_inter_dict_new = {"K":0, "C":0, "KC":0}

			energy_psum_dict_new = {"K":0, "C":0, "KC":0}
			delay_psum_dict_new = {"K":0, "C":0, "KC":0}
			edp_psum_dict_new = {"K":0, "C":0, "KC":0}
			energy_psum =  {"K":0, "C":0, "KC":0}
			delay_psum =  {"K":0, "C":0, "KC":0}
			edp_psum =  {"K":0, "C":0, "KC":0}

			line = "layer " + str(i) + "--------------------------------------"
			print(line, file = f)
			excel_per_layer = []
			excel_per_layer.append(i)

			for para_type_1 in para_type:
				EDP = EDP_intra_layer[i][para_type_1]
				EDP_NoC_dict[para_type_1] = EDP
				energy_NoC_dict[para_type_1] = energy_intra_layer[i][para_type_1]
				delay_NoC_dict[para_type_1] = delay_intra_layer[i][para_type_1]
				#EDP_sum_dict[para_type_1] += EDP
				data_list[para_type_1] = [EDP]
				energy_psum[para_type_1] = Psum_energy_dict[i][para_type_1]
				delay_psum[para_type_1] = Psum_delay_dict[i][para_type_1]
				edp_psum[para_type_1] = Psum_edp_dict[i][para_type_1]
				line = "parallel " + str(para_type_1) + "  noc_EDP " + str(EDP) + "  psum_EDP " + str(Psum_edp_dict[i][para_type_1]) + "  psum_energy " + str(Psum_energy_dict[i][para_type_1]) + "  psum_delay " + str(Psum_delay_dict[i][para_type_1])
				print(line, file = f)

			# 计算层间的EDP
			for para_type_1 in para_type:
				EDP_num = {}
				energy_num = {}
				delay_num = {}
				energy_inter_num = {}
				delay_inter_num = {}
				EDP_num_only = {}
				inter_layer_edp_dict[i][para_type_1] = {}
				par_C_pre = para_each_layer[i][para_type_1]["C"]
				par_K_pre = para_each_layer[i][para_type_1]["K"]
				for para_type_2 in para_type:
					par_C_cur = para_each_layer[i+1][para_type_2]["C"]
					par_K_cur = para_each_layer[i+1][para_type_2]["K"]

					worstCommFlitNum, d2d_e , dram_e, EDP = calInterComm_simba(par_C_pre, par_K_pre, par_C_cur, par_K_cur, network_param_2)
					
					EDP_num[para_type_2] = EDP + EDP_sum_dict[para_type_2]
					EDP_num_only[para_type_2] = EDP
					energy_num[para_type_2] = d2d_e + dram_e + energy_sum_dict[para_type_2]
					delay_num[para_type_2] = worstCommFlitNum + delay_sum_dict[para_type_2]
					energy_inter_num[para_type_2] = d2d_e + dram_e + energy_inter_dict[para_type_2]
					delay_inter_num[para_type_2] = worstCommFlitNum + delay_inter_dict[para_type_2]

					data_list[para_type_1].append(d2d_e)
					data_list[para_type_1].append(dram_e)
					data_list[para_type_1].append(worstCommFlitNum)
					data_list[para_type_1].append(EDP)

					inter_layer_edp_dict[i][para_type_1][para_type_2] = [EDP, d2d_e, dram_e, worstCommFlitNum]
				
				line = "layer " + str(i) + " : parallel " + str(para_type_1)
				print(line, file = f)

				EDP_num_min = EDP_num["K"]
				para_min = "K"
				for pa in EDP_num:
					if EDP_num[pa] < EDP_num_min:
						EDP_num_min = EDP_num[pa]
						para_min = pa
				
				EDP_sum_dict_new[para_type_1] = EDP_num_min + EDP_NoC_dict[para_type_1] + edp_psum[para_type_1]
				EDP_inter_dict_new[para_type_1] = EDP_inter_dict[para_min] + EDP_num_only[para_min]
				energy_sum_dict_new[para_type_1] = energy_num[para_min] + energy_NoC_dict[para_type_1] + energy_psum[para_type_1]
				delay_sum_dict_new[para_type_1] = delay_num[para_min] + delay_NoC_dict[para_type_1] + delay_psum[para_type_1]
				energy_inter_dict_new[para_type_1] = energy_inter_num[para_min]
				delay_inter_dict_new[para_type_1] = delay_inter_num[para_min]

				energy_psum_dict_new[para_type_1] = energy_psum_dict[para_min] +  energy_psum[para_type_1]
				delay_psum_dict_new[para_type_1] = delay_psum_dict[para_min] +  delay_psum[para_type_1]
				edp_psum_dict_new[para_type_1] = edp_psum_dict[para_min] +  edp_psum[para_type_1]

				parallel_type_dict_new[para_type_1] = [para_min]
				parallel_type_dict_new[para_type_1] += parallel_type_dict[para_min]
				
				line2 = "---------nop_EDP = " +str(EDP_num_only)
				print(line2, file = f)
				line2 = "---------Psum_EDP = " +str(edp_psum)
				print(line2, file = f)
				line2 = "---------EDP_num = " +str(EDP_num)
				print(line2, file = f)


			parallel_type_dict = copy.deepcopy(parallel_type_dict_new)
			EDP_sum_dict = copy.deepcopy(EDP_sum_dict_new)
			EDP_inter_dict = copy.deepcopy(EDP_inter_dict_new)
			energy_sum_dict = copy.deepcopy(energy_sum_dict_new)
			delay_sum_dict = copy.deepcopy(delay_sum_dict_new)
			energy_inter_dict = copy.deepcopy(energy_inter_dict_new)
			delay_inter_dict = copy.deepcopy(delay_inter_dict_new)
			edp_psum_dict = copy.deepcopy(edp_psum_dict_new)
			energy_psum_dict = copy.deepcopy(energy_psum_dict_new)
			delay_psum_dict = copy.deepcopy(delay_psum_dict_new)

			excel_per_layer = ["layer"+str(i)] + data_list["P"] +data_list["PK"] + data_list["K"] + [str(energy_sum_dict),str(delay_sum_dict), str(EDP_inter_dict), str(EDP_sum_dict), str(parallel_type_dict)]

			excel_datas.append(excel_per_layer)
			
		line3 = "----EDP_sum_dict = " + str(EDP_sum_dict)
		line1 = "----parallel_type_dict = " +str(parallel_type_dict)
		line4 = "----energy_sum_dict = " + str(energy_sum_dict)
		line5 = "----delay_sum_dict = " + str(delay_sum_dict)
		line6 = "----edp_inter_layer = " + str(EDP_inter_dict)
		line7 = "----energy_inter_layer = " + str(energy_inter_dict)
		line8 = "----delay_inter_layer = " + str(delay_inter_dict)
		line9 = "----edp_psum_dict = " + str(edp_psum_dict)
		line10 = "----energy_psum_dict = " + str(energy_psum_dict)
		line11 = "----delay_psum_dict = " + str(delay_psum_dict)
		print(line3, file = f)
		print(line1, file = f)
		print(line4, file = f)
		print(line5, file = f)
		print(line6, file = f)
		print(line7, file = f)
		print(line8, file = f)
		print(line9, file = f)
		print(line10, file = f)
		print(line11, file = f)

	
	# 计算全K, 全P, 全PK
	print(inter_layer_edp_dict)
	edp_K, edp_K_inter = oneParallelTest(inter_layer_edp_dict, EDP_intra_layer, "K")
	edp_C, edp_C_inter = oneParallelTest(inter_layer_edp_dict, EDP_intra_layer, "C")
	edp_KC, edp_KC_inter = oneParallelTest(inter_layer_edp_dict, EDP_intra_layer, "KC")
	lineK = "EDP_ALL_K = " + str(edp_K) + "  ;  K_inter_layer = " + str(edp_K_inter)
	lineC = "EDP_ALL_C = " + str(edp_C) + "  ;  C_inter_layer = " + str(edp_C_inter)
	lineKC = "EDP_ALL_KC = " + str(edp_KC) + "  ;  KC_inter_layer = " + str(edp_KC_inter)
	print(lineK, file = f)
	print(lineC, file = f)
	print(lineKC, file = f)
	print("Psum_edp_dict : ", Psum_edp_dict, file = f)
	print("Psum_energy_dict : ", Psum_energy_dict, file = f)
	print("Psum_delay_dict : ", Psum_delay_dict, file = f)


	workbook = openpyxl.Workbook()
	sheet = workbook.get_sheet_by_name('Sheet') 
	# 写入标题
	column_tite = ["layer_id", "P_noc_edp", "P_P_d2d_e","P_P_dram_e","P_P_comm_d", "P_P_nop_edp", "P_K_d2d_e","P_K_dram_e","P_K_comm_d", "P_K_nop_edp", "P_PK_d2d_e","P_PK_dram_e","P_PK_comm_d", "P_PK_nop_edp"]
	column_tite += ["PK_noc_edp", "PK_P_d2d_e","PK_P_dram_e","PK_P_comm_d", "PK_P_nop_edp", "PK_K_d2d_e","PK_K_dram_e","PK_K_comm_d", "PK_K_nop_edp", "PK_PK_d2d_e","PK_PK_dram_e","PK_PK_comm_d", "PK_PK_nop_edp"]
	column_tite += ["K_noc_edp", "K_P_d2d_e","K_P_dram_e","K_P_comm_d", "K_P_nop_edp", "K_K_d2d_e","K_K_dram_e","K_K_comm_d", "K_K_nop_edp", "K_PK_d2d_e","K_PK_dram_e","K_PK_comm_d", "K_PK_nop_edp"]
	column_tite += ["energy_sum", "delay_sum", "EDP_inter_layer", "edp_sum", "par_index"]
	for col,column in enumerate(column_tite):
		sheet.cell(1, col+1, column)
	# 写入每一行
	for row, data in enumerate(excel_datas):
		for col, column_data in enumerate(data):
			sheet.cell(row+2, col+1, column_data)

	setname = str(network_param_2["P"]) + "_" + str(network_param_2["Q"]) + "_" + str(network_param_2["K"]) + "_" + str(network_param_2["C"]) + "_" + str(NoC_w) + "_" + str(NOC_NODE_NUM)

	workbook.save("./final_test/noc_nop_xls/"+struct+"_"+app_name+"_noc_nop.xls")

def nop_noc_test_ours_simba(app_name):

	EDP_sum_dict = {"K":0, "C":0, "KC":0}
	EDP_inter_dict = {"K":0, "C":0, "KC":0}
	energy_sum_dict = {"K":0, "C":0, "KC":0}
	delay_sum_dict = {"K":0, "C":0, "KC":0}
	energy_inter_dict = {"K":0, "C":0, "KC":0}
	delay_inter_dict = {"K":0, "C":0, "KC":0}
	inter_layer_edp_dict = {}
	delay_psum_dict = {"K":0, "C":0, "KC":0}
	energy_psum_dict = {"K":0, "C":0, "KC":0}
	edp_psum_dict = {"K":0, "C":0, "KC":0}
	parallel_type_dict = {"K":[],"C":[],"KC":[]}

	para_type = ["K","KC","C"]
	parallel_type_1 = {"K":{"K":16,"C":1}, "KC":{"K":4,"C":4}, "C":{"K":1,"C":16}}
	parallel_type_2 = {"K":{"K":16,"C":1}, "KC":{"K":4,"C":4}, "C":{"K":1,"C":16}}
	
	file_name = "./final_test/noc_nop/ours_simba_" + app_name + "_noc_nop.txt"
	f = open(file_name,'w')

	# --- 获得神经网络参数
	layer_dict, layer_id_list = getLayerParam(app_name)
	# --- 获得片内EDP数值
	EDP_intra_layer = getIntraLayerEDP_simba(app_name)
	energy_intra_layer = getIntraLayerEDP_simba(app_name, "intra_layer_energy")
	delay_intra_layer = getIntraLayerEDP_simba(app_name, "intra_layer_delay")
	para_each_layer = getIntraLayerEDP_simba(app_name, "intra_layer_parallel")

	excel_datas = []
	data_list = {"P":[],"PK":[],"K":[]}

	# 计算psum
	Psum_edp_dict = {}
	Psum_energy_dict = {}
	Psum_delay_dict = {}
	for i in layer_id_list:
		network_param = layer_dict[i]
		output_num = network_param["P"]*network_param["Q"] *network_param["K"]
		Psum_edp_dict[i] = {}
		Psum_energy_dict[i] = {}
		Psum_delay_dict[i] = {}
		for par_type in para_type:
			PC3 = para_each_layer[i][par_type]["C"]
			chiplet_num = para_each_layer[i][par_type]["C"] * para_each_layer[i][par_type]["K"]
			psum_delay, psum_energy_list = calPSumAllReduce(output_num, chiplet_num, PC3)
			Psum_edp_dict[i][par_type] = psum_delay * psum_energy_list[2] / freq_1G / PE_freq
			Psum_energy_dict[i][par_type] = psum_energy_list[2]
			Psum_delay_dict[i][par_type] = psum_delay

	# ---初始（最下面一层的）NoC edp + psum
	for para_type_2 in para_type:
		layer_id = layer_id_list[-1]
		EDP = EDP_intra_layer[layer_id][para_type_2]
		energy = energy_intra_layer[layer_id][para_type_2]
		delay = delay_intra_layer[layer_id][para_type_2]
		EDP_sum_dict[para_type_2] += EDP + Psum_edp_dict[layer_id][para_type_2]
		energy_sum_dict[para_type_2] += energy + Psum_energy_dict[layer_id][para_type_2]
		delay_sum_dict[para_type_2] += delay + Psum_delay_dict[layer_id][para_type_2]

		delay_psum_dict[para_type_2] = Psum_delay_dict[layer_id][para_type_2]
		energy_psum_dict[para_type_2] = Psum_energy_dict[layer_id][para_type_2]
		edp_psum_dict[para_type_2] = Psum_edp_dict[layer_id][para_type_2]

		line = "parallel " + str(para_type_2) + "  noc_EDP " + str(EDP) + "  psum_EDP " + str(Psum_edp_dict[layer_id][para_type_2]) + "  psum_energy " + str(Psum_energy_dict[layer_id][para_type_2]) + "  psum_delay " + str(Psum_delay_dict[layer_id][para_type_2])
		print(line, file = f)
		data_list[para_type_2] = [EDP]
	print("energy_sum_dict",energy_sum_dict)
	print("delay_sum_dict",delay_sum_dict)

	for ii in range(12):
			data_list["P"].append(0)
			data_list["PK"].append(0)
			data_list["K"].append(0)
	
	excel_per_layer = ["layer"+str(layer_id_list[-1])] + data_list["P"] +data_list["PK"] + data_list["K"] + [str(energy_sum_dict), str(delay_sum_dict), str(EDP_inter_dict), str(EDP_sum_dict), str(parallel_type_dict)]

	excel_datas.append(excel_per_layer)
	
	for i in reversed(layer_id_list):
		if i < layer_id_list[-1]:
			inter_layer_edp_dict[i] = {}
			network_param_2 = layer_dict[i+1]
			parallel_type_dict_new = {"K":[], "C":[], "KC":[]}
			edp_on_noc_dict = {"K":0, "C":0, "KC":0}
			EDP_sum_dict_new = {"K":0, "C":0, "KC":0}
			EDP_inter_dict_new = {"K":0, "C":0, "KC":0}
			EDP_NoC_dict = {"K":0, "C":0, "KC":0}
			energy_NoC_dict = {"K":0, "C":0, "KC":0}
			delay_NoC_dict = {"K":0, "C":0, "KC":0}
			energy_sum_dict_new = {"K":0, "C":0, "KC":0}
			delay_sum_dict_new = {"K":0, "C":0, "KC":0}
			energy_inter_dict_new = {"K":0, "C":0, "KC":0}
			delay_inter_dict_new = {"K":0, "C":0, "KC":0}

			energy_psum_dict_new = {"K":0, "C":0, "KC":0}
			delay_psum_dict_new = {"K":0, "C":0, "KC":0}
			edp_psum_dict_new = {"K":0, "C":0, "KC":0}
			energy_psum =  {"K":0, "C":0, "KC":0}
			delay_psum =  {"K":0, "C":0, "KC":0}
			edp_psum =  {"K":0, "C":0, "KC":0}

			line = "layer " + str(i) + "--------------------------------------"
			print(line, file = f)
			excel_per_layer = []
			excel_per_layer.append(i)

			for para_type_1 in para_type:
				EDP = EDP_intra_layer[i][para_type_1]
				EDP_NoC_dict[para_type_1] = EDP
				energy_NoC_dict[para_type_1] = energy_intra_layer[i][para_type_1]
				delay_NoC_dict[para_type_1] = delay_intra_layer[i][para_type_1]
				#EDP_sum_dict[para_type_1] += EDP
				data_list[para_type_1] = [EDP]
				energy_psum[para_type_1] = Psum_energy_dict[i][para_type_1]
				delay_psum[para_type_1] = Psum_delay_dict[i][para_type_1]
				edp_psum[para_type_1] = Psum_edp_dict[i][para_type_1]
				line = "parallel " + str(para_type_1) + "  noc_EDP " + str(EDP) + "  psum_EDP " + str(Psum_edp_dict[i][para_type_1]) + "  psum_energy " + str(Psum_energy_dict[i][para_type_1]) + "  psum_delay " + str(Psum_delay_dict[i][para_type_1])
				print(line, file = f)

			# 计算层间的EDP
			for para_type_1 in para_type:
				EDP_num = {}
				energy_num = {}
				delay_num = {}
				energy_inter_num = {}
				delay_inter_num = {}
				EDP_num_only = {}
				inter_layer_edp_dict[i][para_type_1] = {}
				par_C_pre = para_each_layer[i][para_type_1]["C"]
				par_K_pre = para_each_layer[i][para_type_1]["K"]
				for para_type_2 in para_type:
					par_C_cur = para_each_layer[i+1][para_type_2]["C"]
					par_K_cur = para_each_layer[i+1][para_type_2]["K"]

					worstCommFlitNum, d2d_e , dram_e, EDP = calInterComm_simba(par_C_pre, par_K_pre, par_C_cur, par_K_cur, network_param_2)
					
					EDP_num[para_type_2] = EDP + EDP_sum_dict[para_type_2]
					EDP_num_only[para_type_2] = EDP
					energy_num[para_type_2] = d2d_e + dram_e + energy_sum_dict[para_type_2]
					delay_num[para_type_2] = worstCommFlitNum + delay_sum_dict[para_type_2]
					energy_inter_num[para_type_2] = d2d_e + dram_e + energy_inter_dict[para_type_2]
					delay_inter_num[para_type_2] = worstCommFlitNum + delay_inter_dict[para_type_2]

					data_list[para_type_1].append(d2d_e)
					data_list[para_type_1].append(dram_e)
					data_list[para_type_1].append(worstCommFlitNum)
					data_list[para_type_1].append(EDP)

					inter_layer_edp_dict[i][para_type_1][para_type_2] = [EDP, d2d_e, dram_e, worstCommFlitNum]
				
				line = "layer " + str(i) + " : parallel " + str(para_type_1)
				print(line, file = f)

				EDP_num_min = EDP_num["K"]
				para_min = "K"
				for pa in EDP_num:
					if EDP_num[pa] < EDP_num_min:
						EDP_num_min = EDP_num[pa]
						para_min = pa
				
				EDP_sum_dict_new[para_type_1] = EDP_num_min + EDP_NoC_dict[para_type_1] + edp_psum[para_type_1]
				EDP_inter_dict_new[para_type_1] = EDP_inter_dict[para_min] + EDP_num_only[para_min]
				energy_sum_dict_new[para_type_1] = energy_num[para_min] + energy_NoC_dict[para_type_1] + energy_psum[para_type_1]
				delay_sum_dict_new[para_type_1] = delay_num[para_min] + delay_NoC_dict[para_type_1] + delay_psum[para_type_1]
				energy_inter_dict_new[para_type_1] = energy_inter_num[para_min]
				delay_inter_dict_new[para_type_1] = delay_inter_num[para_min]

				energy_psum_dict_new[para_type_1] = energy_psum_dict[para_min] +  energy_psum[para_type_1]
				delay_psum_dict_new[para_type_1] = delay_psum_dict[para_min] +  delay_psum[para_type_1]
				edp_psum_dict_new[para_type_1] = edp_psum_dict[para_min] +  edp_psum[para_type_1]

				parallel_type_dict_new[para_type_1] = [para_min]
				parallel_type_dict_new[para_type_1] += parallel_type_dict[para_min]
				
				line2 = "---------nop_EDP = " +str(EDP_num_only)
				print(line2, file = f)
				line2 = "---------Psum_EDP = " +str(edp_psum)
				print(line2, file = f)
				line2 = "---------EDP_num = " +str(EDP_num)
				print(line2, file = f)


			parallel_type_dict = copy.deepcopy(parallel_type_dict_new)
			EDP_sum_dict = copy.deepcopy(EDP_sum_dict_new)
			EDP_inter_dict = copy.deepcopy(EDP_inter_dict_new)
			energy_sum_dict = copy.deepcopy(energy_sum_dict_new)
			delay_sum_dict = copy.deepcopy(delay_sum_dict_new)
			energy_inter_dict = copy.deepcopy(energy_inter_dict_new)
			delay_inter_dict = copy.deepcopy(delay_inter_dict_new)
			edp_psum_dict = copy.deepcopy(edp_psum_dict_new)
			energy_psum_dict = copy.deepcopy(energy_psum_dict_new)
			delay_psum_dict = copy.deepcopy(delay_psum_dict_new)

			excel_per_layer = ["layer"+str(i)] + data_list["P"] +data_list["PK"] + data_list["K"] + [str(energy_sum_dict),str(delay_sum_dict), str(EDP_inter_dict), str(EDP_sum_dict), str(parallel_type_dict)]

			excel_datas.append(excel_per_layer)
			
		line3 = "----EDP_sum_dict = " + str(EDP_sum_dict)
		line1 = "----parallel_type_dict = " +str(parallel_type_dict)
		line4 = "----energy_sum_dict = " + str(energy_sum_dict)
		line5 = "----delay_sum_dict = " + str(delay_sum_dict)
		line6 = "----edp_inter_layer = " + str(EDP_inter_dict)
		line7 = "----energy_inter_layer = " + str(energy_inter_dict)
		line8 = "----delay_inter_layer = " + str(delay_inter_dict)
		line9 = "----edp_psum_dict = " + str(edp_psum_dict)
		line10 = "----energy_psum_dict = " + str(energy_psum_dict)
		line11 = "----delay_psum_dict = " + str(delay_psum_dict)
		print(line3, file = f)
		print(line1, file = f)
		print(line4, file = f)
		print(line5, file = f)
		print(line6, file = f)
		print(line7, file = f)
		print(line8, file = f)
		print(line9, file = f)
		print(line10, file = f)
		print(line11, file = f)

	
	# 计算全K, 全P, 全PK
	print(inter_layer_edp_dict)
	edp_K, edp_K_inter = oneParallelTest(inter_layer_edp_dict, EDP_intra_layer, "K")
	edp_C, edp_C_inter = oneParallelTest(inter_layer_edp_dict, EDP_intra_layer, "C")
	edp_KC, edp_KC_inter = oneParallelTest(inter_layer_edp_dict, EDP_intra_layer, "KC")
	lineK = "EDP_ALL_K = " + str(edp_K) + "  ;  K_inter_layer = " + str(edp_K_inter)
	lineC = "EDP_ALL_C = " + str(edp_C) + "  ;  C_inter_layer = " + str(edp_C_inter)
	lineKC = "EDP_ALL_KC = " + str(edp_KC) + "  ;  KC_inter_layer = " + str(edp_KC_inter)
	print(lineK, file = f)
	print(lineC, file = f)
	print(lineKC, file = f)
	print("Psum_edp_dict : ", Psum_edp_dict, file = f)
	print("Psum_energy_dict : ", Psum_energy_dict, file = f)
	print("Psum_delay_dict : ", Psum_delay_dict, file = f)


	workbook = openpyxl.Workbook()
	sheet = workbook.get_sheet_by_name('Sheet') 
	# 写入标题
	column_tite = ["layer_id", "P_noc_edp", "P_P_d2d_e","P_P_dram_e","P_P_comm_d", "P_P_nop_edp", "P_K_d2d_e","P_K_dram_e","P_K_comm_d", "P_K_nop_edp", "P_PK_d2d_e","P_PK_dram_e","P_PK_comm_d", "P_PK_nop_edp"]
	column_tite += ["PK_noc_edp", "PK_P_d2d_e","PK_P_dram_e","PK_P_comm_d", "PK_P_nop_edp", "PK_K_d2d_e","PK_K_dram_e","PK_K_comm_d", "PK_K_nop_edp", "PK_PK_d2d_e","PK_PK_dram_e","PK_PK_comm_d", "PK_PK_nop_edp"]
	column_tite += ["K_noc_edp", "K_P_d2d_e","K_P_dram_e","K_P_comm_d", "K_P_nop_edp", "K_K_d2d_e","K_K_dram_e","K_K_comm_d", "K_K_nop_edp", "K_PK_d2d_e","K_PK_dram_e","K_PK_comm_d", "K_PK_nop_edp"]
	column_tite += ["energy_sum", "delay_sum", "EDP_inter_layer", "edp_sum", "par_index"]
	for col,column in enumerate(column_tite):
		sheet.cell(1, col+1, column)
	# 写入每一行
	for row, data in enumerate(excel_datas):
		for col, column_data in enumerate(data):
			sheet.cell(row+2, col+1, column_data)

	workbook.save("./final_test/noc_nop_xls/simba_"+app_name+"_noc_nop.xls")

def nop_noc_test_nnbaton(app_name):
	NOC_NODE_NUM = 8
	NoC_w = 2

	EDP_sum_dict = {"P":0, "K":0, "PK_1":0, "PK_2":0}
	EDP_inter_dict = {"P":0, "K":0, "PK_1":0, "PK_2":0}
	energy_sum_dict = {"P":0, "K":0, "PK_1":0, "PK_2":0}
	delay_sum_dict = {"P":0, "K":0, "PK_1":0, "PK_2":0}
	inter_layer_edp_dict = {}
	energy_inter_dict = {"P":0, "K":0, "PK_1":0, "PK_2":0}
	delay_inter_dict = {"P":0, "K":0, "PK_1":0, "PK_2":0}
	parallel_type_dict = {"P":[],"K":[],"PK_1":[], "PK_2":[]}

	para_type = ["P","K","PK_1", "PK_2"]
	parallel_type_1 = {"P":{"P":8,"Q":1,"K":1},'K':{"P":1,"Q":1,"K":8},'PK_1':{"P":4,"Q":1,"K":2},'PK_2':{"P":2,"Q":1,"K":4}}
	parallel_type_2 = {"P":{"P":8,"Q":1,"K":1},'K':{"P":1,"Q":1,"K":8},'PK_1':{"P":4,"Q":1,"K":2},'PK_2':{"P":2,"Q":1,"K":4}}

	file_name = "./final_test/noc_nop/nnbaton_" + app_name + "_noc_nop.txt"
	f = open(file_name,'w')

	# --- 获得神经网络参数
	layer_dict, layer_id_list = getLayerParam(app_name)
	# --- 获得片内EDP数值
	EDP_intra_layer = getIntraLayerEDP_nnbaton(app_name, struct_type="nnbaton_")
	energy_intra_layer = getIntraLayerEDP_nnbaton(app_name, "intra_layer_energy",struct_type="nnbaton_")
	delay_intra_layer = getIntraLayerEDP_nnbaton(app_name, "intra_layer_delay",struct_type="nnbaton_")

	excel_datas = []
	data_list = {"P":[],"PK_1":[],"PK_2":[],"K":[]}

	# ---初始（最下面一层的）NoC edp
	for para_type_2 in para_type:
		layer_id = layer_id_list[-1]
		EDP = EDP_intra_layer[layer_id][para_type_2]
		energy = energy_intra_layer[layer_id][para_type_2]
		delay = delay_intra_layer[layer_id][para_type_2]
		EDP_sum_dict[para_type_2] += EDP
		energy_sum_dict[para_type_2] += energy
		delay_sum_dict[para_type_2] += delay
		line = "parallel " + str(para_type_2) + "  noc_EDP " + str(EDP)
		print(line, file = f)
		data_list[para_type_2] = [EDP]
	print("energy_sum_dict",energy_sum_dict)
	print("delay_sum_dict",delay_sum_dict)

	for ii in range(16):
			data_list["P"].append(0)
			data_list["PK_1"].append(0)
			data_list["PK_2"].append(0)
			data_list["K"].append(0)
	
	excel_per_layer = ["layer"+str(layer_id_list[-1])] + data_list["P"] +data_list["PK_1"] + data_list["PK_2"] + data_list["K"] + [str(energy_sum_dict), str(delay_sum_dict), str(EDP_inter_dict), str(EDP_sum_dict), str(parallel_type_dict)]

	excel_datas.append(excel_per_layer)
	

	for i in reversed(layer_id_list):
		if i < layer_id_list[-1]:
			inter_layer_edp_dict[i] = {}
			network_param_2 = layer_dict[i+1]
			parallel_type_dict_new = {"P":[], "K":[], "PK_1":[], "PK_2":[]}
			EDP_sum_dict_new = {"P":0, "K":0, "PK_1":0, "PK_2":0}
			EDP_inter_dict_new = {"P":0, "K":0, "PK_1":0, "PK_2":0}
			EDP_NoC_dict = {"P":0, "K":0, "PK_1":0, "PK_2":0}
			energy_NoC_dict = {"P":0, "K":0, "PK_1":0, "PK_2":0}
			delay_NoC_dict = {"P":0, "K":0, "PK_1":0, "PK_2":0}
			energy_sum_dict_new = {"P":0, "K":0, "PK_1":0, "PK_2":0}
			delay_sum_dict_new = {"P":0, "K":0, "PK_1":0, "PK_2":0}
			energy_inter_dict_new = {"P":0, "K":0, "PK_1":0, "PK_2":0}
			delay_inter_dict_new = {"P":0, "K":0, "PK_1":0, "PK_2":0}

			line = "layer " + str(i) + "--------------------------------------"
			print(line, file = f)
			excel_per_layer = []
			excel_per_layer.append(i)

			for para_type_1 in para_type:
				EDP = EDP_intra_layer[i][para_type_1]
				EDP_NoC_dict[para_type_1] = EDP
				energy_NoC_dict[para_type_1] = energy_intra_layer[i][para_type_1]
				delay_NoC_dict[para_type_1] = delay_intra_layer[i][para_type_1]
				#EDP_sum_dict[para_type_1] += EDP
				line = "parallel " + str(para_type_1) + "  noc_EDP " + str(EDP)
				print(line, file = f)
				data_list[para_type_1] = [EDP]

			# 计算层间的EDP
			for para_type_1 in para_type:
				EDP_num = {}
				energy_num = {}
				delay_num = {}
				energy_inter_num = {}
				delay_inter_num = {}
				EDP_num_only = {}
				inter_layer_edp_dict[i][para_type_1] = {}
				for para_type_2 in para_type:
					parallel_1 = parallel_type_1[para_type_1]
					parallel_2 = parallel_type_2[para_type_2]
					worstCommFlitNum, d2d_e , dram_e, EDP = getInterLayer(network_param_2, parallel_1, parallel_2, NOC_NODE_NUM, NoC_w, topology="ring")
					EDP_num[para_type_2] = EDP + EDP_sum_dict[para_type_2]
					EDP_num_only[para_type_2] = EDP
					energy_num[para_type_2] = d2d_e + dram_e + energy_sum_dict[para_type_2]
					delay_num[para_type_2] = worstCommFlitNum + delay_sum_dict[para_type_2]
					energy_inter_num[para_type_2] = d2d_e + dram_e + energy_inter_dict[para_type_2]
					delay_inter_num[para_type_2] = worstCommFlitNum + delay_inter_dict[para_type_2]


					data_list[para_type_1].append(d2d_e)
					data_list[para_type_1].append(dram_e)
					data_list[para_type_1].append(worstCommFlitNum)
					data_list[para_type_1].append(EDP)

					inter_layer_edp_dict[i][para_type_1][para_type_2] = [EDP, d2d_e, dram_e, worstCommFlitNum]
				
				line = "layer " + str(i) + " : parallel " + str(para_type_1)
				print(line, file = f)

				EDP_num_min = EDP_num["P"]
				para_min = "P"
				for pa in EDP_num:
					if EDP_num[pa] < EDP_num_min:
						EDP_num_min = EDP_num[pa]
						para_min = pa
				
				EDP_sum_dict_new[para_type_1] = EDP_num_min + EDP_NoC_dict[para_type_1]
				EDP_inter_dict_new[para_type_1] = EDP_inter_dict[para_min] + EDP_num_only[para_min]
				energy_sum_dict_new[para_type_1] = energy_num[para_min] + energy_NoC_dict[para_type_1]
				delay_sum_dict_new[para_type_1] = delay_num[para_min] + delay_NoC_dict[para_type_1]
				energy_inter_dict_new[para_type_1] = energy_inter_num[para_min]
				delay_inter_dict_new[para_type_1] = delay_inter_num[para_min]

				parallel_type_dict_new[para_type_1] = [para_min]
				parallel_type_dict_new[para_type_1] += parallel_type_dict[para_min]
				
				line2 = "---------nop_EDP = " +str(EDP_num_only)
				print(line2, file = f)
				line2 = "---------EDP_num = " +str(EDP_num)
				print(line2, file = f)

			parallel_type_dict = copy.deepcopy(parallel_type_dict_new)
			EDP_sum_dict = copy.deepcopy(EDP_sum_dict_new)
			EDP_inter_dict = copy.deepcopy(EDP_inter_dict_new)
			energy_sum_dict = copy.deepcopy(energy_sum_dict_new)
			delay_sum_dict = copy.deepcopy(delay_sum_dict_new)
			energy_inter_dict = copy.deepcopy(energy_inter_dict_new)
			delay_inter_dict = copy.deepcopy(delay_inter_dict_new)

			excel_per_layer = ["layer"+str(i)] + data_list["P"] +data_list["PK_1"] +data_list["PK_2"] + data_list["K"] + [str(energy_sum_dict),str(delay_sum_dict), str(EDP_inter_dict), str(EDP_sum_dict), str(parallel_type_dict)]

			excel_datas.append(excel_per_layer)
			
		line3 = "----EDP_sum_dict = " + str(EDP_sum_dict)
		line1 = "----parallel_type_dict = " +str(parallel_type_dict)
		line4 = "----energy_sum_dict = " + str(energy_sum_dict)
		line5 = "----delay_sum_dict = " + str(delay_sum_dict)
		line6 = "----edp_inter_layer = " + str(EDP_inter_dict)
		line7 = "----energy_inter_layer = " + str(energy_inter_dict)
		line8 = "----delay_inter_layer = " + str(delay_inter_dict)
		print(line3, file = f)
		print(line1, file = f)
		print(line4, file = f)
		print(line5, file = f)
		print(line6, file = f)
		print(line7, file = f)
		print(line8, file = f)
	
	# 计算全K, 全P, 全PK
	print(inter_layer_edp_dict)
	edp_K, edp_K_inter = oneParallelTest(inter_layer_edp_dict, EDP_intra_layer, "K")
	edp_P, edp_P_inter = oneParallelTest(inter_layer_edp_dict, EDP_intra_layer, "P")
	edp_PK_1, edp_PK_1_inter = oneParallelTest(inter_layer_edp_dict, EDP_intra_layer, "PK_1")
	edp_PK_2, edp_PK_2_inter = oneParallelTest(inter_layer_edp_dict, EDP_intra_layer, "PK_2")
	lineK = "EDP_ALL_K = " + str(edp_K) + "  ;  K_inter_layer = " + str(edp_K_inter)
	lineP = "EDP_ALL_P = " + str(edp_P) + "  ;  P_inter_layer = " + str(edp_P_inter)
	linePK_1 = "EDP_ALL_PK = " + str(edp_PK_1) + "  ;  PK_inter_layer = " + str(edp_PK_1_inter)
	linePK_2 = "EDP_ALL_PK = " + str(edp_PK_2) + "  ;  PK_inter_layer = " + str(edp_PK_2_inter)
	print(lineP, file = f)
	print(lineK, file = f)
	print(linePK_1, file = f)
	print(linePK_2, file = f)


	workbook = openpyxl.Workbook()
	sheet = workbook.get_sheet_by_name('Sheet') 
	# 写入标题
	column_tite = ["layer_id", "P_noc_edp", "P_P_d2d_e","P_P_dram_e","P_P_comm_d", "P_P_nop_edp", "P_K_d2d_e","P_K_dram_e","P_K_comm_d", "P_K_nop_edp", "P_PK_1_d2d_e","P_PK_1_dram_e","P_PK_1_comm_d", "P_PK_1_nop_edp", "P_PK_2_d2d_e","P_PK_2_dram_e","P_PK_2_comm_d", "P_PK_2_nop_edp"]
	column_tite += ["PK_1_noc_edp", "PK_P_d2d_e","PK_P_dram_e","PK_P_comm_d", "PK_P_nop_edp", "PK_K_d2d_e","PK_K_dram_e","PK_K_comm_d", "PK_K_nop_edp", "PK_PK_1_d2d_e","PK_PK_1_dram_e","PK_PK_1_comm_d", "PK_PK_1_nop_edp","PK_PK_2_d2d_e","PK_PK_2_dram_e","PK_PK_2_comm_d", "PK_PK_2_nop_edp"]
	column_tite += ["PK_2_noc_edp", "PK_P_d2d_e","PK_P_dram_e","PK_P_comm_d", "PK_P_nop_edp", "PK_K_d2d_e","PK_K_dram_e","PK_K_comm_d", "PK_K_nop_edp", "PK_PK_1_d2d_e","PK_PK_1_dram_e","PK_PK_1_comm_d", "PK_PK_1_nop_edp","PK_PK_2_d2d_e","PK_PK_2_dram_e","PK_PK_2_comm_d", "PK_PK_2_nop_edp"]
	column_tite += ["K_noc_edp", "K_P_d2d_e","K_P_dram_e","K_P_comm_d", "K_P_nop_edp", "K_K_d2d_e","K_K_dram_e","K_K_comm_d", "K_K_nop_edp", "K_PK_1_d2d_e","K_PK_1_dram_e","K_PK_1_comm_d", "K_PK_1_nop_edp", "K_PK_2_d2d_e","K_PK_2_dram_e","K_PK_2_comm_d", "K_PK_2_nop_edp"]
	column_tite += ["energy_sum", "delay_sum", "EDP_inter_layer", "edp_sum", "par_index"]
	for col,column in enumerate(column_tite):
		sheet.cell(1, col+1, column)
	# 写入每一行
	for row, data in enumerate(excel_datas):
		for col, column_data in enumerate(data):
			sheet.cell(row+2, col+1, column_data)

	setname = str(network_param_2["P"]) + "_" + str(network_param_2["Q"]) + "_" + str(network_param_2["K"]) + "_" + str(network_param_2["C"]) + "_" + str(NoC_w) + "_" + str(NOC_NODE_NUM)

	workbook.save("./final_test/noc_nop_xls/nnbaton_"+app_name+"_noc_nop.xls")

	print(EDP_intra_layer)

def oneParallelTest(inter_layer_edp_dict, intra_layer_edp_dict, para):
	edp_sum = 0
	edp_inter = 0
	for i in inter_layer_edp_dict:
		edp_inter += inter_layer_edp_dict[i][para][para][0]
	edp_sum += edp_inter
	for i in intra_layer_edp_dict:
		edp_sum += intra_layer_edp_dict[i][para]
	return edp_sum, edp_inter
		
def noc_test_ours(app_name):
	NOC_NODE_NUM = 16
	NoC_w = 4

	EDP_sum_dict = {"P":0, "K":0, "PK":0}

	para_type = ["P","K","PK"]
	parallel_type = {"P":{"P":16,"Q":1,"K":1},'K':{"P":1,"Q":1,"K":16},'PK':{"P":4,"Q":1,"K":4}}
	
	file_name = "./final_test/noc_nop/" + app_name + "_noc.txt"
	f = open(file_name,'w')

	# --- 获得神经网络参数
	layer_dict, layer_id_list = getLayerParam(app_name)
	# --- 获得片内EDP数值
	EDP_intra_layer = getIntraLayerEDP(app_name)
	energy_intra_layer = getIntraLayerEDP(app_name, "intra_layer_energy")
	delay_intra_layer = getIntraLayerEDP(app_name, "intra_layer_delay")


	para_list = []

	edp_sum = 0
	energy_sum = 0
	delay_sum = 0

	inter_edp = 0
	inter_energy = 0
	inter_delay = 0
	layer_num = 0

	for i in layer_id_list:
		line = "layer " + str(i) + "-----------------"
		print(line, file = f)
		EDP_min = 0
		for para_type_2 in para_type:
			EDP = EDP_intra_layer[i][para_type_2]
			EDP_sum_dict[para_type_2] += EDP
			line = "parallel " + str(para_type_2) + "  EDP " + str(EDP)
			print(line, file = f)
			if EDP_min == 0 or EDP_min > EDP:
				EDP_min = EDP
				EDP_min_id = para_type_2
	
		line = "layer " + str(i) + " :  EDP_min_id = " + str(EDP_min_id) + "  EDP_Min =  " + str(EDP_min)
		print(line, file = f)
		para_list.append(EDP_min_id)
		edp_sum += EDP_min
		energy_sum += energy_intra_layer[i][EDP_min_id]
		delay_sum += delay_intra_layer[i][EDP_min_id]

		if i > layer_id_list[0]:
			network_param_2 = layer_dict[i]
			print(network_param_2)
			worstCommFlitNum, d2d_e , dram_e, EDP = getInterLayer(network_param_2, parallel_type[para_list[layer_num - 1]], parallel_type[para_list[layer_num]], NOC_NODE_NUM, NoC_w)
			edp_sum += EDP
			inter_edp += EDP
			inter_energy += d2d_e + dram_e
			energy_sum += d2d_e + dram_e
			delay_sum += worstCommFlitNum
			inter_delay += worstCommFlitNum

			line = "layer " + str(i-1) + " to " + str(i) + " : nop_EDP =  " + str(EDP)
			print(line, file = f)
		layer_num += 1
		
	line1 = "---- EDP ---- "
	line1_1 = str(inter_edp) + "\t" + str(edp_sum) + "\t"
	print(line1, file = f)
	print(line1_1, file = f)

	line2 = "---- energy ---- "
	line2_1 = str(inter_energy) + "\t" + str(energy_sum) + "\t"
	print(line2, file = f)
	print(line2_1, file = f)

	line3 = "---- delay ---- "
	line3_1 = str(inter_delay) + "\t" + str(delay_sum) + "\t"
	print(line3, file = f)
	print(line3_1, file = f)

	line4 = "----parallel_type_dict = " +str(para_list)
	print(line4, file = f)
	
def noc_test_ours_on_simba(app_name):
	NOC_NODE_NUM = 64
	NoC_w = 8

	EDP_sum_dict = {"P":0, "K":0, "PK":0}
	
	para_type = ["P","K","PK"]
	parallel_type = {"P":{"P":64,"Q":1,"K":1},'K':{"P":1,"Q":1,"K":64},'PK':{"P":8,"Q":1,"K":8}}
	
	file_name = "./final_test/simba_noc/ours_simba_" + app_name + "_noc.txt"
	f = open(file_name,'w')

	# --- 获得神经网络参数
	layer_dict, layer_id_list = getLayerParam(app_name)
	# --- 获得片内EDP数值
	EDP_intra_layer = getIntraLayerEDP(app_name, struct_type="simba_ours_")
	energy_intra_layer = getIntraLayerEDP(app_name, "intra_layer_energy", struct_type="simba_ours_")
	delay_intra_layer = getIntraLayerEDP(app_name, "intra_layer_delay", struct_type="simba_ours_")

	excel_datas = []
	data_list = {"P":[],"PK":[],"K":[]}

	para_list = []
	edp_sum = 0
	energy_sum = 0
	delay_sum = 0

	inter_edp = 0
	inter_energy = 0
	inter_delay = 0
	layer_num = 0

	for i in layer_id_list:
		line = "layer " + str(i) + "-----------------"
		print(line, file = f)
		EDP_min = 0
		for para_type_2 in para_type:
			EDP = EDP_intra_layer[i][para_type_2]
			energy = energy_intra_layer[i][para_type_2]
			delay = delay_intra_layer[i][para_type_2]

			EDP_sum_dict[para_type_2] += EDP

			line = "parallel " + str(para_type_2) + "  EDP " + str(EDP) + "  energy " + str(energy) + "  delay " + str(delay)
			print(line, file = f)
			if EDP_min == 0 or EDP_min > EDP:
				EDP_min = EDP
				EDP_min_id = para_type_2
	
		line = "layer " + str(i) + " :  EDP_min_id = " + str(EDP_min_id) + "  EDP_Min =  " + str(EDP_min)
		print(line, file = f)
		para_list.append(EDP_min_id)
		edp_sum += EDP_min
		energy_sum += energy_intra_layer[i][EDP_min_id]
		delay_sum += delay_intra_layer[i][EDP_min_id]

		if i > layer_id_list[0]:
			network_param_2 = layer_dict[i]
			print(network_param_2)
			worstCommFlitNum, d2d_e , dram_e, EDP = getInterLayer(network_param_2, parallel_type[para_list[layer_num - 1]], parallel_type[para_list[layer_num]], NOC_NODE_NUM, NoC_w)
			edp_sum += EDP
			inter_edp += EDP
			inter_energy += d2d_e + dram_e
			energy_sum += d2d_e + dram_e
			delay_sum += worstCommFlitNum
			inter_delay += worstCommFlitNum

			line = "layer " + str(i-1) + " to " + str(i) + " : nop_EDP =  " + str(EDP) + " nop_energy=" + str(d2d_e + dram_e) + " nop_delay=" + str(worstCommFlitNum)
			print(line, file = f)
		layer_num += 1
		
	line1 = "---- EDP ---- "
	line1_1 = str(inter_edp) + "\t" + str(edp_sum) + "\t"
	print(line1, file = f)
	print(line1_1, file = f)

	line2 = "---- energy ---- "
	line2_1 = str(inter_energy) + "\t" + str(energy_sum) + "\t"
	print(line2, file = f)
	print(line2_1, file = f)

	line3 = "---- delay ---- "
	line3_1 = str(inter_delay) + "\t" + str(delay_sum) + "\t"
	print(line3, file = f)
	print(line3_1, file = f)

	line4 = "----parallel_type_dict = " +str(para_list)
	print(line4, file = f)

def noc_test_nnbaton(app_name):
	NOC_NODE_NUM = 8
	NoC_w = 2

	EDP_sum_dict = {"P":0, "K":0, "PK_1":0, "PK_2":0}

	para_type = ["P","K","PK_1", "PK_2"]
	parallel_type = {"P":{"P":8,"Q":1,"K":1},'K':{"P":1,"Q":1,"K":8},'PK_1':{"P":4,"Q":1,"K":2},'PK_2':{"P":2,"Q":1,"K":4}}
	
	file_name = "./final_test/nnbaton_noc/mem_nnbaton_" + app_name + "_noc.txt"
	f = open(file_name,'w')

		# --- 获得神经网络参数
	layer_dict, layer_id_list = getLayerParam(app_name)
	# --- 获得片内EDP数值
	EDP_intra_layer = getIntraLayerEDP_nnbaton(app_name, struct_type="mem_nnbaton_")
	energy_intra_layer = getIntraLayerEDP_nnbaton(app_name, "intra_layer_energy",struct_type="mem_nnbaton_")
	delay_intra_layer = getIntraLayerEDP_nnbaton(app_name, "intra_layer_delay",struct_type="mem_nnbaton_")

	excel_datas = []
	data_list = {"P":[],"PK_1":[],"PK_2":[],"K":[]}

	para_list = []
	edp_sum = 0
	energy_sum = 0
	delay_sum = 0

	inter_edp = 0
	inter_energy = 0
	inter_delay = 0
	layer_num = 0

	for i in layer_id_list:
		line = "layer " + str(i) + "-----------------"
		print(line, file = f)
		EDP_min = 0
		for para_type_2 in para_type:
			EDP = EDP_intra_layer[i][para_type_2]
			energy = energy_intra_layer[i][para_type_2]
			delay = delay_intra_layer[i][para_type_2]

			EDP_sum_dict[para_type_2] += EDP

			line = "parallel " + str(para_type_2) + "  EDP " + str(EDP) + "  energy " + str(energy) + "  delay " + str(delay)
			print(line, file = f)
			if EDP_min == 0 or EDP_min > EDP:
				EDP_min = EDP
				EDP_min_id = para_type_2
	
		line = "layer " + str(i) + " :  EDP_min_id = " + str(EDP_min_id) + "  EDP_Min =  " + str(EDP_min)
		print(line, file = f)
		para_list.append(EDP_min_id)
		edp_sum += EDP_min
		energy_sum += energy_intra_layer[i][EDP_min_id]
		delay_sum += delay_intra_layer[i][EDP_min_id]

		if i > layer_id_list[0]:
			network_param_2 = layer_dict[i]
			print(network_param_2)
			worstCommFlitNum, d2d_e , dram_e, EDP = getInterLayer(network_param_2, parallel_type[para_list[layer_num - 1]], parallel_type[para_list[layer_num]], NOC_NODE_NUM, NoC_w, topology="ring")
			edp_sum += EDP
			inter_edp += EDP
			inter_energy += d2d_e + dram_e
			energy_sum += d2d_e + dram_e
			delay_sum += worstCommFlitNum
			inter_delay += worstCommFlitNum

			line = "layer " + str(i-1) + " to " + str(i) + " : nop_EDP =  " + str(EDP) + " nop_energy=" + str(d2d_e + dram_e) + " nop_delay=" + str(worstCommFlitNum)
			print(line, file = f)
		layer_num += 1
		
	line1 = "---- EDP ---- "
	line1_1 = str(inter_edp) + "\t" + str(edp_sum) + "\t"
	print(line1, file = f)
	print(line1_1, file = f)

	line2 = "---- energy ---- "
	line2_1 = str(inter_energy) + "\t" + str(energy_sum) + "\t"
	print(line2, file = f)
	print(line2_1, file = f)

	line3 = "---- delay ---- "
	line3_1 = str(inter_delay) + "\t" + str(delay_sum) + "\t"
	print(line3, file = f)
	print(line3_1, file = f)

	line4 = "----parallel_type_dict = " +str(para_list)
	print(line4, file = f)

def noc_test_simba(app_name, struct):
	EDP_sum_dict = {"K":0, "C":0, "KC":0}

	para_type = ["K","KC","C"]
	
	file_name = "./final_test/simba_noc/" + struct +"_"+ app_name + "_noc.txt"
	f = open(file_name,'w')

	# --- 获得神经网络参数
	layer_dict, layer_id_list = getLayerParam(app_name)
	# --- 获得片内EDP数值
	EDP_intra_layer = getIntraLayerEDP_simba(app_name, struct_type = struct)
	energy_intra_layer =  getIntraLayerEDP_simba(app_name, "intra_layer_energy", struct_type = struct)
	delay_intra_layer = getIntraLayerEDP_simba(app_name, "intra_layer_delay", struct_type = struct)
	para_each_layer = getIntraLayerEDP_simba(app_name, "intra_layer_parallel", struct_type = struct)

	excel_datas = []
	data_list = {"P":[],"PK":[],"K":[]}

	# 计算psum
	Psum_edp_dict = {}
	Psum_energy_dict = {}
	Psum_delay_dict = {}
	for i in layer_id_list:
		network_param = layer_dict[i]
		output_num = network_param["P"]*network_param["Q"] *network_param["K"]
		Psum_edp_dict[i] = {}
		Psum_energy_dict[i] = {}
		Psum_delay_dict[i] = {}
		for par_type in para_type:
			PC3 = para_each_layer[i][par_type]["C"]
			chiplet_num = para_each_layer[i][par_type]["C"] * para_each_layer[i][par_type]["K"]
			psum_delay, psum_energy_list = calPSumAllReduce(output_num, chiplet_num, PC3)
			Psum_edp_dict[i][par_type] = psum_delay * psum_energy_list[2] / freq_1G / PE_freq
			Psum_energy_dict[i][par_type] = psum_energy_list[2]
			Psum_delay_dict[i][par_type] = psum_delay

	para_list = []
	edp_sum = 0
	energy_sum = 0
	delay_sum = 0

	psum_edp_sum = 0
	psum_energy_sum = 0
	psum_delay_sum = 0

	inter_edp = 0
	inter_energy = 0
	inter_delay = 0
	layer_num = 0

	for i in layer_id_list:
		line = "layer " + str(i) + "-----------------"
		print(line, file = f)
		EDP_min = 0
		for para_type_2 in para_type:
			EDP = EDP_intra_layer[i][para_type_2]
			energy = energy_intra_layer[i][para_type_2]
			delay = delay_intra_layer[i][para_type_2]
			psum_edp = Psum_edp_dict[i][para_type_2]
			psum_energy = Psum_energy_dict[i][para_type_2]
			psum_delay = Psum_delay_dict[i][para_type_2]

			EDP_sum_dict[para_type_2] += EDP + psum_edp

			line = "parallel " + str(para_type_2) + "  EDP " + str(EDP+psum_edp) + "  energy " + str(energy+psum_energy) + "  delay " + str(delay+psum_delay)
			print(line, file = f)
			if EDP_min == 0 or EDP_min > (EDP + psum_edp):
				EDP_min = EDP + psum_edp
				EDP_min_id = para_type_2
	
		line = "layer " + str(i) + " :  EDP_min_id = " + str(EDP_min_id) + "  EDP_Min =  " + str(EDP_min)
		print(line, file = f)
		para_list.append(EDP_min_id)
		edp_sum += EDP_min
		energy_sum += energy_intra_layer[i][EDP_min_id] + Psum_energy_dict[i][EDP_min_id]
		delay_sum += delay_intra_layer[i][EDP_min_id] + Psum_delay_dict[i][EDP_min_id]
		psum_edp_sum += Psum_edp_dict[i][EDP_min_id]
		psum_energy_sum += Psum_energy_dict[i][EDP_min_id]
		psum_delay_sum += Psum_delay_dict[i][EDP_min_id]

		if i > layer_id_list[0]:
			network_param_2 = layer_dict[i]
			print(network_param_2)
			par_C_pre = para_each_layer[i-1][para_list[layer_num - 1]]["C"]
			par_K_pre = para_each_layer[i-1][para_list[layer_num - 1]]["K"]

			par_C_cur = para_each_layer[i][para_list[layer_num]]["C"]
			par_K_cur = para_each_layer[i][para_list[layer_num]]["K"]

			worstCommFlitNum, d2d_e , dram_e, EDP = calInterComm_simba(par_C_pre, par_K_pre, par_C_cur, par_K_cur, network_param_2)

			edp_sum += EDP
			inter_edp += EDP
			inter_energy += d2d_e + dram_e
			energy_sum += d2d_e + dram_e
			delay_sum += worstCommFlitNum
			inter_delay += worstCommFlitNum

			line = "layer " + str(i-1) + " to " + str(i) + " : nop_EDP =  " + str(EDP) + " nop_energy=" + str(d2d_e + dram_e) + " nop_delay=" + str(worstCommFlitNum)
			print(line, file = f)
		layer_num += 1
		
	line1 = "---- EDP ---- "
	line1_1 = str(psum_edp_sum) + "\t" + str(inter_edp) + "\t" + str(edp_sum) + "\t"
	print(line1, file = f)
	print(line1_1, file = f)

	line2 = "---- energy ---- "
	line2_1 = str(psum_energy_sum) + "\t" + str(inter_energy) + "\t" + str(energy_sum) + "\t"
	print(line2, file = f)
	print(line2_1, file = f)

	line3 = "---- delay ---- "
	line3_1 = str(psum_delay_sum) + "\t" + str(inter_delay) + "\t" + str(delay_sum) + "\t"
	print(line3, file = f)
	print(line3_1, file = f)

	line4 = "----parallel_type_dict = " +str(para_list)
	print(line4, file = f)

def noc_test_simba_on_ours(app_name, struct):
	EDP_sum_dict = {"K":0, "C":0, "KC":0}

	para_type = ["K","KC","C"]
	
	file_name = "./final_test/simba_noc/" + struct +"_"+ app_name + "_noc.txt"
	f = open(file_name,'w')

	# --- 获得神经网络参数
	layer_dict, layer_id_list = getLayerParam(app_name)
	# --- 获得片内EDP数值
	EDP_intra_layer = getIntraLayerEDP_simba(app_name, struct_type = struct)
	energy_intra_layer =  getIntraLayerEDP_simba(app_name, "intra_layer_energy", struct_type = struct)
	delay_intra_layer = getIntraLayerEDP_simba(app_name, "intra_layer_delay", struct_type = struct)
	para_each_layer = getIntraLayerEDP_simba(app_name, "intra_layer_parallel", struct_type = struct)

	excel_datas = []
	data_list = {"P":[],"PK":[],"K":[]}

	# 计算psum
	Psum_edp_dict = {}
	Psum_energy_dict = {}
	Psum_delay_dict = {}
	for i in layer_id_list:
		network_param = layer_dict[i]
		output_num = network_param["P"]*network_param["Q"] *network_param["K"]
		Psum_edp_dict[i] = {}
		Psum_energy_dict[i] = {}
		Psum_delay_dict[i] = {}
		for par_type in para_type:
			PC3 = para_each_layer[i][par_type]["C"]
			chiplet_num = para_each_layer[i][par_type]["C"] * para_each_layer[i][par_type]["K"]
			psum_delay, psum_energy_list = calPSumAllReduce(output_num, chiplet_num, PC3)
			Psum_edp_dict[i][par_type] = psum_delay * psum_energy_list[2] / freq_1G / PE_freq
			Psum_energy_dict[i][par_type] = psum_energy_list[2]
			Psum_delay_dict[i][par_type] = psum_delay

	para_list = []
	edp_sum = 0
	energy_sum = 0
	delay_sum = 0

	psum_edp_sum = 0
	psum_energy_sum = 0
	psum_delay_sum = 0

	inter_edp = 0
	inter_energy = 0
	inter_delay = 0
	layer_num = 0

	for i in layer_id_list:
		line = "layer " + str(i) + "-----------------"
		print(line, file = f)
		EDP_min = 0
		for para_type_2 in para_type:
			EDP = EDP_intra_layer[i][para_type_2]
			energy = energy_intra_layer[i][para_type_2]
			delay = delay_intra_layer[i][para_type_2]
			psum_edp = Psum_edp_dict[i][para_type_2]
			psum_energy = Psum_energy_dict[i][para_type_2]
			psum_delay = Psum_delay_dict[i][para_type_2]

			EDP_sum_dict[para_type_2] += EDP + psum_edp

			line = "parallel " + str(para_type_2) + "  EDP " + str(EDP+psum_edp) + "  energy " + str(energy+psum_energy) + "  delay " + str(delay+psum_delay)
			print(line, file = f)
			if EDP_min == 0 or EDP_min > (EDP + psum_edp):
				EDP_min = EDP + psum_edp
				EDP_min_id = para_type_2
	
		line = "layer " + str(i) + " :  EDP_min_id = " + str(EDP_min_id) + "  EDP_Min =  " + str(EDP_min)
		print(line, file = f)
		para_list.append(EDP_min_id)
		edp_sum += EDP_min
		energy_sum += energy_intra_layer[i][EDP_min_id] + Psum_energy_dict[i][EDP_min_id]
		delay_sum += delay_intra_layer[i][EDP_min_id] + Psum_delay_dict[i][EDP_min_id]
		psum_edp_sum += Psum_edp_dict[i][EDP_min_id]
		psum_energy_sum += Psum_energy_dict[i][EDP_min_id]
		psum_delay_sum += Psum_delay_dict[i][EDP_min_id]

		if i > layer_id_list[0]:
			network_param_2 = layer_dict[i]
			print(network_param_2)
			par_C_pre = para_each_layer[i-1][para_list[layer_num - 1]]["C"]
			par_K_pre = para_each_layer[i-1][para_list[layer_num - 1]]["K"]

			par_C_cur = para_each_layer[i][para_list[layer_num]]["C"]
			par_K_cur = para_each_layer[i][para_list[layer_num]]["K"]

			worstCommFlitNum, d2d_e , dram_e, EDP = calInterComm_simba(par_C_pre, par_K_pre, par_C_cur, par_K_cur, network_param_2)

			edp_sum += EDP
			inter_edp += EDP
			inter_energy += d2d_e + dram_e
			energy_sum += d2d_e + dram_e
			delay_sum += worstCommFlitNum
			inter_delay += worstCommFlitNum

			line = "layer " + str(i-1) + " to " + str(i) + " : nop_EDP =  " + str(EDP) + " nop_energy=" + str(d2d_e + dram_e) + " nop_delay=" + str(worstCommFlitNum)
			print(line, file = f)
		layer_num += 1
		
	line1 = "---- EDP ---- "
	line1_1 = str(psum_edp_sum) + "\t" + str(inter_edp) + "\t" + str(edp_sum) + "\t"
	print(line1, file = f)
	print(line1_1, file = f)

	line2 = "---- energy ---- "
	line2_1 = str(psum_energy_sum) + "\t" + str(inter_energy) + "\t" + str(energy_sum) + "\t"
	print(line2, file = f)
	print(line2_1, file = f)

	line3 = "---- delay ---- "
	line3_1 = str(psum_delay_sum) + "\t" + str(inter_delay) + "\t" + str(delay_sum) + "\t"
	print(line3, file = f)
	print(line3_1, file = f)

	line4 = "----parallel_type_dict = " +str(para_list)
	print(line4, file = f)


def get_set_comm_num_test(app_name):

	para_type = ["P","K","PK"]
	dim_seq_1 = ["K","P","Q"]
	dim_seq_2 = ["K","P","Q"]
	parallel_type_1 = {"P":{"P":16,"Q":1,"K":1},'K':{"P":1,"Q":1,"K":16},'PK':{"P":4,"Q":1,"K":4}}
	parallel_type_2 = {"P":{"P":16,"Q":1,"K":1},'K':{"P":1,"Q":1,"K":16},'PK':{"P":4,"Q":1,"K":4}}

	file_name = "./final_test/nop_set_comm/" + app_name + "_set_comm.txt"
	f = open(file_name,'w')
	line = "app_name " + app_name + "--------------------------------------"
	print(line, file = f)

	# --- 获得神经网络参数
	layer_dict, layer_id_list = getLayerParam(app_name)

	comm_set_layer = {}
	comm_chip_set_layer = {}
	
	for i in reversed(layer_id_list):
		if i < layer_id_list[-1]:
			network_param_2 = layer_dict[i+1]

			line = "layer " + str(i) + "-------------------------"
			print(line, file = f)

			comm_set_layer[i] = {}
			comm_chip_set_layer[i] = {}

			for para_type_1 in para_type:
				for para_type_2 in para_type:
					parallel_1 = parallel_type_1[para_type_1]
					parallel_2 = parallel_type_2[para_type_2]
					comm_num_dict, comm_type_dict, comm_type_times_dict, chiplet_num = getInterLayerComm(dim_seq_1, dim_seq_2, parallel_1, network_param_2, parallel_2, 0)
					comm_chip_set_dict, comm_set_dict = getSetComm(comm_num_dict)

					line0 = "parallel type : " + str(para_type_1) + " to " + str(para_type_2)
					line1 = "---comm_chip_set_dict: " + str(comm_chip_set_dict)
					line2 = "---comm_set_dict: " + str(comm_set_dict)
					print(line0, file = f)
					print(line1, file = f)
					print(line2, file = f)

					par_type_str = str(para_type_1) + "_" + str(para_type_2)
					comm_chip_set_layer[i][par_type_str] = comm_chip_set_dict
					comm_set_layer[i][par_type_str] = comm_set_dict
	
	print("comm_chip_set_layer ", comm_chip_set_layer, file = f)
	print("comm_set_layer ", comm_set_layer, file = f)
	f.close()
	return comm_chip_set_layer, comm_set_layer # [layer_num][parallel_type][pac_id][packet_num, recv_id_list] , parallel_type = parallel_layer_i + "_" + parallel_layer_(i+1)

if __name__ == '__main__':

	# python3 inter_layer_test.py ours resnet50
	app_name = str(sys.argv[2])
	struct_name = str(sys.argv[1])
	if struct_name == "ours":
		nop_noc_test_ours_iodie(app_name)
		#noc_test_ours(app_name)
	elif struct_name == "nnbaton":
		#nop_noc_test_nnbaton(app_name)
		noc_test_nnbaton(app_name)
	elif struct_name == "simba" or struct_name == "mem_simba":
		#nop_noc_test_simba(app_name, struct_name)
		noc_test_simba(app_name, struct_name)
	elif struct_name == "ours_simba":
		noc_test_simba_on_ours(app_name, struct_name)
	elif struct_name == "simba_ours":
		noc_test_ours_on_simba(app_name)