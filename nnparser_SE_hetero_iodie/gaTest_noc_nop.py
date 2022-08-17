import math
import os
import sys
import random
import numpy as np
import copy
from enum import Enum
from single_engine_predict_intralayer_iodie import *
from mesh_hetero import *
from matplotlib import pyplot as plt
from config import *
import openpyxl
from randomTest import *

# Parameter
# 
# GA parameter
pm = 0.2
pc = 0.8
num_children = int(50)
num_children_r = int(50)
num_iter = 10

class GASolver():
	def __init__(self, num_children, num_iter, memory_param, NoC_param, if_multicast, record_filename, input_act_enough=0, fuse_tag = "initial", debug = 0):
		self.GAGen = None
		self.memory_param = memory_param
		self.NoC_param = NoC_param
		self.if_multicast = if_multicast
		self.input_act_enough = input_act_enough
		self.fuse_tag = fuse_tag
		self.debug = debug

		self.num_children = num_children
		self.num_iter = num_iter

		self.parent_ratio = 1
		self.elite_ratio = 0.05

		self.generation = None
		self.fitness = None
		self.fitness_elite = None
		self.select_table = None

		# total per layer
		self.best_out = {"fitness":None, "e_sum":None, "delay":None, "code":None, "degrade_ratio":None, "compuation_cycles":None}

		self.parent = None

		# total per spatial parallel method
		self.best_fitness_record = None
		self.best_evalatuion_record = None
		self.generation_record = {}
		self.record = []

		self.record_filename = record_filename

	def setGAGen(self, GAGen):
		self.GAGen = GAGen
		self.total_reset()
	
	def generation_reset(self):
		self.generation = None
		self.fitness = None
		self.select_table = None
		self.parent = None
		self.best_fitness_record = None

	def total_reset(self):
		self.generation_reset()
		self.fitness_elite = None
		self.best_out = {"fitness":None, "e_sum":None, "delay":None, "code":None, "degrade_ratio":None, "compuation_cycles":None}
		self.best_evalatuion_record = None
		self.generation_record = {}
		self.record = []

	def calFitnessAll(self, Ga_code, flag = "ours"):
		for_list, act_wgt_dict, out_dict, parallel_dim_list, partition_list, code = self.GAGen.GaGetChild(Ga_code)
		#---计算适应度---
		delay, degrade_ratio, compuation_cycles, runtime_list,cp_list,utilization_ratio_list, energy_dram_list, energy_L2_list, energy_L1_list, energy_die2die, energy_MAC, energy_psum_list, delay_psum, worstlinks = \
			calFitness(for_list, act_wgt_dict, out_dict, parallel_dim_list, partition_list, self.GAGen.network_param, self.GAGen.HW_param, self.memory_param, self.NoC_param, self.if_multicast, self.input_act_enough, self.fuse_tag, flag = flag)
		#---比较适应度，并记录相关变量---
		e_mem = sum(energy_dram_list)+sum(energy_L2_list)+sum(energy_L1_list)
		e_sum = e_mem + energy_die2die+energy_MAC + energy_psum_list[2]
		edp_res = (delay + delay_psum) * e_sum  /(PE_freq * freq_1G) # pJ*s
		fitness = edp_res
		return fitness, e_sum, delay, code, degrade_ratio, compuation_cycles, runtime_list,cp_list,utilization_ratio_list, energy_dram_list, energy_L2_list, energy_L1_list, energy_die2die, energy_MAC, energy_psum_list, delay_psum, worstlinks

	def getFirstGeneration(self):
		self.generation = []
		self.fitness = np.zeros(self.num_children)
		self.generation_record = {}
		for i in range(self.num_children):
			if self.GAGen.GaType == "index":
				child_gen = self.GAGen.getGaCode_index()
			elif self.GAGen.GaType == "num":
				child_gen = self.GAGen.getGaCode_num()
			self.generation.append(child_gen)
			fitness_i, e_sum, delay, code, degrade_ratio, compuation_cycles, runtime_list,cp_list,utilization_ratio_list, energy_dram_list, energy_L2_list, energy_L1_list, energy_die2die, energy_MAC, energy_psum_list, delay_psum, worstlinks = self.calFitnessAll(self.generation[i])
			self.fitness[i] = fitness_i
			output_record = {	"delay":delay, "degrade_ratio":degrade_ratio, "compuation_cycles":compuation_cycles, "runtime_list":runtime_list, "cp_list":cp_list, "utilization_ratio_list":utilization_ratio_list, \
								"energy_dram_list":energy_dram_list, "energy_L2_list":energy_L2_list, "energy_L1_list":energy_L1_list, "energy_die2die":energy_die2die, "energy_MAC":energy_MAC, \
								"energy_psum_list":energy_psum_list, "delay_psum":delay_psum, "worstlinks":worstlinks}
			self.record.append(output_record)
			self.generation_record[i] = output_record 
			if self.best_out["fitness"] == None or self.best_out["fitness"] > fitness_i:
				self.best_out["fitness"] = fitness_i
				self.best_out["e_sum"] = e_sum
				self.best_out["delay"] = delay
				self.best_out["code"] = code
				self.best_out["degrade_ratio"] = degrade_ratio
				self.best_out["compuation_cycles"] = compuation_cycles

	def mutate_index(self, code):
		for i in range(len(code)):
			ran1 = random.random()
			if ran1 < pm:
				if i < 2:
					ran = 1 - code[i]
				elif i < 8+2:
					if i % 2 == 0:
						dim = dim_list[int((i-2)/2)]
						num_max = len(self.GAGen.loop_tile_dict[dim])
					else:
						num_max = len(self.GAGen.tile_order_dict)
					ran = random.randint(0,num_max-1)
				else:
					ran = random.random()
				code[i] = ran
		return code
	
	def mutate_num(self, code):
		for i in range(len(code)):
			ran1 = random.random()
			if ran1 < pm:
				if i < 2:
					ran = 1 - code[i]
				elif i < 8+2:
					num = code[i]
					ran = random.randint(math.ceil(num*0.5),math.ceil(num*2))
				else:
					ran = random.random()
				code[i] = ran
		return code
	
	def crossover(self, code1, code2):
		childCode_1 = copy.deepcopy(code1)
		childCode_2 = copy.deepcopy(code2)

		index_list = list(range(len(code1)))
		random.shuffle(index_list)
		index_1 = index_list[0]
		index_2 = index_list[1]
		if index_1 > index_2:
			index_1, index_2 = index_2, index_1
		
		childCode_1[index_1:index_2], childCode_2[index_1:index_2] = code2[index_1:index_2], code1[index_1:index_2]

		return childCode_1, childCode_2

	def selectGeneration(self):
		num_ran1 = random.random()
		flag = 0
		time1 = 0
		while flag == 0:
			if num_ran1 < self.select_table[time1]:
				flag = 1
				break
			time1 += 1
		
		return time1
	
	def orderGeneration(self):
		fitness_order = sorted(enumerate(self.fitness), key=lambda x:x[1])
		fitness_index = [f[0] for f in fitness_order]
		self.fitness = [f[1] for f in fitness_order]
		generation_order = []

		for i in range(len(fitness_index)):
			id = fitness_index[i]
			generation_order.append(self.generation[id])
		self.generation = generation_order

		self.parent = copy.deepcopy(self.generation[:int(self.parent_ratio*self.num_children)])

	def calProbability(self):
		num_parent = int(self.parent_ratio * self.num_children)
		fitness_select = self.fitness[:num_parent]
		fitness_prob = np.zeros(num_parent)
		for i in range(len(fitness_select)):
			fitness_prob[i] = 1 / fitness_select[i]
		total = fitness_prob.sum(axis = 0)
		self.select_table = np.zeros(num_parent)
		probability_sum = 0
		for i in range(0, num_parent):
			probability_sum += fitness_prob[i] / total
			self.select_table[i] = probability_sum

	def getNpMin(self, list):
		min_num = min(list)
		min_index = list.index(min(list))
		return min_num, min_index

	def getNextGeneration(self):
		childNum = int(self.num_children/2)
		fitness_parent = []

		self.generation = []
		for i in range(childNum):
			parent_id_1 = self.selectGeneration()
			parent_id_2 = self.selectGeneration()
			while parent_id_1 == parent_id_2:
				parent_id_2 = self.selectGeneration()

			parent1 = self.parent[parent_id_1]
			parent2 = self.parent[parent_id_2]

			fitness_parent.append(self.fitness[parent_id_1])
			fitness_parent.append(self.fitness[parent_id_2])

			ran = random.random()
			if ran < pc:
				child1, child2 = self.crossover(parent1, parent2)
			else:
				if self.GAGen.GaType == "index":
					child1 = self.GAGen.getGaCode_index()
					child2 = self.GAGen.getGaCode_index()
				elif self.GAGen.GaType == "num":
					child1 = self.GAGen.getGaCode_num()
					child2 = self.GAGen.getGaCode_num()

			if self.GAGen.GaType == "index":
				child1 = self.mutate_index(child1)
				child2 = self.mutate_index(child2)
			else:
				child1 = self.mutate_num(child1)
				child2 = self.mutate_num(child2)

			#self.generation[2*i] = child1
			#self.generation[2*i+1] = child2
			self.generation.append(child1)
			self.generation.append(child2)
		
		self.generation = self.generation + self.parent[:math.ceil(self.elite_ratio*self.num_children)]
		self.fitness_elite = self.fitness[:math.ceil(self.elite_ratio*self.num_children)]

		return fitness_parent

	def evaluationRecord(self):
		filename = self.record_filename
		file_data = []
		title = []

		for name in self.best_evalatuion_record[0]:
			title.append(name)
		for index in reversed(range(len(self.best_evalatuion_record))):
			record = self.best_evalatuion_record[index]
			record_list = []
			for item in record:
				record_list.append(str(record[item]))
			file_data.append(record_list)
		
		workbook = openpyxl.Workbook()
		sheet = workbook.get_sheet_by_name('Sheet') 
		# 写入标题
		for col,column in enumerate(title):
			sheet.cell(1, col+1, column)
		# 写入每一行
		for row, data in enumerate(file_data):
			for col, column_data in enumerate(data):
				sheet.cell(row+2, col+1, column_data)

		workbook.save(filename)

	def gaIter(self, first_flag = 1):
		print("---- Begin GA")

		self.best_fitness_record = []
		self.best_evalatuion_record = []

		if first_flag == 1:
			self.reset()
			self.getFirstGeneration()

		for i in range(self.num_iter):
			print("iter = ", str(i))
			self.orderGeneration()
			self.calProbability()
			fitness_father = self.getNextGeneration()
			self.fitness = []
			self.generation_record = {}
			for j in range(len(self.generation)):
				code = self.generation[j]
				fitness_i, e_sum, delay, code, degrade_ratio, compuation_cycles, runtime_list,cp_list,utilization_ratio_list, energy_dram_list, energy_L2_list, energy_L1_list, \
				energy_die2die, energy_MAC, energy_psum_list, delay_psum, worstlinks = self.calFitnessAll(code)
				output_record = {"spatial parallel":self.GAGen.spatial_parallel, "delay":delay, "degrade_ratio":degrade_ratio, "compuation_cycles":compuation_cycles, "runtime_list":runtime_list, "cp_list":cp_list, "utilization_ratio_list":utilization_ratio_list, \
									"energy_dram_list":energy_dram_list, "energy_L2_list":energy_L2_list, "energy_L1_list":energy_L1_list, "energy_die2die":energy_die2die, "energy_MAC":energy_MAC, \
									"energy_psum_list":energy_psum_list, "delay_psum":delay_psum, "worstlinks":worstlinks}
				self.fitness.append(fitness_i)
				self.record.append(output_record)
				self.generation_record[j] = output_record

				if self.best_out["fitness"] == None or self.best_out["fitness"] > fitness_i:
					self.best_out["fitness"] = fitness_i
					self.best_out["e_sum"] = e_sum
					self.best_out["delay"] = delay
					self.best_out["code"] = code
					self.best_out["degrade_ratio"] = degrade_ratio
					self.best_out["compuation_cycles"] = compuation_cycles

			print("--- times = ",str(i+1))

			min_num, min_index = self.getNpMin(self.fitness)
			code_best = self.generation[min_index]
			evaluation_best = self.generation_record[min_index]

			print("fitness_min = ",min_num)
			self.best_fitness_record.append(min_num)
			self.best_evalatuion_record.append(evaluation_best)

		print("GA_FINAL_RESULT : ", self.best_fitness_record[-1])
		print(code_best)
		

def fitnessPlot(fitness_list, GaType):
	x = []
	for i in range(len(fitness_list)):
		x.append(i)
	plt.figure("ga_"+GaType)
	plt.plot(x,fitness_list)
	plt.scatter(x[-1],fitness_list[-1],s=10)
	xy = (x[-1], round(fitness_list[-1]))
	plt.annotate("(%s,%s)" % xy, xy=xy, xytext=(-70, 10), textcoords='offset points')
	plt.title("GA_fitness_iter"+GaType)
	plt.savefig("./GaTest/GaTest_"+GaType+".png")

def fitnessPlot_random(fitness_list, GaType):
	x = list(range(num_iter))
	y = []
	for i in range(len(fitness_list)):
		if (i+1) % num_children_r == 0:
			y.append(fitness_list[i])
	plt.figure("random"+GaType)
	plt.plot(x,y)
	plt.title("random_fitness_iter_"+GaType)
	plt.scatter(x[-1],fitness_list[-1],s=10)
	xy = (x[-1], round(fitness_list[-1]))
	plt.annotate("(%s,%s)" % xy, xy=xy, xytext=(-70, 10), textcoords='offset points')
	plt.savefig("./GaTest/randomTest_"+GaType+".png")
	#plt.legend()
	print("final_fitness : ", fitness_list[-1])

def gaTest_NoC_ours(GaType, app_name, layer_name, network_param):
	# --- 硬件参数
	HW_param = {"Chiplet":[2,2],"PE":[4,4],"intra_PE":{"C":16,"K":16}}       	# from granularity exploration
	spatial_parallel = {"pe":{"P":1, "Q":1, "C":16, "K":16, "R":1, "S":1}, "chiplet":{"P":2, "Q":1, "C":1, "K":8, "R":1, "S":1}, "package":{"P":2, "Q":1, "C":1, "K":2, "R":1, "S":1}}
	# memory_param = {"OL1":1.5,"OL2":1.5*16,"AL1":800/1024,"AL2":64,"WL1":18,"WL2":18*16} 	from nnbaton
	memory_param = {"OL1":8 ,"OL2":128,"AL1":16,"AL2":256,"WL1":64,"WL2":1024}		# from granularity exploration
	NoC_w = HW_param["PE"][1] + 1
	NOC_NODE_NUM = NoC_w * HW_param["PE"][0]
	NoP_w = HW_param["Chiplet"][1] + 1
	NOP_SIZE = NoP_w * HW_param["Chiplet"][0]
	TOPO_param = {"NoC_w":NoC_w, "NOC_NODE_NUM": NOC_NODE_NUM, "NoP_w": NoP_w, "NOP_SIZE": NOP_SIZE,"nop_scale_ratio": nop_bandwidth/noc_bandwidth}
	
	# --- 生成noc-nop结构图
	NoC_param, all_sim_node_num = construct_noc_nop_topo(TOPO_param["NOC_NODE_NUM"],TOPO_param["NoC_w"], TOPO_param["NOP_SIZE"],TOPO_param["NoP_w"], TOPO_param["nop_scale_ratio"], topology = 'Ring')
	debug = 0
	if_multicast = 1

	# --- 神经网络参数
	#layer_dict = getLayerParam(app_name)

	edp_res_min_dict = {}
	energy_min_dict = {}
	delay_min_dict = {}
	code_min_dict = {}

	GA_Solver = GASolver(memory_param, NoC_param, if_multicast)

	filename = './GaTest/'+app_name+"_"+layer_name+"_random.xls"
	GATest = GaEncode(GaType, network_param, HW_param, spatial_parallel, debug)
	GA_Solver.setGAGen(GATest)

	GA_Solver.gaIter()
	fitness_list = GA_Solver.best_fitness_record

	return fitness_list

def fitnessPlot_compare(fitness_dict_ga, fitness_dict_random):
	x = list(range(num_iter))
	fitness_dict_random_new = {}
		
	for item in fitness_dict_random:
		fitness_dict_random_new[item] = []
		for i in range(len(fitness_dict_random[item])):
			if (i+1) % num_children_r == 0:
				fitness_dict_random_new[item].append(fitness_dict_random[item][i])
	plt.figure("compare")
	for item in fitness_dict_ga:
		plt.plot(x, fitness_dict_ga[item], label=item)
		plt.scatter(x[-1],fitness_dict_ga[item][-1],s=10)
		xy = (x[-1], round(fitness_dict_ga[item][-1]))
		plt.annotate("(%s,%s)" % xy, xy=xy, xytext=(-70, 10), textcoords='offset points')
	for item in fitness_dict_random_new:
		plt.plot(x, fitness_dict_random_new[item], label=item)
		plt.scatter(x[-1],fitness_dict_random_new[item][-1],s=10)
		xy = (x[-1], round(fitness_dict_random_new[item][-1]))
		plt.annotate("(%s,%s)" % xy, xy=xy, xytext=(-70, 10), textcoords='offset points')
	plt.legend()
	plt.title("compare_fitness_iter")
	plt.savefig("./GaTest/compare.png")
	plt.show

if __name__ == '__main__':
	#224 224 64 256 3 3
	iterTime = num_children_r * num_iter
	app_name = str(sys.argv[1])

	fitness_dict_random = {}
	fitness_dict_ga = {}

	flag = [1,1,1,1]

	output_file = open("./GaTest/fitness_record_"+app_name+".txt", 'w')

	layer_dict = getLayerParam(app_name)

	title = ["layer_name", "ga_index", "ga_num", "random_index", "random_num"]
	excel_data = []

	for layer_name in layer_dict:
		network_param = layer_dict[layer_name]
		print("-------- {} -------".format(layer_name), file = output_file)
		fitness_data = [layer_name]
		if flag[0] == 1:
			GaType = "index"
			fitness_list_ga_index = gaTest_NoC_ours(GaType, app_name, layer_name, network_param)
			#fitnessPlot(fitness_list_ga_index, GaType)
			fitness_dict_ga["ga_index"] = fitness_list_ga_index
			print("ga_index_fitness : ", fitness_dict_ga["ga_index"][-1], file = output_file)
			fitness_data.append(fitness_dict_ga["ga_index"][-1])

		if flag[1] == 1:
			GaType = "num"
			fitness_list_ga_num = gaTest_NoC_ours(GaType, app_name, layer_name, network_param)
			#fitnessPlot(fitness_list_ga_num, GaType)
			fitness_dict_ga["ga_num"] = fitness_list_ga_num
			print("ga_num_fitness : ", fitness_dict_ga["ga_num"][-1], file = output_file)
			fitness_data.append(fitness_dict_ga["ga_num"][-1])

		if flag[2] == 1:
			GaType = "index"
			fitness_list_random_index = randomTest_NoC_ours(GaType, app_name, iterTime, layer_name, network_param)
			#fitnessPlot_random(fitness_list_random_index, GaType)
			fitness_dict_random["random_index"] = fitness_list_random_index
			print("random_index_fitness : ", fitness_dict_random["random_index"][-1], file = output_file)
			fitness_data.append(fitness_dict_random["random_index"][-1])

		if flag[3] == 1:
			GaType = "num"
			fitness_list_random_num = randomTest_NoC_ours(GaType, app_name, iterTime, layer_name, network_param)
			#fitnessPlot_random(fitness_list_random_num, GaType)
			fitness_dict_random["random_num"] = fitness_list_random_num
			print("random_num_fitness : ", fitness_dict_random["random_num"][-1], file = output_file)
			fitness_data.append(fitness_dict_random["random_num"][-1])

		print("", file = output_file)
		excel_data.append(fitness_data)
	fitnessPlot_compare(fitness_dict_ga,fitness_dict_random)

	filename = "./GaTest/fitness_excel_"+app_name+".xls"
	workbook = openpyxl.Workbook()
	sheet = workbook.get_sheet_by_name('Sheet')
	for col,column in enumerate(title):
		sheet.cell(1, col+1, column)
	
	for row, data in enumerate(excel_data):
		for col, column_data in enumerate(data):
			sheet.cell(row+2, col+1, column_data)

	workbook.save(filename)

	#print("fitness_random_index: ", fitness_dict_random["random_index"][-1])
	#print("fitness_random_num: ", fitness_dict_random["random_num"][-1])


