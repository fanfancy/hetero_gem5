import math
import os
from platform import architecture
import sys
import random
import copy
import argparse
from matplotlib import pyplot as plt
import shutil
import openpyxl

# Parameter
chiplet_parallel_list = ["P_stable", "PK_stable", "K_stable"]
PE_Frequency = 1000 * 1000 * 1000

# degub signal
debug_in_workload_extract = 0
debug_in_layer_fuse = 0
debug_in_main = 1
debug_in_txt_extract = 0
debug_in_workload_partition = 1

# getLayerParam: 根据优化目标objective，输出对应的层的理论评估结果
def getLayerParam(SE_path_dir, app_name, objective):
	layer_num = 0
	layer_dict = {}
	layer_id_list = []
	layer_name_list = []
	f = open(SE_path_dir + "/nn_input_noc_nop/" + app_name + ".txt")

	lines = f.readlines()
	for line in lines:
		if line.startswith("#") or line.startswith("*"):
			pass
		elif line != "\n":
			line_item = line.split(" ")
			layer_num += 1
			H = int(line_item[1])
			M = int(line_item[2])
			C = int(line_item[3])
			R = int(line_item[4])
			S = int(line_item[4])
			stride = int(line_item[5])
			padding = int(line_item[6])
			K = int(line_item[7])
			P = int(line_item[8])
			Q = int(line_item[9])

			iact_num = H * M * C
			oact_num = P * Q * K
			weight_num = R * S * K * C
			compute_num = P * Q * K * R * S * C
			iparam_num = iact_num + weight_num
			param_num = iparam_num + oact_num

			iparam_vs_compute = iparam_num / compute_num
			param_vs_compute = param_num / compute_num

			iparam_mux_compute = iparam_num * compute_num
			param_mux_compute = param_num * compute_num

			layer_id_list.append(layer_num)
			layer_name = "layer" + str(layer_num)
			layer_name_list.append(layer_name)
			layer_dict[layer_name] = {"iact_num":iact_num, "oact_num":oact_num, "weight_num":weight_num, \
									"compute_num":compute_num, "iparam_num":iparam_num, "param_num":param_num, \
									"iparam_vs_compute":iparam_vs_compute, "param_vs_compute":param_vs_compute, \
									"iparam_mux_compute":iparam_mux_compute, "param_mux_compute":param_mux_compute}

	if objective == "noPar" or objective == "totalPar":
		return layer_id_list, layer_name_list

	f.close()
	result_dict = {}
	for layer_name, result_items in layer_dict.items():
		result = result_items[objective]
		result_dict[layer_name] = result
	return result_dict

# partition: 获得评估目标下的最优划分方案
# --- 方案迭代方法：随机，partition_num为方案数目
# --- objective_tag：评估目标（min、max、balance），todo：min和max的考量没有想的很清楚
# --- method_best：输出为划分点的层id，相当于每个workload的第一层的id
def partition(fitness_list, partition_num, objective_tag="balance"):
	def partition_max(list, partition_num):
		list_order = sorted(list, key = lambda x: x, reverse=True)
		TH = list_order[partition_num-1]
		select_index_list = []
		TH_fitness_list = []
		rest_partiton_num = partition_num
		for index, fitness in enumerate(list):
			if fitness > TH:
				select_index_list.append(index)
				rest_partiton_num -= 1
			elif fitness == TH:
				TH_fitness_list.append(index)
			else:
				pass
		
		TH_num = len(TH_fitness_list)
		assert(TH_num >= rest_partiton_num)
		TH_list = list(range(TH_num))
		random.shuffle(TH_list)
		for i in range(rest_partiton_num):
			select_index_list.append(TH_list[i])	
		select_index_list.sort()
		return select_index_list
	
	def partition_min(list, partition_num):
		list_order = sorted(list, key = lambda x: x, reverse=False)
		TH = list_order[partition_num-1]
		select_index_list = []
		TH_fitness_list = []
		rest_partiton_num = partition_num
		for index, fitness in enumerate(list):
			if fitness < TH:
				select_index_list.append(index)
				rest_partiton_num -= 1
			elif fitness == TH:
				TH_fitness_list.append(index)
			else:
				pass
		
		TH_num = len(TH_fitness_list)
		assert(TH_num >= rest_partiton_num)
		TH_list = list(range(TH_num))
		random.shuffle(TH_list)
		for i in range(rest_partiton_num):
			select_index_list.append(TH_list[i])
		select_index_list.sort()
		return select_index_list
		
	def partition_balance(fitness_list, partition_num, method_num):
		# cal_unbalance: 评估方案的不平衡度，不平衡度越高，方案越不好
		def cal_unbalance(list):
			list.sort()
			min = list[0]
			max = list[-1]
			unbalance = 1 - min/max
			return unbalance

		# get_method: 获取方案，随机获取
		def get_method(layer_num, partition_num, method_num):
			index_list = list(range(layer_num))
			method_list = []
			num = 0
			unchange_times = 0
			while num < method_num:
				random.shuffle(index_list)
				select_index_list = index_list[0: partition_num-1]
				assert(len(select_index_list) == partition_num-1)
				select_index_list.sort()
				if select_index_list not in method_list:
					method_list.append(select_index_list)
					num += 1
					unchange_times = 0
				else:
					unchange_times += 1
				
				if unchange_times > 100:
					break
			return method_list
		
		layer_num = len(fitness_list)
		method_list = get_method(layer_num, partition_num, method_num)
		unbalance_best = None
		method_best = None
		for method in method_list:
			unbalance = 0
			index_pre = 0
			for index in method:
				unbalance += cal_unbalance(fitness_list[index_pre: index+1])
			if unbalance_best == None or unbalance < unbalance_best:
				unbalance_best = unbalance
				method_best = method

		return method_best
	
	if objective_tag == "balance":
		method_num = 1000
		select_index_list = partition_balance(fitness_list, partition_num, method_num)
	elif objective_tag == "max":
		select_index_list = partition_max(fitness_list, partition_num)
	elif objective_tag == "min":
		select_index_list = partition_min(fitness_list, partition_num)
	
	return select_index_list

# getWorkloadPartition: 任务划分
# --- objective: "noPar"(以模型为粒度)、"totalPar"(以层为粒度)、"iact_num"、"oact_num"...(省略的详见代码75行layer_dict内容)
def getWorkloadPartition(app_name, SE_path_dir, objective, select_min_num):
	if objective == "noPar":
		workload_dict = {}
		layer_id_list, layer_name_list = getLayerParam(SE_path_dir, app_name, objective)
		workload_dict[app_name+"w0"] = {"layer_id_list":layer_id_list, "layer_name_list":layer_name_list, "workload_name": "w0"}
		return workload_dict
	elif objective == "totalPar":
		workload_dict = {}
		layer_id_list, layer_name_list = getLayerParam(SE_path_dir, app_name, objective)
		for id in range(len(layer_id_list)):
			w_name = app_name + "w" + str(id)
			layer_id = layer_id_list[id]
			layer_name = layer_name_list[id]
			workload_name = "w" + str(id)
			workload_dict[w_name] = {"layer_id_list": [layer_id], "layer_name_list": [layer_name], "workload_name": workload_name}
		return workload_dict
	
	# 1. 获得各层的优化对象objective的理论评估结果
	result_dict = getLayerParam(SE_path_dir, app_name, objective)
	result_list = list(result_dict.values())
	if debug_in_workload_partition == 1:
		print("DEBUG IN WORKLOAD PARTITION")
		print("---objective:{}, select_min_num:{}".format(objective, select_min_num))
		print("---result_dict: ", result_dict)

	# 2. 根据理论评估结果，进行任务划分探索
	# --- select_index_list: 划分点的层id
	select_index_list = partition(result_list, select_min_num, objective_tag="balance")

	# 3. 根据划分点，进行任务列表的整合
	workload_dict = {}
	workload_id = 0
	layer_index = 0
	for layer_name, result in result_dict.items():
		if layer_index != 0 and layer_index-1 in select_index_list:
			workload_id += 1
		else:
			pass
		workload_name = app_name + "w" + str(workload_id)
		if workload_name not in workload_dict:
			workload_dict[workload_name] = {"layer_id_list":[], "layer_name_list":[], "workload_name": "w{}".format(workload_id)}
		
		layer_id = layer_index + 1
		workload_dict[workload_name]["layer_id_list"].append(layer_id)
		workload_dict[workload_name]["layer_name_list"].append(layer_name)
		assert(layer_name == "layer"+str(layer_id))
		layer_index += 1
	
	if debug_in_workload_partition == 1:
		print("---workload_dict: ", workload_dict)
	return workload_dict

# getLayerList：获得每一层的对应的单网络仿真的layer id
def getLayerList(layer_name_dict_line):
	#layer_name_dict:  {'layer1': 'layer1', 'layer2': 'layer2'
	line = layer_name_dict_line.replace("layer_name_dict:  ", "")
	line = line.replace("{","")
	line = line.replace("}","")
	line = line.replace("\'","")
	line = line.replace(":","")
	line = line.replace(", ",",")
	line_list = line.split(",")
	layer_list = []
	layer_dict = {}
	id = 0
	for item in line_list:
		if item.startswith("layer"):
			item_list = item.split(" ")
			layer_init = item_list[0]
			layer_real = item_list[1]
			layer_id = str(layer_real.replace("layer", "").replace("\n",""))
			layer_list.append(layer_id)
			layer_dict[layer_init] = layer_real
	return layer_list, layer_dict

def txt_extract(result_indir_p, result_indir_pk, result_indir_k, app_name, fuse_flag, layer_position):
	
	def lineParse(line, ratio_list = {}):
		line_split = line.split("{")
		line = line_split[1]
		line = line.replace(")","(")
		line = line.replace(",","")
		line = line.replace("\'","")
		line = line.replace(":","")
		line = line.replace("{","")
		line = line.replace("}","")
		line_list = line.split(" ")
		item_num = 0
		result_dict = {}
		for line_item in line_list:
			if item_num % 2 == 1:
				layer_id = line_list[item_num-1]
				layer_id = layer_id.replace(" ","")
				if len(ratio_list) == 0:
					result_dict[layer_id] = float(line_item)
				else:
					result_dict[layer_id] = float(line_item) * ratio_list[layer_id]
			item_num += 1
		return result_dict
	
	def lineParseList(line, ratio_list = {}):
		line_split = line.split("{")
		line = line_split[1]
		line = line.replace("\n","")
		line = line.replace("}","")
		line_item = line.split("], ")
		result_dict = {}
		for item in line_item:
			item_list = item.split(": ")
			layer_id = item_list[0]
			layer_id = layer_id.replace('\'','')
			parse_list = item_list[1]
			parse_list = parse_list.replace("]","")
			parse_list += "]"
			result_dict[layer_id] = parse_list
		return result_dict

	def extractFitness_initial(result_indir, layer_position): 
		# --- read files ---
		result_file_init = os.path.join(result_indir, "final_result_record_initial_{}Layer.txt".format(layer_position))
		f_init = open(result_file_init)

		lines_init = f_init.readlines()
		line_init_edp = lines_init[0]
		line_init_energy = lines_init[1]
		line_init_latency = lines_init[2]

		line_init_noc_DR = lines_init[5]
		line_init_L2_to_DRAM_DR = lines_init[6]
		line_init_DRAM_to_L2_DR = lines_init[7]

		line_init_input_DRAM_flit_needed = lines_init[9]
		line_init_weight_DRAM_flit_needed = lines_init[10]
		line_init_input_L2_flit_needed = lines_init[11]
		line_init_weight_L2_flit_needed = lines_init[12]
		line_init_output_rd_flit_needed = lines_init[13]
		line_init_output_wr_flit_needed = lines_init[14]
		line_init_chiplet_spatial_parallel = lines_init[15]

		line_iact_enough = lines_init[18]
		line_layer_name = lines_init[22]

		# --- extract from lines ---
		result_init_edp = lineParse(line_init_edp)
		result_init_energy = lineParse(line_init_energy)
		result_init_latency = lineParse(line_init_latency)
		result_init_NoC_DR = lineParse(line_init_noc_DR)
		result_init_L2_to_DRAM_DR = lineParse(line_init_L2_to_DRAM_DR)
		result_init_DRAM_to_L2_DR = lineParse(line_init_DRAM_to_L2_DR)

		result_init_input_DRAM_flit_needed = lineParse(line_init_input_DRAM_flit_needed)
		result_init_weight_DRAM_flit_needed = lineParse(line_init_weight_DRAM_flit_needed)
		result_init_input_L2_flit_needed = lineParse(line_init_input_L2_flit_needed)
		result_init_weight_L2_flit_needed = lineParse(line_init_weight_L2_flit_needed)
		result_init_output_rd_flit_needed = lineParse(line_init_output_rd_flit_needed)
		result_init_output_wr_flit_needed = lineParse(line_init_output_wr_flit_needed)
		result_init_chiplet_spatial_parallel = lineParseList(line_init_chiplet_spatial_parallel)
		result_iact_enough = lineParse(line_iact_enough)
		result_layer_name_list, result_layer_name_dict = getLayerList(line_layer_name)
		
		fitness_init_dict = {}

		for layer_id in result_init_edp:
			fitness_init_dict[layer_id] = { \
				"edp": result_init_edp[layer_id], \
				"energy": result_init_energy[layer_id], \
				"latency": result_init_latency[layer_id], \
				"NoC_DR": result_init_NoC_DR[layer_id], \
				"L2_to_DRAM_DR": result_init_L2_to_DRAM_DR[layer_id], \
				"DRAM_to_L2_DR": result_init_DRAM_to_L2_DR[layer_id], \
				"input_DRAM_flit_needed": result_init_input_DRAM_flit_needed[layer_id], \
				"weight_DRAM_flit_needed": result_init_weight_DRAM_flit_needed[layer_id], \
				"input_L2_flit_needed": result_init_input_L2_flit_needed[layer_id], \
				"weight_L2_flit_needed": result_init_weight_L2_flit_needed[layer_id], \
				"output_rd_flit_needed": result_init_output_rd_flit_needed[layer_id], \
				"output_wr_flit_needed": result_init_output_wr_flit_needed[layer_id], \
				"chiplet_spatial_parallel": result_init_chiplet_spatial_parallel[layer_id], \
				"iact_enough": result_iact_enough[layer_id]
				}

		return fitness_init_dict, result_layer_name_list

	def extractFitness_fuse_head_tail(result_indir, layer_position):
		# --- read files ---
		result_file_init = os.path.join(result_indir, "final_result_record_initial_{}.txt".format(layer_position))
		result_file_head = os.path.join(result_indir, "final_result_record_headFuse_{}.txt".format(layer_position))
		result_file_tail = os.path.join(result_indir, "final_result_record_tailFuse_{}.txt".format(layer_position))

		f_init = open(result_file_init)
		lines_init = f_init.readlines()
		line_init_edp = lines_init[0]
		result_init_edp = lineParse(line_init_edp)

		if not os.path.exists(result_file_head):
			fitness_head_dict = {}
			fitness_tail_dict = {}
			for layer_id in result_init_edp:
				fitness_head_dict[layer_id] = [None, None, None, None, None, None, None]
				fitness_tail_dict[layer_id] = [None, None, None, None, None, None, None]
			return fitness_head_dict, fitness_tail_dict
		f_head = open(result_file_head)
		f_tail = open(result_file_tail)

		lines_head = f_head.readlines()
		line_head_edp = lines_head[0]
		line_head_energy = lines_head[1]
		line_head_latency = lines_head[2]
		line_head_noc_DR = lines_head[5]
		line_head_L2_to_DRAM_DR = lines_head[6]
		line_head_DRAM_to_L2_DR = lines_head[7]

		line_head_input_DRAM_flit_needed = lines_head[9]
		line_head_weight_DRAM_flit_needed = lines_head[10]
		line_head_input_L2_flit_needed = lines_head[11]
		line_head_weight_L2_flit_needed = lines_head[12]
		line_head_output_rd_flit_needed = lines_head[13]
		line_head_output_wr_flit_needed = lines_head[14]
		line_head_chiplet_spatial_parallel = lines_head[15]

		line_head_iact_enough = lines_head[18]
		line_head_fuse_parNum = lines_head[20].replace("par_num:  ", "")

		lines_tail = f_tail.readlines()
		line_tail_edp = lines_tail[0]
		line_tail_energy = lines_tail[1]
		line_tail_latency = lines_tail[2]
		line_tail_noc_DR = lines_tail[5]
		line_tail_L2_to_DRAM_DR = lines_tail[6]
		line_tail_DRAM_to_L2_DR = lines_tail[7]

		line_tail_input_DRAM_flit_needed = lines_tail[9]
		line_tail_weight_DRAM_flit_needed = lines_tail[10]
		line_tail_input_L2_flit_needed = lines_tail[11]
		line_tail_weight_L2_flit_needed = lines_tail[12]
		line_tail_output_rd_flit_needed = lines_tail[13]
		line_tail_output_wr_flit_needed = lines_tail[14]
		line_tail_chiplet_spatial_parallel = lines_tail[15]

		line_tail_iact_enough = lines_tail[18]
		line_tail_fuse_parNum = lines_tail[20]#.replace("par_num:  ", "")

		# --- extract from lines ---
		result_head_edp = lineParse(line_head_edp)
		result_tail_edp = lineParse(line_tail_edp)

		result_head_energy = lineParse(line_head_energy)
		result_tail_energy = lineParse(line_tail_energy)

		result_head_latency = lineParse(line_head_latency)
		result_tail_latency = lineParse(line_tail_latency)

		result_head_NoC_DR = lineParse(line_head_noc_DR)
		result_tail_NoC_DR = lineParse(line_tail_noc_DR)

		result_head_L2_to_DRAM_DR = lineParse(line_head_L2_to_DRAM_DR)
		result_tail_L2_to_DRAM_DR = lineParse(line_tail_L2_to_DRAM_DR)

		result_head_DRAM_to_L2_DR = lineParse(line_head_DRAM_to_L2_DR)
		result_tail_DRAM_to_L2_DR = lineParse(line_tail_DRAM_to_L2_DR)

		result_head_input_DRAM_flit_needed = lineParse(line_head_input_DRAM_flit_needed)
		result_tail_input_DRAM_flit_needed = lineParse(line_tail_input_DRAM_flit_needed)

		result_head_weight_DRAM_flit_needed = lineParse(line_head_weight_DRAM_flit_needed)
		result_tail_weight_DRAM_flit_needed = lineParse(line_tail_weight_DRAM_flit_needed)

		result_head_input_L2_flit_needed = lineParse(line_head_input_L2_flit_needed)
		result_tail_input_L2_flit_needed = lineParse(line_tail_input_L2_flit_needed)

		result_head_weight_L2_flit_needed = lineParse(line_head_weight_L2_flit_needed)
		result_tail_weight_L2_flit_needed = lineParse(line_tail_weight_L2_flit_needed)

		result_head_output_rd_flit_needed = lineParse(line_head_output_rd_flit_needed)
		result_tail_output_rd_flit_needed = lineParse(line_tail_output_rd_flit_needed)

		result_head_output_wr_flit_needed = lineParse(line_head_output_wr_flit_needed)
		result_tail_output_wr_flit_needed = lineParse(line_tail_output_wr_flit_needed)

		result_head_chiplet_spatial_parallel = lineParse(line_head_chiplet_spatial_parallel)
		result_tail_chiplet_spatial_parallel = lineParse(line_tail_chiplet_spatial_parallel)

		result_head_iact_enough = lineParse(line_head_iact_enough)
		result_tail_iact_enough = lineParse(line_tail_iact_enough)

		result_head_fuse_parNum = lineParse(line_head_fuse_parNum)
		result_tail_fuse_parNum = lineParse(line_tail_fuse_parNum)

		fitness_head_dict = {}
		fitness_tail_dict = {}
		
		for layer_id in result_init_edp:
			if layer_id in result_head_edp:
				tile_ratio_head = result_head_fuse_parNum[layer_id]
				result_head_edp[layer_id] *= tile_ratio_head * tile_ratio_head
				result_head_energy[layer_id] *= tile_ratio_head
				result_head_latency[layer_id] *= tile_ratio_head
				result_head_NoC_DR[layer_id] *= 1
				result_head_L2_to_DRAM_DR[layer_id] *= 1
				result_head_DRAM_to_L2_DR[layer_id] *= 1
				result_head_iact_enough[layer_id] *= 1
				result_head_input_DRAM_flit_needed[layer_id] *= tile_ratio_head
				result_head_weight_DRAM_flit_needed[layer_id] *= tile_ratio_head
				result_head_input_L2_flit_needed[layer_id] *= tile_ratio_head
				result_head_weight_L2_flit_needed[layer_id] *= tile_ratio_head
				result_head_output_rd_flit_needed[layer_id] *= tile_ratio_head
				result_head_output_wr_flit_needed[layer_id] *= tile_ratio_head
				result_head_chiplet_spatial_parallel[layer_id] = result_head_chiplet_spatial_parallel[layer_id]
			else:
				result_head_edp[layer_id] = None
				result_head_energy[layer_id] = None
				result_head_latency[layer_id] = None
				result_head_NoC_DR[layer_id] = None
				result_head_L2_to_DRAM_DR[layer_id] = None
				result_head_DRAM_to_L2_DR[layer_id] = None
				result_head_iact_enough[layer_id] = None
				result_head_input_DRAM_flit_needed[layer_id] = None
				result_head_weight_DRAM_flit_needed[layer_id] = None
				result_head_input_L2_flit_needed[layer_id] = None
				result_head_weight_L2_flit_needed[layer_id] = None
				result_head_output_rd_flit_needed[layer_id] = None
				result_head_output_wr_flit_needed[layer_id] = None
				result_head_chiplet_spatial_parallel[layer_id] = None

			if layer_id in result_tail_edp:
				tile_ratio_tail = result_tail_fuse_parNum[layer_id]
				result_tail_edp[layer_id] *= tile_ratio_tail * tile_ratio_tail
				result_tail_energy[layer_id] *= tile_ratio_tail
				result_tail_latency[layer_id] *= tile_ratio_tail
				result_tail_NoC_DR[layer_id] *= 1
				result_tail_L2_to_DRAM_DR[layer_id] *= 1
				result_tail_DRAM_to_L2_DR[layer_id] *= 1
				result_tail_iact_enough[layer_id] *= 1
				result_tail_input_DRAM_flit_needed[layer_id] *= tile_ratio_tail
				result_tail_weight_DRAM_flit_needed[layer_id] *= tile_ratio_tail
				result_tail_input_L2_flit_needed[layer_id] *= tile_ratio_tail
				result_tail_weight_L2_flit_needed[layer_id] *= tile_ratio_tail
				result_tail_output_rd_flit_needed[layer_id] *= tile_ratio_tail
				result_tail_output_wr_flit_needed[layer_id] *= tile_ratio_tail
				result_tail_chiplet_spatial_parallel[layer_id] = result_tail_chiplet_spatial_parallel[layer_id]
			else:
				result_tail_edp[layer_id] = None
				result_tail_energy[layer_id] = None
				result_tail_latency[layer_id] = None
				result_tail_NoC_DR[layer_id] = None
				result_tail_L2_to_DRAM_DR[layer_id] = None
				result_tail_DRAM_to_L2_DR[layer_id] = None
				result_tail_iact_enough[layer_id] = None
				result_tail_input_DRAM_flit_needed[layer_id] = None
				result_tail_weight_DRAM_flit_needed[layer_id] = None
				result_tail_input_L2_flit_needed[layer_id] = None
				result_tail_weight_L2_flit_needed[layer_id] = None
				result_tail_output_rd_flit_needed[layer_id] = None
				result_tail_output_wr_flit_needed[layer_id] = None
				result_tail_chiplet_spatial_parallel[layer_id] = result_tail_chiplet_spatial_parallel[layer_id]
			
			fitness_head_dict[layer_id] = { \
				"edp": result_head_edp[layer_id], \
				"energy": result_head_energy[layer_id], \
				"latency": result_head_latency[layer_id], \
				"NoC_DR": result_head_NoC_DR[layer_id], \
				"L2_to_DRAM_DR": result_head_L2_to_DRAM_DR[layer_id], \
				"DRAM_to_L2_DR": result_head_DRAM_to_L2_DR[layer_id], \
				"input_DRAM_flit_needed": result_head_input_DRAM_flit_needed[layer_id], \
				"weight_DRAM_flit_needed": result_head_weight_DRAM_flit_needed[layer_id], \
				"input_L2_flit_needed": result_head_input_L2_flit_needed[layer_id], \
				"weight_L2_flit_needed": result_head_weight_L2_flit_needed[layer_id], \
				"output_rd_flit_needed": result_head_output_rd_flit_needed[layer_id], \
				"output_wr_flit_needed": result_head_output_wr_flit_needed[layer_id], \
				"chiplet_spatial_parallel": result_head_chiplet_spatial_parallel[layer_id], \
				"iact_enough": result_head_iact_enough[layer_id]
			}
			
			fitness_tail_dict[layer_id] = { \
				"edp": result_tail_edp[layer_id], \
				"energy": result_tail_energy[layer_id], \
				"latency": result_tail_latency[layer_id], \
				"NoC_DR": result_tail_NoC_DR[layer_id], \
				"L2_to_DRAM_DR": result_tail_L2_to_DRAM_DR[layer_id], \
				"DRAM_to_L2_DR": result_tail_DRAM_to_L2_DR[layer_id], \
				"input_DRAM_flit_needed": result_tail_input_DRAM_flit_needed[layer_id], \
				"weight_DRAM_flit_needed": result_tail_weight_DRAM_flit_needed[layer_id], \
				"input_L2_flit_needed": result_tail_input_L2_flit_needed[layer_id], \
				"weight_L2_flit_needed": result_tail_weight_L2_flit_needed[layer_id], \
				"output_rd_flit_needed": result_tail_output_rd_flit_needed[layer_id], \
				"output_wr_flit_needed": result_tail_output_wr_flit_needed[layer_id], \
				"chiplet_spatial_parallel": result_tail_chiplet_spatial_parallel[layer_id], \
				"iact_enough": result_tail_iact_enough[layer_id]
			}

		return fitness_head_dict, fitness_tail_dict

	# Fitness per chiplet_parallel extract and file out
	# 获取每层在三种片间并行度方案下的适应度数值
	fitness_init_dict_p, layer_list_in_p = extractFitness_initial(result_indir_p, layer_position)
	fitness_init_dict_pk, layer_list_in_pk = extractFitness_initial(result_indir_pk, layer_position)
	fitness_init_dict_k, layer_list_in_k = extractFitness_initial(result_indir_k, layer_position)
	assert(layer_list_in_p == layer_list_in_pk)
	assert(layer_list_in_p == layer_list_in_k)
	if fuse_flag == 1:
		fitness_head_dict_p, fitness_tail_dict_p = extractFitness_fuse_head_tail(result_indir_p)
		fitness_head_dict_pk, fitness_tail_dict_pk = extractFitness_fuse_head_tail(result_indir_pk)
		fitness_head_dict_k, fitness_tail_dict_k = extractFitness_fuse_head_tail(result_indir_k)

	def getMinFrom3(a1, a2, a3, dim_list):
		if a1 == a2 == a3 == None:
			return None, 'p'
		def getMinFrom2(a1, a2):
			if a1 <= a2:
				return a1, 0
			else:
				return a2, 1
		num, index = getMinFrom2(a1, a2)
		num2, index2 = getMinFrom2(num, a3)
		if index2 == 1:
			return num2, dim_list[2]
		else:
			return num2, dim_list[index]
	
	def getFitnessMinDict(dict_p, dict_k, dict_pk, app_name, layer_list_in):
		assert(len(dict_p) == len(dict_k) == len(dict_pk))
		fitness_dict = {"layer_id":{}, "edp":[], "energy":[], "latency":[], "NoC_DR":[], "L2_to_DRAM_DR":[], "DRAM_to_L2_DR":[], "select_dim":[], \
						"input_DRAM_flit_needed":[], "weight_DRAM_flit_needed":[], "input_L2_flit_needed":[], "weight_L2_flit_needed":[], \
						"output_rd_flit_needed":[], "output_wr_flit_needed":[], "chiplet_spatial_parallel":[],\
						"iact_enough":[]}
		
		layer_id_list = []
		edp_dict = {}
		energy_dict = {}
		latency_dict = {}
		NoC_DR_dict = {}
		L2_to_DRAM_DR_dict = {}
		DRAM_to_L2_DR_dict = {}
		input_DRAM_flit_needed_dict = {}
		weight_DRAM_flit_needed_dict = {}
		input_L2_flit_needed_dict = {}
		weight_L2_flit_needed_dict = {}
		output_rd_flit_needed_dict = {}
		output_wr_flit_needed_dict = {}
		chiplet_spatial_parallel_dict = {}
		iact_enough_dict = {}
		select_dim_dict = {}

		for layer_id in dict_p:
			edp_p = dict_p[layer_id]["edp"]
			edp_k = dict_k[layer_id]["edp"]
			edp_pk = dict_pk[layer_id]["edp"]
			edp_min, dim = getMinFrom3(edp_p, edp_k, edp_pk, ["p",'k','pk'])

			if dim == "p":
				edp = dict_p[layer_id]["edp"]
				energy = dict_p[layer_id]["energy"]
				latency = dict_p[layer_id]["latency"]
				NoC_DR = dict_p[layer_id]["NoC_DR"]
				L2_to_DRAM_DR = dict_p[layer_id]["L2_to_DRAM_DR"]
				DRAM_to_L2_DR = dict_p[layer_id]["DRAM_to_L2_DR"]
				input_DRAM_flit_needed = dict_p[layer_id]["input_DRAM_flit_needed"]
				weight_DRAM_flit_needed = dict_p[layer_id]["weight_DRAM_flit_needed"]
				input_L2_flit_needed = dict_p[layer_id]["input_L2_flit_needed"]
				weight_L2_flit_needed = dict_p[layer_id]["weight_L2_flit_needed"]
				output_rd_flit_needed = dict_p[layer_id]["output_rd_flit_needed"]
				output_wr_flit_needed = dict_p[layer_id]["output_wr_flit_needed"]
				chiplet_spatial_parallel = dict_p[layer_id]["chiplet_spatial_parallel"]
				iact_enough = dict_p[layer_id]["iact_enough"]
			elif dim == "k":
				edp = dict_k[layer_id]["edp"]
				energy = dict_k[layer_id]["energy"]
				latency = dict_k[layer_id]["latency"]
				NoC_DR = dict_k[layer_id]["NoC_DR"]
				L2_to_DRAM_DR = dict_k[layer_id]["L2_to_DRAM_DR"]
				DRAM_to_L2_DR = dict_k[layer_id]["DRAM_to_L2_DR"]
				input_DRAM_flit_needed = dict_k[layer_id]["input_DRAM_flit_needed"]
				weight_DRAM_flit_needed = dict_k[layer_id]["weight_DRAM_flit_needed"]
				input_L2_flit_needed = dict_k[layer_id]["input_L2_flit_needed"]
				weight_L2_flit_needed = dict_k[layer_id]["weight_L2_flit_needed"]
				output_rd_flit_needed = dict_k[layer_id]["output_rd_flit_needed"]
				output_wr_flit_needed = dict_k[layer_id]["output_wr_flit_needed"]
				chiplet_spatial_parallel = dict_k[layer_id]["chiplet_spatial_parallel"]
				iact_enough = dict_k[layer_id]["iact_enough"]
			elif dim == "pk":
				edp = dict_pk[layer_id]["edp"]
				energy = dict_pk[layer_id]["energy"]
				latency = dict_pk[layer_id]["latency"]
				NoC_DR = dict_pk[layer_id]["NoC_DR"]
				L2_to_DRAM_DR = dict_pk[layer_id]["L2_to_DRAM_DR"]
				DRAM_to_L2_DR = dict_pk[layer_id]["DRAM_to_L2_DR"]
				input_DRAM_flit_needed = dict_pk[layer_id]["input_DRAM_flit_needed"]
				weight_DRAM_flit_needed = dict_pk[layer_id]["weight_DRAM_flit_needed"]
				input_L2_flit_needed = dict_pk[layer_id]["input_L2_flit_needed"]
				weight_L2_flit_needed = dict_pk[layer_id]["weight_L2_flit_needed"]
				output_rd_flit_needed = dict_pk[layer_id]["output_rd_flit_needed"]
				output_wr_flit_needed = dict_pk[layer_id]["output_wr_flit_needed"]
				chiplet_spatial_parallel = dict_pk[layer_id]["chiplet_spatial_parallel"]
				iact_enough = dict_pk[layer_id]["iact_enough"]
			else:
				print("dim error")
				exit()
			
			layer_id_list.append(layer_id)

			edp_dict[layer_id] = edp
			energy_dict[layer_id] = energy
			latency_dict[layer_id] = latency
			NoC_DR_dict[layer_id] = NoC_DR
			L2_to_DRAM_DR_dict[layer_id] = L2_to_DRAM_DR
			DRAM_to_L2_DR_dict[layer_id] = DRAM_to_L2_DR
			input_DRAM_flit_needed_dict[layer_id] = input_DRAM_flit_needed
			weight_DRAM_flit_needed_dict[layer_id] = weight_DRAM_flit_needed
			input_L2_flit_needed_dict[layer_id] = input_L2_flit_needed
			weight_L2_flit_needed_dict[layer_id] = weight_L2_flit_needed
			output_rd_flit_needed_dict[layer_id] = output_rd_flit_needed
			output_wr_flit_needed_dict[layer_id] = output_wr_flit_needed
			chiplet_spatial_parallel_dict[layer_id] = chiplet_spatial_parallel
			iact_enough_dict[layer_id] = iact_enough
			select_dim_dict[layer_id] = dim
		if debug_in_txt_extract == 1:
			print("DEBUG IN TXT EXTRACT")
			print('---edp_dict =', edp_dict)
			print("---layer_list_in: ", layer_list_in)
		for i in range(len(layer_list_in)):
			layer_id = "layer" + str(i+1)
			layer_index = layer_list_in[i]
			layer_id_select = "layer" + str(layer_index)
			fitness_dict["layer_id"][layer_id] = i
			fitness_dict["edp"].append(edp_dict[layer_id_select])
			fitness_dict["energy"].append(energy_dict[layer_id_select])
			fitness_dict["latency"].append(latency_dict[layer_id_select])
			fitness_dict["NoC_DR"].append(NoC_DR_dict[layer_id_select])
			fitness_dict["L2_to_DRAM_DR"].append(L2_to_DRAM_DR_dict[layer_id_select])
			fitness_dict["DRAM_to_L2_DR"].append(DRAM_to_L2_DR_dict[layer_id_select])
			fitness_dict["input_DRAM_flit_needed"].append(input_DRAM_flit_needed_dict[layer_id_select])
			fitness_dict["weight_DRAM_flit_needed"].append(weight_DRAM_flit_needed_dict[layer_id_select])
			fitness_dict["input_L2_flit_needed"].append(input_L2_flit_needed_dict[layer_id_select])
			fitness_dict["weight_L2_flit_needed"].append(weight_L2_flit_needed_dict[layer_id_select])
			fitness_dict["output_rd_flit_needed"].append(output_rd_flit_needed_dict[layer_id_select])
			fitness_dict["output_wr_flit_needed"].append(output_wr_flit_needed_dict[layer_id_select])
			fitness_dict["chiplet_spatial_parallel"].append(chiplet_spatial_parallel_dict[layer_id_select])
			fitness_dict["iact_enough"].append(iact_enough_dict[layer_id_select])
			fitness_dict["select_dim"].append(select_dim_dict[layer_id_select])

		return fitness_dict

	fitness_init_dict = getFitnessMinDict(fitness_init_dict_p, fitness_init_dict_k, fitness_init_dict_pk, app_name, layer_list_in_p)
	if fuse_flag == 1:
		fitness_headFuse_dict = getFitnessMinDict(fitness_head_dict_p, fitness_head_dict_k, fitness_head_dict_pk, app_name, layer_list_in_p)
		fitness_tailFuse_dict = getFitnessMinDict(fitness_tail_dict_p, fitness_tail_dict_k, fitness_tail_dict_pk, app_name, layer_list_in_p)
	else:
		fitness_headFuse_dict = {"edp":None, "energy":None, "latency":None}
		fitness_tailFuse_dict = {"edp":None, "energy":None, "latency":None}

	return fitness_init_dict, fitness_headFuse_dict, fitness_tailFuse_dict

def workload_extract(workload_dict, fitness_mid_dict, fitness_head_dict, fitness_tail_dict, tile_tag): # tail_tag = (0:alone; 1:start; 2:mid; 3:end)
	workload_fitness_dict = {}

	layer_name_index_dict = fitness_mid_dict["layer_id"]

	if tile_tag == 0:
		start_tile_tag = 1
		end_tile_tag = 1
	elif tile_tag == 1:
		start_tile_tag = 1
		end_tile_tag = 0
	elif tile_tag == 2:
		start_tile_tag = 0
		end_tile_tag = 0
	elif tile_tag == 3:
		start_tile_tag = 0
		end_tile_tag = 1

	for workload_name, workload_list in workload_dict.items():
		workload_fitness_dict[workload_name] = {}
		layer_name_list = workload_list["layer_name_list"]
		start_layer = layer_name_list[0]
		end_layer = layer_name_list[-1]
		for fitness_name in fitness_mid_dict:
			workload_fitness_dict[workload_name][fitness_name] = []
		for layer_name in layer_name_list:
			layer_index = layer_name_index_dict[layer_name]
			for fitness_name in workload_fitness_dict[workload_name]:
				if fitness_name == "layer_id":
					workload_fitness_dict[workload_name][fitness_name].append(layer_name)
				else:
					if layer_name == start_layer and start_tile_tag == 1:
						workload_fitness_dict[workload_name][fitness_name].append(fitness_head_dict[fitness_name][layer_index])
					elif layer_name == end_layer and end_tile_tag == 1:
						workload_fitness_dict[workload_name][fitness_name].append(fitness_tail_dict[fitness_name][layer_index])
					else:
						workload_fitness_dict[workload_name][fitness_name].append(fitness_mid_dict[fitness_name][layer_index])
	
	if debug_in_workload_extract == 1:
		print("DEBUG IN  F(workload_extract)-------------")
		print("---workload_dict: ", workload_dict)
		print("---fitness_mid_dict: ", fitness_mid_dict)
		print("---fitness_head_dict: ", fitness_head_dict)
		print("---fitness_tail_dict: ", fitness_tail_dict)
		print("---tile_tag: {}  start_tag:{}  end_tag:{}".format(tile_tag, start_tile_tag, end_tile_tag))
		print("---workload_fitness_dict: ", workload_fitness_dict)

	return workload_fitness_dict

# layer_fuse：多层融合
def layer_fuse(fitness_init_dict, fitness_head_dict, fitness_tail_dict):
	
	# --- get which layers fuse is useful
	fuse_tag = []
	fitness_fuse_dict = {"edp":[], "energy":[], "latency":[]}
	fitness_layer_dict = {"edp":fitness_init_dict["edp"], "energy":fitness_init_dict["energy"], "latency":fitness_init_dict["latency"]}
	edp_i_list = fitness_init_dict["edp"]
	energy_i_list = fitness_init_dict["energy"]
	latency_i_list = fitness_init_dict["latency"]
	energy_h_list = fitness_head_dict["energy"]
	latency_h_list = fitness_head_dict["latency"]
	energy_t_list = fitness_tail_dict["energy"]
	latency_t_list = fitness_tail_dict["latency"]

	for i in range(len(edp_i_list)-1):
		energy_i_h = energy_i_list[i]
		energy_i_t = energy_i_list[i+1]
		latency_i_h = latency_i_list[i]
		latency_i_t = latency_i_list[i+1]
		energy_h = energy_h_list[i]
		energy_t = energy_t_list[i+1]
		latency_h = latency_h_list[i]
		latency_t = latency_t_list[i+1]

		edp_i_fuse = (energy_i_h + energy_i_t) * (latency_i_h + latency_i_t) / PE_Frequency

		if energy_h != None and energy_t != None:
			energy_sum = energy_h + energy_t
			latency_sum = latency_h + latency_t
			edp_fuse = energy_sum * latency_sum / PE_Frequency
			if edp_fuse <= edp_i_fuse:
				fuse_tag.append(1)
			else:
				fuse_tag.append(0)
			fitness_fuse_dict["edp"].append(edp_fuse)
			fitness_fuse_dict["energy"].append(energy_sum)
			fitness_fuse_dict["latency"].append(latency_sum)
		else:
			fuse_tag.append(0)
			fitness_fuse_dict["edp"].append(None)
			fitness_fuse_dict["energy"].append(None)
			fitness_fuse_dict["latency"].append(None)
	
	# --- get need to decide layer fuse id
	one_num = 0
	fuse_code = [0 for _ in range(len(fuse_tag))]
	no_fuse_code = [0 for _ in range(len(fuse_tag))]
	layer_fuse_select_list = []
	for id in range(len(fuse_tag)):
		tag = fuse_tag[id]
		if tag == 1:
			one_num += 1
		else:
			if one_num > 1:
				layer_fuse_select_list.append([id-one_num, id-1])
			elif one_num == 1:
				fuse_code[id-1] = 1
			one_num = 0
	
	if debug_in_layer_fuse == 1:
		print("DEBUG IN F(Layer_fuse)----------------------")
		print("---fuse_tag:", fuse_tag)
		print("---fuse_code:", fuse_code)
	
	def getFuseFitness(fuse_fitness_dict, layer_fitness_dict, code, final_tag = 0):
		
		print("fuse_fitness_dict:", fuse_fitness_dict)
		print("layer_fitness_dict:", layer_fitness_dict)
		print("code:", code)

		layer_num = len(fuse_fitness_dict["latency"]) + 1
		layer_code = [0 for _ in range(layer_num)]
		layer_tag_list = ['i' for _ in range(layer_num)]
		if final_tag == 1 and debug_in_layer_fuse == 1:
			print("---FINAL RESULT IN DETAIL")
			print("---final_code: ", code)
			print("---code_lenth: ", len(code))
		for id in range(len(code)):
			fuse_num = code[id]
			if fuse_num == 1:
				assert(layer_code[id] == 0 and layer_code[id+1] == 0)
				layer_code[id] = 1
				layer_code[id+1] = 1
				layer_tag_list[id] = 'h'
				layer_tag_list[id+1] = 't'
			else:
				pass
		
		if final_tag == 1 and debug_in_layer_fuse == 1:
			print("---final_layer_code: ", layer_code)
		
		latency_all = 0
		energy_all = 0
		for id in range(len(code)):
			if code[id] == 1:
				latency_all += code[id] * fuse_fitness_dict["latency"][id]
				energy_all += code[id] * fuse_fitness_dict["energy"][id]

		for id in range(layer_num):
			latency_all += (1-layer_code[id]) * layer_fitness_dict["latency"][id]
			energy_all += (1-layer_code[id]) * layer_fitness_dict["energy"][id]
		
		edp_all = latency_all * energy_all / PE_Frequency 
			
		return edp_all, latency_all, energy_all, layer_tag_list

	def exploreFuse(fuse_fitness_dict, layer_fitness_dict, num):
		assert(num == len(fuse_fitness_dict["latency"]))

		def generateCode(code_lenth):
			max_1_num = math.ceil(code_lenth/2)
			code_list = []
			one_num = max_1_num
			if max_1_num * 2 == num + 1:
				code = [0 for _ in range(code_lenth)]
				for i in range(max_1_num):
					code[i*2] = 1
				code_list.append(code)
				one_num -= 1
			
			add_zero_num = code_lenth - one_num * 2 + 1
			code_init = [1]
			for i in range(one_num-1):
				code_init.append(0)
				code_init.append(1)
			
			if add_zero_num == 1:
				for i in range(max_1_num+1):
					insert_id = i*2
					code = code_init[0:insert_id] + [0] + code_init[insert_id:]
					code_list.append(code)
			else:
				assert(add_zero_num == 2)
				for i1 in range(max_1_num+1):
					for i2 in range(i1, max_1_num+1):
						insert_id1 = i1 * 2
						insert_id2 = i2 * 2
						code = code_init[0:insert_id1] + [0] + code_init[insert_id1:insert_id2] + [0] + code_init[insert_id2:]
						code_list.append(code)
						assert(len(code) == code_lenth)
			return code_list		
		
		code_list = generateCode(num)
		
		fitness_record_list = []
		fitness_best = None
		code_best = None
		for code in code_list:
			fitness, latency_all, energy_all, layer_tag_list = getFuseFitness(fuse_fitness_dict, layer_fitness_dict, code)
			fitness_record_list.append(fitness)
			if fitness_best == None or fitness <= fitness_best:
				fitness_best = fitness
				code_best = code
		
		return fitness_best, code_best

	for [start_id, end_id] in layer_fuse_select_list:
		num = end_id - start_id + 1
		fuse_fitness_dict = {}
		fuse_fitness_dict["latency"] = fitness_fuse_dict["latency"][start_id : end_id + 1]
		fuse_fitness_dict["energy"] = fitness_fuse_dict["energy"][start_id : end_id + 1]
		layer_fitness_dict = {}
		layer_fitness_dict["latency"] = fitness_layer_dict["latency"][start_id : end_id + 2]
		layer_fitness_dict["energy"] = fitness_layer_dict["energy"][start_id : end_id + 2]

		fitness, code = exploreFuse(fuse_fitness_dict, layer_fitness_dict, num)

		fuse_code[start_id: end_id+1] = code[:]

	edp_all, latency_all, energy_all, layer_tag_list = getFuseFitness(fitness_fuse_dict, fitness_layer_dict, fuse_code, final_tag = 1)
	edp_all_no_fuse, latency_all_no_fuse, energy_all_no_fuse, layer_tag_list_no_fuse = getFuseFitness(fitness_fuse_dict, fitness_layer_dict, no_fuse_code)
	
	if debug_in_layer_fuse == 1:
		print("---FINAL RESULT")
		print("---fuse_code: ", fuse_code)
		print("--result_fitness: ", edp_all)
		print("---result_fitness_no_fuse: ", edp_all_no_fuse)

	layer_fuse_use = 0
	layer_fuse_BW_dict = {"latency":[], "NoC_DR":[], "L2_to_DRAM_DR":[], "DRAM_to_L2_DR":[], "iact_enough":[]}
	layer_fuse_FN_dict = {"input_DRAM_flit_needed":[], "weight_DRAM_flit_needed":[], "input_L2_flit_needed":[], \
						"weight_L2_flit_needed":[], "output_rd_flit_needed":[], "output_wr_flit_needed":[],\
						"chiplet_spatial_parallel":[]}
	for i in range(len(layer_tag_list)):
		tag = layer_tag_list[i]
		if tag == 'i':
			latency = fitness_init_dict["latency"][i]
			NoC_DR = fitness_init_dict["NoC_DR"][i]
			L2_to_DRAM_DR = fitness_init_dict["L2_to_DRAM_DR"][i]
			DRAM_to_L2_DR = fitness_init_dict["DRAM_to_L2_DR"][i]
			input_DRAM_flit_needed = fitness_init_dict["input_DRAM_flit_needed"][i]
			weight_DRAM_flit_needed = fitness_init_dict["weight_DRAM_flit_needed"][i]
			input_L2_flit_needed = fitness_init_dict["input_L2_flit_needed"][i]
			weight_L2_flit_needed = fitness_init_dict["weight_L2_flit_needed"][i]
			output_rd_flit_needed = fitness_init_dict["output_rd_flit_needed"][i]
			output_wr_flit_needed = fitness_init_dict["output_wr_flit_needed"][i]
			chiplet_spatial_parallel = fitness_init_dict["chiplet_spatial_parallel"][i]
			iact_enough = int(fitness_init_dict["iact_enough"][i])
		elif tag == 'h':
			layer_fuse_use = 1
			latency = fitness_head_dict["latency"][i]
			NoC_DR = fitness_head_dict["NoC_DR"][i]
			L2_to_DRAM_DR = fitness_head_dict["L2_to_DRAM_DR"][i]
			DRAM_to_L2_DR = fitness_head_dict["DRAM_to_L2_DR"][i]
			input_DRAM_flit_needed = fitness_head_dict["input_DRAM_flit_needed"][i]
			weight_DRAM_flit_needed = fitness_head_dict["weight_DRAM_flit_needed"][i]
			input_L2_flit_needed = fitness_head_dict["input_L2_flit_needed"][i]
			weight_L2_flit_needed = fitness_head_dict["weight_L2_flit_needed"][i]
			output_rd_flit_needed = fitness_head_dict["output_rd_flit_needed"][i]
			output_wr_flit_needed = fitness_head_dict["output_wr_flit_needed"][i]
			chiplet_spatial_parallel = fitness_head_dict["chiplet_spatial_parallel"][i]
			iact_enough = int(fitness_init_dict["iact_enough"][i])
		elif tag == 't':
			latency = fitness_tail_dict["latency"][i]
			NoC_DR = fitness_tail_dict["NoC_DR"][i]
			L2_to_DRAM_DR = fitness_tail_dict["L2_to_DRAM_DR"][i]
			DRAM_to_L2_DR = fitness_tail_dict["DRAM_to_L2_DR"][i]
			input_DRAM_flit_needed = fitness_tail_dict["input_DRAM_flit_needed"][i]
			weight_DRAM_flit_needed = fitness_tail_dict["weight_DRAM_flit_needed"][i]
			input_L2_flit_needed = fitness_tail_dict["input_L2_flit_needed"][i]
			weight_L2_flit_needed = fitness_tail_dict["weight_L2_flit_needed"][i]
			output_rd_flit_needed = fitness_tail_dict["output_rd_flit_needed"][i]
			output_wr_flit_needed = fitness_tail_dict["output_wr_flit_needed"][i]
			chiplet_spatial_parallel = fitness_tail_dict["chiplet_spatial_parallel"][i]
			iact_enough = 1
		layer_fuse_BW_dict["latency"].append(latency)
		layer_fuse_BW_dict["NoC_DR"].append(NoC_DR)
		layer_fuse_BW_dict["L2_to_DRAM_DR"].append(L2_to_DRAM_DR)
		layer_fuse_BW_dict["DRAM_to_L2_DR"].append(DRAM_to_L2_DR)
		layer_fuse_FN_dict["input_DRAM_flit_needed"].append(input_DRAM_flit_needed)
		layer_fuse_FN_dict["weight_DRAM_flit_needed"].append(weight_DRAM_flit_needed)
		layer_fuse_FN_dict["input_L2_flit_needed"].append(input_L2_flit_needed)
		layer_fuse_FN_dict["weight_L2_flit_needed"].append(weight_L2_flit_needed)
		layer_fuse_FN_dict["output_rd_flit_needed"].append(output_rd_flit_needed)
		layer_fuse_FN_dict["output_wr_flit_needed"].append(output_wr_flit_needed)
		layer_fuse_FN_dict["chiplet_spatial_parallel"].append(chiplet_spatial_parallel)
		layer_fuse_BW_dict["iact_enough"].append(iact_enough)

	return edp_all, latency_all, energy_all, layer_fuse_BW_dict, layer_fuse_FN_dict, layer_fuse_use, layer_tag_list

def layer_no_fuse(fitness_init_dict):
	edp_all = 0
	latency_all = 0
	energy_all = 0
	edp_i_list = fitness_init_dict["edp"]
	energy_i_list = fitness_init_dict["energy"]
	latency_i_list = fitness_init_dict["latency"]

	for layer_id in range(len(edp_i_list)):
		energy_all += energy_i_list[layer_id]
		latency_all += latency_i_list[layer_id]
	
	edp_all = energy_all * latency_all / PE_Frequency

	print("new --------------------------")
	print("edp_all: ", edp_all)
	print("energy_all: ", energy_all)
	print("latency_all: ", latency_all)
	
	layer_tag_list = []
	layer_fuse_use = 0
	layer_fuse_BW_dict = {"latency":[], "NoC_DR":[], "L2_to_DRAM_DR":[], "DRAM_to_L2_DR":[], "iact_enough":[]}
	layer_fuse_FN_dict = {"input_DRAM_flit_needed":[], "weight_DRAM_flit_needed":[], "input_L2_flit_needed":[], \
						"weight_L2_flit_needed":[], "output_rd_flit_needed":[], "output_wr_flit_needed":[],\
						"chiplet_spatial_parallel":[]}
	for i in range(len(edp_i_list)):
		latency = fitness_init_dict["latency"][i]
		NoC_DR = fitness_init_dict["NoC_DR"][i]
		L2_to_DRAM_DR = fitness_init_dict["L2_to_DRAM_DR"][i]
		DRAM_to_L2_DR = fitness_init_dict["DRAM_to_L2_DR"][i]
		iact_enough = int(fitness_init_dict["iact_enough"][i])
		input_DRAM_flit_needed = fitness_init_dict["input_DRAM_flit_needed"][i]
		weight_DRAM_flit_needed = fitness_init_dict["weight_DRAM_flit_needed"][i]
		input_L2_flit_needed = fitness_init_dict["input_L2_flit_needed"][i]
		weight_L2_flit_needed = fitness_init_dict["weight_L2_flit_needed"][i]
		output_rd_flit_needed = fitness_init_dict["output_rd_flit_needed"][i]
		output_wr_flit_needed = fitness_init_dict["output_wr_flit_needed"][i]
		chiplet_spatial_parallel = fitness_init_dict["chiplet_spatial_parallel"][i]
		
		layer_fuse_BW_dict["latency"].append(latency)
		layer_fuse_BW_dict["NoC_DR"].append(NoC_DR)
		layer_fuse_BW_dict["L2_to_DRAM_DR"].append(L2_to_DRAM_DR)
		layer_fuse_BW_dict["DRAM_to_L2_DR"].append(DRAM_to_L2_DR)
		layer_fuse_FN_dict["input_DRAM_flit_needed"].append(input_DRAM_flit_needed)
		layer_fuse_FN_dict["weight_DRAM_flit_needed"].append(weight_DRAM_flit_needed)
		layer_fuse_FN_dict["input_L2_flit_needed"].append(input_L2_flit_needed)
		layer_fuse_FN_dict["weight_L2_flit_needed"].append(weight_L2_flit_needed)
		layer_fuse_FN_dict["output_rd_flit_needed"].append(output_rd_flit_needed)
		layer_fuse_FN_dict["output_wr_flit_needed"].append(output_wr_flit_needed)
		layer_fuse_FN_dict["chiplet_spatial_parallel"].append(chiplet_spatial_parallel)
		layer_fuse_BW_dict["iact_enough"].append(iact_enough)
		layer_tag_list.append('iiii')

	return edp_all, latency_all, energy_all, layer_fuse_BW_dict, layer_fuse_FN_dict, layer_fuse_use, layer_tag_list

def main(app_name, fuse_flag, chiplet_num_min_TH, chiplet_num_max_TH, workload_dict, architecture="simba", alg="GA", encode_type="index", objective="iact_num", tile_tag = 1, PE_parallel="All", debug_open=0, save_all_records=0):
	# Variable
	# --- result_indir: 单网络逐层性能结果文件地址
	# --- result_outdir: 多层融合后整个网络性能结果文件地址
	abs_path = os.path.dirname(os.path.abspath(__file__))
	SE_abs_path = os.path.join(abs_path, "../nnparser_SE_hetero_iodie")
	result_outdir = os.path.join(abs_path, "SE_result")
	os.makedirs(result_outdir, exist_ok=True)
	result_outdir = os.path.join(result_outdir, architecture) # + "_" + alg + "_" + encode_type)
	os.makedirs(result_outdir, exist_ok=True)

	BW_result_outdir = os.path.join(result_outdir, "BW_result")
	os.makedirs(BW_result_outdir, exist_ok=True)

	FN_result_outdir = os.path.join(result_outdir, "FN_result")
	os.makedirs(FN_result_outdir, exist_ok=True)

	extract_out_xls = os.path.join(abs_path, "extarct_out.xls")
	excel_data_dict = {}
	excel_title = []

	workload_edp_dict = {}
	workload_latency_dict = {}
	workload_energy_dict = {}
	workload_latency_BW_dict = {}
	workload_FN_dict = {}
	workload_layer_fuse_use_dict = {}
	workload_layer_tag_list_dict = {}

	edp_chiplet_num_dict = {}

	if tile_tag == 0:
		start_name = "_aloneTile"
	elif tile_tag == 1:
		start_name = "_headTile"
	elif tile_tag == 2:
		start_name = "_midTile"
	elif tile_tag == 3:
		start_name = "_tailTile"

	for workload_name in workload_dict:
		result_out_file = os.path.join(result_outdir, "{}{}.txt".format(workload_name, start_name))
		file = open(result_out_file, "w")
		file.close()

		workload_edp_dict[workload_name] = {}
		workload_latency_dict[workload_name] = {}
		workload_energy_dict[workload_name] = {}
		workload_latency_BW_dict[workload_name] = {}
		workload_FN_dict[workload_name] = {}
		workload_layer_fuse_use_dict[workload_name] = {}
		workload_layer_tag_list_dict[workload_name] = {}
	
		result_out_file = os.path.join(result_outdir, "{}_{}.txt".format(workload_name, "fuse_result"))
		file = open(result_out_file, "w")
		file.close()

	for i in range(chiplet_num_min_TH-1, chiplet_num_max_TH):
		chiplet_num = i + 1
		result_indir = os.path.join(SE_abs_path, "result")
		result_indir = os.path.join(result_indir, "intraLayer")
		result_indir = os.path.join(result_indir, architecture + "_" + app_name)
		result_indir = os.path.join(result_indir, "chiplet_num_"+str(chiplet_num))
		result_indir = os.path.join(result_indir, alg + "_" + encode_type)
		result_indir_p = os.path.join(result_indir, chiplet_parallel_list[0] + "_and_" + PE_parallel)
		result_indir_pk = os.path.join(result_indir, chiplet_parallel_list[1] + "_and_" + PE_parallel)
		result_indir_k = os.path.join(result_indir, chiplet_parallel_list[2] + "_and_" + PE_parallel)

		# 1. 获得模型的各层相关参数
		fitness_init_mid_dict, fitness_headFuse_mid_dict, fitness_tailFuse_mid_dict = txt_extract(result_indir_p, result_indir_pk, result_indir_k, app_name, fuse_flag, "mid")
		fitness_init_head_dict, fitness_headFuse_head_dict, fitness_tailFuse_head_dict = txt_extract(result_indir_p, result_indir_pk, result_indir_k, app_name, fuse_flag, "head")
		fitness_init_tail_dict, fitness_headFuse_tail_dict, fitness_tailFuse_tail_dict = txt_extract(result_indir_p, result_indir_pk, result_indir_k, app_name, fuse_flag, "tail")
		
		# record layer fitness per chiplet_num
		'''
		for layer_i, edp in enumerate(fitness_init_mid_dict["edp"]):
			layer_index = layer_i + 1
			if layer_index not in edp_chiplet_num_dict:
				edp_chiplet_num_dict[layer_index] = {}
			edp_i = edp * chiplet_num
			edp_chiplet_num_dict[layer_index][chiplet_num] = edp_i
		'''
		
		if debug_in_main == 1:
			if excel_title == []:
				excel_title = ["layer_id"]
				dict_name_list = ["init_flex", "head_flex", "tail_flex", "init_DRAM", "head_DRAM", "tail_DRAM"]
				for target in fitness_init_mid_dict:
					if target != "layer_id":
						for name in dict_name_list:
							target_name = target + "_" + name
							excel_title.append(target_name)
			print("--chiplet_num = ", chiplet_num)
			print("--Txt Extract")
			print("---fitness_init_mid_dict: ", fitness_init_mid_dict)
			print("---fitness_headFuse_mid_dict: ", fitness_headFuse_mid_dict)
			print("---fitness_tailFuse_mid_dict: ", fitness_tailFuse_mid_dict)
			print("---fitness_init_head_dict: ", fitness_init_head_dict)
			print("---fitness_headFuse_head_dict: ", fitness_headFuse_head_dict)
			print("---fitness_tailFuse_head_dict: ", fitness_tailFuse_head_dict)
			print("---fitness_init_tail_dict: ", fitness_init_tail_dict)
			print("---fitness_headFuse_tail_dict: ", fitness_headFuse_tail_dict)
			print("---fitness_tailFuse_tail_dict: ", fitness_tailFuse_tail_dict)
			print("")

			excel_data_dict[chiplet_num] = {}
			layer_name_list = []
			for target in fitness_init_mid_dict:
				if target == "layer_id":
					dict_i_f = fitness_init_mid_dict[target]
					for layer_name, id in dict_i_f.items():
						if layer_name not in excel_data_dict[chiplet_num]:
							excel_data_dict[chiplet_num][layer_name] = []
							layer_name_list.append(layer_name)
				else:
					list_i_m = fitness_init_mid_dict[target]
					list_i_h = fitness_init_head_dict[target]
					list_i_t = fitness_init_tail_dict[target]
					
					if target not in fitness_headFuse_mid_dict:
						list_h_m = None
						list_h_h = None
						list_h_t = None
					else:
						list_h_m = fitness_headFuse_mid_dict[target]
						list_h_h = fitness_headFuse_head_dict[target]
						list_h_t = fitness_headFuse_tail_dict[target]
					
					if target not in fitness_tailFuse_mid_dict:
						list_t_m = None
						list_t_h = None
						list_t_t = None
					else:
						list_t_m = fitness_tailFuse_mid_dict[target]
						list_t_h = fitness_tailFuse_head_dict[target]
						list_t_t = fitness_tailFuse_tail_dict[target]

					for index in range(len(list_i_m)):
						layer_name = layer_name_list[index]
						if list_h_m == None:
							h_m = None
							h_t = None
							h_h = None
						else:
							h_m = list_h_m[index]
							h_h = list_h_h[index]
							h_t = list_h_t[index]
						
						if list_t_m == None:
							t_m = None
							t_t = None
							t_h = None
						else:
							t_m = list_t_m[index]
							t_h = list_t_h[index]
							t_t = list_t_t[index]
						i_m = list_i_m[index]
						i_h = list_i_h[index]
						i_t = list_i_t[index]
						excel_data_dict[chiplet_num][layer_name].append(i_m)
						excel_data_dict[chiplet_num][layer_name].append(h_m)
						excel_data_dict[chiplet_num][layer_name].append(t_m)
						excel_data_dict[chiplet_num][layer_name].append(i_h)
						excel_data_dict[chiplet_num][layer_name].append(h_h)
						excel_data_dict[chiplet_num][layer_name].append(t_h)
						excel_data_dict[chiplet_num][layer_name].append(i_t)
						excel_data_dict[chiplet_num][layer_name].append(h_t)
						excel_data_dict[chiplet_num][layer_name].append(t_t)

		# 2. 获得模型中各子任务的相关参数
		workload_fitness_init_dict = workload_extract(workload_dict, fitness_init_mid_dict, fitness_init_head_dict, fitness_init_tail_dict, tile_tag)
		if fuse_flag == 1:
			workload_fitness_headFuse_dict = workload_extract(workload_dict, fitness_headFuse_mid_dict, fitness_headFuse_head_dict, fitness_headFuse_tail_dict, tile_tag)
			workload_fitness_tailFuse_dict = workload_extract(workload_dict, fitness_tailFuse_mid_dict, fitness_tailFuse_head_dict, fitness_tailFuse_tail_dict, tile_tag)
		else:
			workload_fitness_headFuse_dict = None
			workload_fitness_tailFuse_dict = None

		if debug_in_main == 1:
			print("--Worklaod Extract")
			print("---workload_fitness_init_dict: ", workload_fitness_init_dict)
			print("---workload_fitness_head_dict: ", workload_fitness_headFuse_dict)
			print("---workload_fitness_tail_dict: ", workload_fitness_tailFuse_dict)
			print("")

		# 3. 以子任务为粒度进行任务内的多层融合探索
		for workload_name in workload_fitness_init_dict:
			fitness_init_dict = workload_fitness_init_dict[workload_name]
			if fuse_flag == 1:
				fitness_headFuse_dict = workload_fitness_headFuse_dict[workload_name]
				fitness_tailFuse_dict = workload_fitness_tailFuse_dict[workload_name]

			if fuse_flag == 1:
				edp_all, latency_all, energy_all, layer_fuse_BW_dict, layer_fuse_FN_dict, layer_fuse_use, layer_tag_list = layer_fuse(fitness_init_dict, fitness_headFuse_dict, fitness_tailFuse_dict)
			else:
				edp_all, latency_all, energy_all, layer_fuse_BW_dict, layer_fuse_FN_dict, layer_fuse_use, layer_tag_list = layer_no_fuse(fitness_init_dict)
			
			if debug_in_main == 1:
				print("--Layer Fuse")
				print("--workload_name: ", workload_name)
				print("---workload-{}: edp_all={}, latency_all={}, energy_all={}".format(workload_name, edp_all, latency_all, energy_all))
				print("---                      : layer_fuse_use={}, layer_tag_list={}".format(layer_fuse_use, layer_tag_list))
				print("")
			
			workload_edp_dict[workload_name][chiplet_num] = edp_all
			workload_latency_dict[workload_name][chiplet_num] = latency_all
			workload_energy_dict[workload_name][chiplet_num] = energy_all
			workload_latency_BW_dict[workload_name][chiplet_num] = copy.deepcopy(layer_fuse_BW_dict)
			workload_FN_dict[workload_name][chiplet_num] = copy.deepcopy(layer_fuse_FN_dict)
			workload_layer_fuse_use_dict[workload_name][chiplet_num] = layer_fuse_use
			workload_layer_tag_list_dict[workload_name][chiplet_num] = layer_tag_list

			result_out_file = os.path.join(result_outdir, "{}{}.txt".format(workload_name, start_name))
			f = open(result_out_file, 'a')
			line = "chiplet_num\t{}\tedp_all\t{}\tlatency_all\t{}\tenergy_all\t{}\tiact_enouogh\t{}\tlayer_fuse_use\t{}\tlayer_fuse_tag\t{}".format(chiplet_num, edp_all, latency_all, energy_all, str(layer_fuse_BW_dict["iact_enough"]), layer_fuse_use, str(layer_tag_list))
			print(line, file=f)
			f.close()

			workload_FN_result_outdir = os.path.join(FN_result_outdir, workload_name + start_name)
			os.makedirs(workload_FN_result_outdir, exist_ok=True)
			f = open(workload_FN_result_outdir + "/chiplet_num_"+str(chiplet_num) + ".txt", 'w')
			line = "layer_id\tiact_DRAM_FN\tweight_DRAM_FN\tiact_L2_FN\tweight_L2_FN\toact_rd_FN\toact_wr_FN\tchiplet_spatial_parallel"
			print(line, file=f)

			for id in range(len(layer_fuse_FN_dict["input_DRAM_flit_needed"])):
				layer_id = id + 1
				input_DRAM_flit_needed = layer_fuse_FN_dict["input_DRAM_flit_needed"][id]
				weight_DRAM_flit_needed = layer_fuse_FN_dict["weight_DRAM_flit_needed"][id]
				input_L2_flit_needed = layer_fuse_FN_dict["input_L2_flit_needed"][id]
				weight_L2_flit_needed = layer_fuse_FN_dict["weight_L2_flit_needed"][id]
				output_rd_flit_needed = layer_fuse_FN_dict["output_rd_flit_needed"][id]
				output_wr_flit_needed = layer_fuse_FN_dict["output_wr_flit_needed"][id]
				chiplet_spatial_parallel = layer_fuse_FN_dict["chiplet_spatial_parallel"][id]
				line = "{}\t{}\t{}\t{}\t{}\t{}\t{}\t{}".format(layer_id, input_DRAM_flit_needed, weight_DRAM_flit_needed, input_L2_flit_needed, weight_L2_flit_needed, output_rd_flit_needed, output_wr_flit_needed, chiplet_spatial_parallel)
				print(line, file=f)
			f.close()

			workload_BW_result_outdir = os.path.join(BW_result_outdir, workload_name + start_name)
			os.makedirs(workload_BW_result_outdir, exist_ok=True)
			f = open(workload_BW_result_outdir + "/chiplet_num_"+str(chiplet_num) + ".txt", 'w')
			line = "layer_id\tlatency\tNoC_DR\tL2_to_DRAM_DR\tDRAM_to_L2_DR\t"
			print(line, file=f)

			if fuse_flag == 1:
				fuse_result_out_file = os.path.join(result_outdir, "{}_{}.txt".format(workload_name, "fuse_result"))
				fuse_f = open(fuse_result_out_file, "a")
				line = "chiplet" + str(chiplet_num) + ":"
				for id, tag in enumerate(layer_tag_list):
					layer_name = "layer" + str(id+1)
					if tag == 'i':
						line += "\t{}({}); ".format(layer_name, "None")
					elif tag == 'h':
						line += "\t{}({}); ".format(layer_name, "Head")
					elif tag == 't':
						line += "\t{}({}); ".format(layer_name, "Tail")
					else:
						print("ERROR TAG: ", tag)
						exit()
				print(line, file = fuse_f)
				fuse_f.close()

			for id in range(len(layer_fuse_BW_dict["latency"])):
				layer_id = id + 1
				latency = layer_fuse_BW_dict["latency"][id]
				NoC_DR = layer_fuse_BW_dict["NoC_DR"][id]
				L2_to_DRAM_DR = layer_fuse_BW_dict["L2_to_DRAM_DR"][id]
				DRAM_to_L2_DR = layer_fuse_BW_dict["DRAM_to_L2_DR"][id]
				line = "{}\t{}\t{}\t{}\t{}".format(layer_id, latency, NoC_DR, L2_to_DRAM_DR, DRAM_to_L2_DR)
				print(line, file=f)
			f.close()
	
	workbook = openpyxl.Workbook()
	for chiplet_num, excel_data in excel_data_dict.items():
		sheet = workbook.create_sheet('ChipletNum {}'.format(chiplet_num))
		for col, t in enumerate(excel_title):
			sheet.cell(1, col+1, t)
		
		row = 2
		for layer_name, items_list in excel_data.items():
			sheet.cell(row, 1, layer_name)
			for col, item in enumerate(items_list):
				sheet.cell(row, col+2, item)
			row += 1
	workbook.save(extract_out_xls)

	# record layer fitness per chiplet_num
	'''
	layer_edp_result_outdir = os.path.join(abs_path, "SE_result/layer_edp")
	os.makedirs(layer_edp_result_outdir, exist_ok=True)
	out_file = layer_edp_result_outdir + "/" + app_name + ".txt"
	out_f = open(out_file, 'w')
	for layer_index, edp_dict in edp_chiplet_num_dict.items():
		edp_list_order = sorted(edp_dict.values())
		edp_min = edp_list_order[0]
		edp_chiplet_num_line = "layer({})\t".format(layer_index)
		for edp in edp_list_order:
			edp_chiplet_num_line += "\t{}".format(edp/edp_min)
		print(edp_chiplet_num_line, file = out_f)

		edp_dict_order = sorted(edp_dict.items(), key = lambda x: x[1])
		for edp in edp_dict.values():
			edp_chiplet_num_line += "\t{:.2f}".format(edp/edp_min)
		print(edp_chiplet_num_line, file = out_f)

		layer_id_list_order = []
		for i in edp_dict_order:
			layer_id_list_order.append(i[0])
		edp_chiplet_num_line = "\t\t{}".format(layer_id_list_order)
		print(edp_chiplet_num_line, file = out_f)
	out_f.close()
	'''
	return workload_dict, workload_edp_dict, workload_latency_dict, workload_energy_dict, workload_latency_BW_dict, workload_FN_dict, workload_layer_fuse_use_dict, workload_layer_tag_list_dict

def fitness_plot(edp_dict, latency_dict, energy_dict, app_name, architecture="simba", alg="GA", encode_type="index"):
	abs_path = os.path.dirname(os.path.abspath(__file__))
	result_outdir = os.path.join(abs_path, "SE_result")
	os.makedirs(result_outdir, exist_ok=True)
	result_outdir = os.path.join(result_outdir, architecture + "_" + alg + "_" + encode_type)
	os.makedirs(result_outdir, exist_ok=True)
	result_outdir = os.path.join(result_outdir, "plot")
	os.makedirs(result_outdir, exist_ok=True)

	y_edp = []
	y_edp_2 = []
	y_energy = []
	y_latency = []
	x = []
	for chiplet_num in edp_dict:
		edp = edp_dict[chiplet_num]
		energy = energy_dict[chiplet_num]
		latency = latency_dict[chiplet_num]

		y_edp.append(edp)
		y_energy.append(energy)
		y_latency.append(latency)
		y_edp_2.append(edp_dict[16] * 16 / chiplet_num)

		x.append(chiplet_num)
	
	plt.figure("Fitness Result " + app_name)
	plt.title(app_name + " Fitness Result", fontsize=12)

	plt.subplot(3, 1, 1)
	plt.ylabel("EDP", fontsize=10)
	plt.bar(x, y_edp, width=0.5,color='rosybrown')
	plt.plot(x,y_edp,color='brown')
	plt.plot(x,y_edp_2,color='blue')
	plt.tick_params(labelsize=8)
	for i in range(len(x)):
		plt.scatter(x[i],y_edp[i],s=8,color='brown')
	
	plt.subplot(3, 1, 2)
	plt.ylabel("energy", fontsize=10)
	plt.bar(x, y_energy, width=0.5,color='rosybrown')
	plt.plot(x,y_energy,color='brown')
	plt.tick_params(labelsize=8)
	for i in range(len(x)):
		plt.scatter(x[i],y_energy[i],s=8,color='brown')

	plt.subplot(3, 1, 3)
	plt.ylabel("latency", fontsize=10)
	plt.bar(x, y_latency, width=0.5,color='rosybrown')
	plt.plot(x,y_latency,color='brown')
	plt.tick_params(labelsize=8)
	for i in range(len(x)):
		plt.scatter(x[i],y_latency[i],s=8,color='brown')
	
	plt.tight_layout(pad=1.1)
	plt_file = "{}/{}_Fitness.png".format(result_outdir, app_name)
	plt.savefig(plt_file, bbox_inches = 'tight')

if __name__ == '__main__':
	parser = argparse.ArgumentParser()
	parser.add_argument('--app_name_line', type=str, default="resnet50", help='app_name_line,using+as split signal')	# simba , nnbaton
	parser.add_argument('--chiplet_num_max_TH', type=int, default=16, help='max chiplet_num')
	parser.add_argument('--chiplet_num_min_TH', type=int, default=1, help='min chiplet_num')
	parser.add_argument('--fuse_flag', type=int, default=1, help='use layer fuse or not')
	parser.add_argument('--alg_list', type=str, default='GA', help='use layer fuse or not')
	parser.add_argument('--architecture', type=str, default='ours', help='hardware architecture')
	parser.add_argument('--encode_type_list', type=str, default='index', help='encode type')
	parser.add_argument('--workload_par_objective', type=str, default='iact_num', help='workload partition objective')
	parser.add_argument('--workload_num_TH', type=int, default=4, help='max workload num per nn')

	opt = parser.parse_args()
	app_name_line = opt.app_name_line
	chiplet_num_max_TH = opt.chiplet_num_max_TH
	chiplet_num_min_TH = opt.chiplet_num_min_TH
	fuse_flag = opt.fuse_flag
	architecture = opt.architecture
	alg_list = opt.alg_list
	encode_type_list = opt.encode_type_list
	workload_num_TH = opt.workload_num_TH

	app_name_list = app_name_line.replace('\n',"").split("+")
	alg_list = alg_list.replace('\n','').split("+")
	encode_type_list = encode_type_list.replace('\n','').split("+")

	abs_path = os.path.dirname(os.path.abspath(__file__))
	result_outdir = os.path.join(abs_path, "SE_result")
	os.makedirs(result_outdir, exist_ok=True)
	result_outdir = os.path.join(result_outdir, architecture) # + "_" + opt.alg + "_" + encode_type)
	if os.path.exists(result_outdir):
		shutil.rmtree(result_outdir)
	os.makedirs(result_outdir, exist_ok=True)
	workload_file = os.path.join(result_outdir, "workload_list.txt")
	workload_f = open(workload_file, "w")

	for i, app_name in enumerate(app_name_list):
		alg = alg_list[i]
		encode_type = encode_type_list[i]

		abs_path = os.path.dirname(os.path.abspath(__file__))
		SE_abs_path = os.path.join(abs_path, "../nnparser_SE_hetero_iodie")

		# 1. 模型的任务划分，以理论计算参数作为评估
		'''
		objective_list = ["iact_num", "oact_num", "weight_num", "compute_num", "iparam_num", "param_num", "iparam_vs_compute", "param_vs_compute", "iparam_mux_compute", "param_mux_compute"]
		layer_result_outdir = os.path.join(abs_path, "SE_result/layer_edp")
		os.makedirs(layer_result_outdir, exist_ok=True)
		out_file = layer_result_outdir + "/" + app_name + "_objective.txt"
		out_f = open(out_file, 'w')
		for objective in objective_list:
			print(objective, file = out_f)
			result_dict = getLayerParam(SE_abs_path, app_name, objective)
			result_order = sorted(result_dict.items(), key = lambda x : x[1])
			result_min = result_order[0][1]
			line_id = "\t"
			line_fitness = "\t"
			for [id, fitness] in result_order:
				line_id += "\t" + str(id)
				line_fitness += "\t" + "{:.2f}".format(fitness/result_min)
			print(line_id, file = out_f)
			print(line_fitness, file = out_f)
		'''

		workload_dict = getWorkloadPartition(app_name, SE_abs_path, opt.workload_par_objective, select_min_num = workload_num_TH)

		if debug_in_main == 1:
			print("DEBUG IN MAIN")
			print("--Workload Partition")
			print("---app_name:{};  objective:{};  workload_num_TH:{}".format(app_name, opt.workload_par_objective, workload_num_TH))
			print("")

		# 2. 任务的多层融合探索与相关参数整合
		workload_dict, workload_edp_dict, workload_latency_dict, workload_energy_dict, workload_latency_BW_dict, workload_FN_dict, \
		workload_layer_fuse_use_dict, workload_layer_tag_list_dict = main(app_name, fuse_flag, chiplet_num_min_TH, chiplet_num_max_TH, workload_dict, architecture = architecture, alg = alg, tile_tag=0, encode_type = encode_type)

		# 3. 输出workload list
		line = app_name + ": "
		workload_line = ""
		for workload_name, workload_list in workload_dict.items():
			workload_line += "{}:{};\t".format(workload_name, workload_list["layer_id_list"])
		line = app_name + ": {"+ workload_line + "}"
		print(line, file = workload_f)

		workload_dict, workload_edp_dict, workload_latency_dict, workload_energy_dict, workload_latency_BW_dict, workload_FN_dict, \
		workload_layer_fuse_use_dict, workload_layer_tag_list_dict = main(app_name, fuse_flag, chiplet_num_min_TH, chiplet_num_max_TH, workload_dict, architecture = architecture, alg = alg, tile_tag=1, encode_type = encode_type)
		
		workload_dict, workload_edp_dict, workload_latency_dict, workload_energy_dict, workload_latency_BW_dict, workload_FN_dict, \
		workload_layer_fuse_use_dict, workload_layer_tag_list_dict = main(app_name, fuse_flag, chiplet_num_min_TH, chiplet_num_max_TH, workload_dict, architecture = architecture, alg = alg, tile_tag=2, encode_type = encode_type)

		workload_dict, workload_edp_dict, workload_latency_dict, workload_energy_dict, workload_latency_BW_dict, workload_FN_dict, \
		workload_layer_fuse_use_dict, workload_layer_tag_list_dict = main(app_name, fuse_flag, chiplet_num_min_TH, chiplet_num_max_TH, workload_dict, architecture = architecture, alg = alg, tile_tag=3, encode_type = encode_type)

	workload_f.close()
