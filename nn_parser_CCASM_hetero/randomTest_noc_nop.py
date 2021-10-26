import math
import os
import sys
import random
import numpy as np
import copy
from enum import Enum
from single_engine_predict_noc_nop import *
from matplotlib import pyplot as plt
import openpyxl

iterTime = 100000
fitness_min_ran = 0
index = range(iterTime)
degrade_ratio_list = []
excel_datas = []

def randomTest():
	fitness_min_ran = 0
	fitness_list = []
	fitness_min_ran_list = []
	for i in range(iterTime):
		#---生成个代---
		for_list, act_wgt_dict, out_dict, parallel_dim_list, partition_list = GaGetChild()
		#---计算适应度---
		fitness, degrade_ratio, compuation_cycles,runtime_list,cp_list,utilization_ratio_list, chip_comm_num_list, core_comm_num_list = calFitness(for_list, act_wgt_dict, out_dict, parallel_dim_list, partition_list)
		#---比较适应度，并记录相关变量---
		if fitness_min_ran == 0 or fitness < fitness_min_ran:
			fitness_min_ran = fitness
			for_list_1 = copy.deepcopy(for_list)
			act_wgt_dict_1 = copy.deepcopy(act_wgt_dict)
			out_dict_1 = copy.deepcopy(out_dict)
			parallel_dim_list_1 = copy.deepcopy(parallel_dim_list)
			partition_list_1 = copy.deepcopy(partition_list)
			compuation_cycles_1 = compuation_cycles
			degrade_ratio_1 = degrade_ratio
		fitness_list.append(fitness)
		fitness_min_ran_list.append(fitness_min_ran)
		degrade_ratio_list.append (degrade_ratio)
		#TODO
		excel_datas.append([i, fitness, degrade_ratio, str(for_list[0]), \  
			parallel_dim_list[0][0],parallel_dim_list[0][1],parallel_dim_list[0][2], \
			parallel_dim_list[1][0],parallel_dim_list[1][1],parallel_dim_list[1][2], \
			parallel_dim_list[0][0]*parallel_dim_list[1][0], \
			parallel_dim_list[0][1]*parallel_dim_list[1][1], \
			parallel_dim_list[0][2]*parallel_dim_list[1][2], \
			parallel_dim_list[0][0]*parallel_dim_list[1][0]*parallel_dim_list[0][1]*parallel_dim_list[1][1], \
			str(partition_list), runtime_list[0], runtime_list[1], runtime_list[2],  \
			runtime_list[3], runtime_list[4], runtime_list[5], runtime_list[6],\
			cp_list[0], cp_list[1], cp_list[2],cp_list[3], cp_list[4], cp_list[5] ,\
		    utilization_ratio_list[0], utilization_ratio_list[1], utilization_ratio_list[2],utilization_ratio_list[3], utilization_ratio_list[4], utilization_ratio_list[5], \
			chip_comm_num_list[0], chip_comm_num_list[1], chip_comm_num_list[2], chip_comm_num_list[3], \
			core_comm_num_list[0], core_comm_num_list[1], core_comm_num_list[2], core_comm_num_list[3], \
			sum(chip_comm_num_list), sum(core_comm_num_list), sum(chip_comm_num_list) * 0.364+sum(core_comm_num_list)* 0.033])
		print("######---------Times = ", i)
		print("fitness_min_ran = ",fitness_min_ran)
		print("compuation_cycles_1 = ",compuation_cycles_1)
		print("degrade_ratio_1 = ",degrade_ratio_1)
		print("######---------over")
		print("")
		
		#---生成task file
	createTaskFile(for_list_1, act_wgt_dict_1, out_dict_1, parallel_dim_list_1, partition_list_1)
	workbook = openpyxl.Workbook()
	sheet = workbook.get_sheet_by_name('Sheet') 
	# 写入标题
	#TODO
	column_tite = ["index","fitness","degrade_ratio", "dataflow", \
		"PP2","PQ2","PK2","PP3","PQ3","PK3","PP","PQ","PKtotal","PPPQtotal", \
		"partition_list",\
		"runtimeP","runtimeQ", "runtimeC", "runtimeK", "runtimeChipNum", "runtimeCoreNum", "runtime_calNum",\
		"ol1_cp_id","al1_cp_id","wl1_cp_id","ol2_cp_id","al2_cp_id","wl2_cp_id", \
		"ol1_util","al1_util","wl1_util","ol2_util","al2_util","wl2_util", \
		"chip_num_wr_opt", "chip_num_rd_opt", "chip_num_rd_act" , "chip_num_rd_wgt",\
		"core_num_wr_opt", "core_num_rd_opt", "core_num_rd_act" , "core_num_rd_wgt"	,\
		"chip_comm", "core_comm", "power"]
	for col,column in enumerate(column_tite):
		sheet.cell(1, col+1, column)
	# 写入每一行
	for row, data in enumerate(excel_datas):
		for col, column_data in enumerate(data):
			sheet.cell(row+2, col+1, column_data)

	workbook.save('./randomTest_result_VGG-16 conv1.xls')
	return compuation_cycles_1,degrade_ratio_1, fitness_min_ran_list, partition_list_1, parallel_dim_list_1

if __name__ == '__main__':
	compuation_cycles_1,degrade_ratio_1, fitness_min_ran_list, partition_list_1, parallel_dim_list_1 = randomTest()
	print("fitness_min_ran = ",fitness_min_ran)
	print("compuation_cycles_1 = ",compuation_cycles_1)
	print("degrade_ratio_1 = ",degrade_ratio_1)
	print("partition_list_1 = ",partition_list_1)
	print("parallel_dim_list_1 = ",parallel_dim_list_1)


	print(fitness_min_ran_list[len(fitness_min_ran_list)-1])

	plt.figure(1)
	plt.scatter(index,degrade_ratio_list)
	plt.savefig("randomTest.png")


	degrade_ratio_list.sort()

	plt.figure(2)
	plt.scatter(index,degrade_ratio_list)
	plt.savefig("randomTest2.png")

	plt.figure(3)
	plt.scatter(index[:900],degrade_ratio_list[:900])
	plt.savefig("randomTest3.png")
